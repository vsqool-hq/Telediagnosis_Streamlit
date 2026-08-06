"""
Router modułu LEKARZY: diagnostyka gotowości, rozliczenie lekarzy i porównanie
(marża) — liczone dla istniejącego, ukończonego zadania jednostek (reużywamy
jego zweryfikowanych danych i snapshotu słownika). Moduł jednostek bez zmian.
"""

import os
import io
import json
import glob
import zipfile
import datetime

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from app import db
from app.storage import job_paths, version_dir

router = APIRouter(prefix="/api/doctors", tags=["doctors"])


def _now():
    return datetime.datetime.now().isoformat(timespec="seconds")


def _lekarze_dir(paths: dict) -> str:
    d = os.path.join(paths["base"], "lekarze")
    os.makedirs(d, exist_ok=True)
    return d


def _cache_path(paths: dict, name: str) -> str:
    return os.path.join(_lekarze_dir(paths), name)


def _has_sprawdzone(paths: dict) -> bool:
    """Czy zadanie ma pliki SPRAWDZONE — czyli czy da się je PRZELICZYĆ (lekarze/
    porównanie liczą wyłącznie ze sprawdzonych). Zadania zaimportowane z chmury
    bywają bez sprawdzonych — wtedy recompute dałby pusty wynik."""
    d = paths.get("sprawdzone")
    if not d or not os.path.isdir(d):
        return False
    return any(not os.path.basename(f).startswith("~$") for f in glob.glob(os.path.join(d, "*.xlsx")))


def _load_cache(paths: dict, name: str):
    from app.engine import ENGINE_VERSION
    p = _cache_path(paths, name)
    if os.path.isfile(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError):
            return None
        # Cache policzony STARSZYM silnikiem unieważniamy (przelicz z bieżącą logiką —
        # spójność z Pulpitem/Porównaniem) — ale TYLKO gdy jest z czego przeliczyć
        # (są pliki sprawdzone). Inaczej recompute dałby pusty wynik i zepsuł działające
        # (choć starsze) rozliczenie, np. zadania zaimportowane z chmury bez sprawdzonych.
        if (isinstance(data, dict) and not data.get("empty")
                and data.get("_engine_version") != ENGINE_VERSION and _has_sprawdzone(paths)):
            return None
        return data
    return None


def _save_cache(paths: dict, name: str, data: dict):
    from app.engine import ENGINE_VERSION
    try:
        if isinstance(data, dict):
            data["_engine_version"] = ENGINE_VERSION
        with open(_cache_path(paths, name), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except OSError:
        pass


def _active_doctor_cennik_csv() -> str | None:
    v = db.get_active_version("cennik_lekarzy")
    if not v:
        return None
    path = os.path.join(version_dir("cennik_lekarzy", v["id"]), v["filename"])
    return path if os.path.isfile(path) else None


def _active_doctor_cennik_csv_for_period(period: str | None) -> str | None:
    """CSV cennika lekarzy dopasowany do OKRESU rozliczenia. Konwertuje source.xlsx
    aktywnej wersji, wybierając blok-aneks OBOWIĄZUJĄCY dla `period` (a nie skrajnie
    prawy — ten bywa przygotowany na przyszłe okresy). Cache per (wersja, okres).
    Bez source.xlsx / bez okresu → zwykły aktywny CSV (skrajnie prawy blok)."""
    base = _active_doctor_cennik_csv()
    if not period:
        return base
    from app.engine.commitments import active_commitments_workbook
    wb_path, _ = active_commitments_workbook()
    if not wb_path or not os.path.isfile(wb_path):
        return base  # aktywny cennik wgrany bez .xlsx — nie ma z czego wybrać okresu
    out = os.path.join(os.path.dirname(wb_path), f"cennik_period_{period}.csv")
    if os.path.isfile(out):
        return out
    try:
        from app.engine.cennik_lekarzy_convert import convert_workbook, rows_to_csv
        res = convert_workbook(wb_path, period=period)
        if not res.get("rows"):
            return base
        with open(out, "w", encoding="utf-8") as f:
            f.write(rows_to_csv(res["rows"]))
        return out
    except Exception:  # noqa: BLE001
        return base


def _slownik_path(wzorcowe_dir: str) -> str | None:
    files = [f for f in glob.glob(os.path.join(wzorcowe_dir, "*.xls*"))
             if not os.path.basename(f).startswith("~$")]
    return files[0] if files else None


def _active_slownik_path() -> str | None:
    """Ścieżka do AKTYWNEJ wersji słownika (a nie kopii sprzed uruchomienia zadania).
    Kolumnę „Rodzaj procedury lekarz" wypełnia się na bieżąco, więc kategorie
    bierzemy z aktualnego słownika, nie ze starego snapshotu zadania."""
    wz = db.get_active_version("wzorcowe")
    if not wz:
        return None
    p = os.path.join(version_dir("wzorcowe", wz["id"]), wz["filename"])
    return p if os.path.isfile(p) else None


@router.get("/coverage")
async def coverage():
    """Czy moduł jest gotowy: cennik lekarzy + wypełniona kolumna w słowniku."""
    from openpyxl import load_workbook

    out = {
        "doctor_cennik": None,
        "slownik_lekarz_filled": 0,
        "slownik_total": 0,
        "ready": False,
    }

    csv_path = _active_doctor_cennik_csv()
    if csv_path:
        n_rows = n_docs = 0
        docs = set()
        with open(csv_path, encoding="utf-8-sig") as f:
            next(f, None)
            for line in f:
                parts = line.rstrip("\n").split(";")
                if len(parts) >= 3 and parts[0]:
                    n_rows += 1
                    docs.add(parts[0])
        out["doctor_cennik"] = {"rows": n_rows, "doctors": len(docs)}

    wz = db.get_active_version("wzorcowe")
    if wz:
        spath = os.path.join(version_dir("wzorcowe", wz["id"]), wz["filename"])
        if os.path.isfile(spath):
            try:
                wb = load_workbook(spath, read_only=True, data_only=True)
                ws = wb["Szczegółowe"] if "Szczegółowe" in wb.sheetnames else wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                header = list(next(rows))
                idx = header.index("Rodzaj procedury lekarz") if "Rodzaj procedury lekarz" in header else None
                total = filled = 0
                for r in rows:
                    total += 1
                    if idx is not None and idx < len(r):
                        v = r[idx]
                        if v not in (None, "", "None"):
                            filled += 1
                out["slownik_total"] = total
                out["slownik_lekarz_filled"] = filled
            except Exception:  # noqa: BLE001
                pass

    out["ready"] = bool(out["doctor_cennik"] and out["slownik_lekarz_filled"] > 0)
    return out


def _resolve_job(job_id: str):
    job = db.get_job(job_id)
    if not job:
        raise HTTPException(404, "Nie znaleziono zadania.")
    # Lekarze liczą się z „sprawdzone": pełne rozliczenie (full) albo tryb „doctors"
    # (sam Etap 1 na nowo wgranym pliku — „tylko lekarze").
    if job["mode"] not in ("full", "doctors"):
        raise HTTPException(400, "Moduł lekarzy działa na pełnym rozliczeniu jednostek lub trybie „tylko lekarze”.")
    from app.storage import heal_job_dirs
    heal_job_dirs(job_id)  # napraw zadania zaimportowane starą paczką (złe nazwy katalogów)
    paths = job_paths(job_id)
    # Słownik bierzemy z AKTYWNEJ wersji (świeże kategorie „Rodzaj procedury lekarz"),
    # a tylko awaryjnie z kopii zapisanej przy zadaniu.
    slownik = _active_slownik_path() or _slownik_path(paths["wzorcowe"])
    # Cennik lekarzy dobrany do OKRESU rozliczanego miesiąca (blok-aneks obowiązujący
    # w tym miesiącu, nie skrajnie prawy). Okres z nazwy pliku wejściowego zadania.
    from app.engine.periods import period_from_filename
    period = period_from_filename(job.get("input_name"))
    cennik_lek = _active_doctor_cennik_csv_for_period(period)
    if not slownik:
        raise HTTPException(400, "Brak słownika w danych zadania.")
    if not cennik_lek:
        raise HTTPException(400, "Brak aktywnego cennika lekarzy. Wgraj go w zakładce 'Cennik lekarzy'.")
    return job, paths, slownik, cennik_lek


def _excluded_keys() -> list:
    """Klucze lekarzy wyłączonych z rozliczenia (z ustawień)."""
    try:
        return sorted(set(db.get_settings().get("doctors_excluded", []) or []))
    except Exception:  # noqa: BLE001
        return []


def _availability_totals(job: dict) -> dict:
    """Gotowość+triaż per lekarz dla miesiąca zadania — do doliczenia w Porównaniu.
    Zwraca {} gdy TeamUp nieskonfigurowany / brak miesiąca (nie blokuje liczenia)."""
    from app.engine.periods import period_from_filename
    period = period_from_filename((job or {}).get("input_name"))
    if not period:
        return {}
    try:
        from app.engine.teamup import compute_availability
        av = compute_availability(period, excluded_keys=_excluded_keys())
        return {lk: d["total"] for lk, d in av["doctors"].items()}
    except RuntimeError:
        return {}


def _compare_cache_fresh(c: dict | None) -> bool:
    """Czy zapisane porównanie jest aktualne względem WYŁĄCZONYCH JEDNOSTEK.
    Zmiana wyłączeń → zapis traktujemy jako „do przeliczenia" (nie kasujemy go).
    Dodatkowo: brak pola „koszt_konsultacje" = zapis SPRZED doliczania konsultacji
    do kosztu lekarzy → wymuszamy przeliczenie (inaczej Porównanie ≠ Rozliczenie)."""
    if not c or c.get("empty"):
        return False
    if "koszt_konsultacje" not in (c.get("totals") or {}):
        return False
    from app.engine.billing import get_excluded_units
    return (c.get("_units_excluded", []) == sorted(get_excluded_units())
            and c.get("_doctors_excluded", []) == _excluded_keys())


@router.get("/billing/{job_id}")
async def doctor_billing(job_id: str, recompute: bool = False, peek: bool = False):
    """
    Zwraca ZAPISANY wynik rozliczenia lekarzy (cache lekarze/billing.json). NIE liczy
    w wątku serwera — ciężkie liczenie biegnie w osobnym procesie (patrz /billing/{id}/run
    + /billing/{id}/status), żeby nie blokować pętli async i nie zrywać długich żądań.

      • peek=true / domyślnie → zwróć cache albo {empty, reason:'not_computed'},
      • recompute=true        → wystartuj liczenie w tle i zwróć {empty, reason:'computing'}.
    Cache jest unieważniany, gdy zmieni się lista wyłączonych lekarzy.
    """
    paths = job_paths(job_id)
    excluded = _excluded_keys()
    cached = None if recompute else _load_cache(paths, "billing.json")
    if cached is not None and cached.get("_excluded_keys", []) == excluded:
        return cached
    if peek:
        return {"empty": True, "reason": "not_computed", "computed_at": None}

    # Zgodność wstecz: stare wywołanie z recompute=true uruchamia teraz zadanie w tle.
    if recompute:
        _resolve_job(job_id)  # walidacja (cennik/słownik) — szybki błąd, zanim wystartujemy
        from app.services import doctors_job
        doctors_job.start(job_id, recompute=True)
    return {"empty": True, "reason": "computing", "computed_at": None}


@router.post("/billing/{job_id}/run")
async def doctor_billing_run(job_id: str, recompute: bool = False):
    """Startuje rozliczenie lekarzy w tle (osobny proces). Zwraca {status}."""
    _resolve_job(job_id)  # walidacja wejść (cennik lekarzy / słownik) — szybki błąd
    from app.services import doctors_job
    return doctors_job.start(job_id, recompute=recompute)


@router.get("/billing/{job_id}/status")
async def doctor_billing_status(job_id: str):
    """Status liczenia w tle do odpytywania przez front: idle|running|done|error."""
    db.get_job(job_id) or _raise_no_job()
    from app.services import doctors_job
    return doctors_job.status(job_id)


def _raise_no_job():
    raise HTTPException(404, "Nie znaleziono zadania.")



@router.get("/compare/months")
async def doctor_compare_months():
    """Miesiące rozliczeniowe do przełącznika na Porównaniu. Każdy miesiąc jest
    spięty z jego OSTATNIM przeliczeniem (latest_jobs_by_month — jak Pulpit/trend).
    Dla każdego zwracamy, czy porównanie jest już policzone (computed/computed_at).
    Musi być zadeklarowane PRZED /compare/{job_id}."""
    from app.routers.stats import latest_jobs_by_month
    best = latest_jobs_by_month()
    months = []
    for period in sorted(best.keys(), reverse=True):
        b = best[period]
        c = _load_cache(job_paths(b["job_id"]), "compare.json")
        computed = _compare_cache_fresh(c)   # nieaktualne wyłączenia jednostek → „policz"
        months.append({
            "period": period,
            "job_id": b["job_id"],
            "revenue": b["revenue"],
            "computed": computed,
            "computed_at": (c or {}).get("computed_at") if computed else None,
        })
    return {"months": months}


@router.get("/revenue-history")
async def doctors_revenue_history():
    """Historia wypłaty per lekarz (badania + gotowość/triaż) za miesiące, dla
    których rozliczenie lekarzy było już policzone (billing.json w cache zadania) —
    do dymków „najedź i zobacz historię" na Pulpicie/Porównaniu. W przeciwieństwie
    do jednostek NIE liczymy tego w locie dla brakujących miesięcy: przeliczenie
    dzisiejszym, aktualnym cennikiem lekarzy dałoby historycznie nieprawdziwe kwoty,
    jeśli cennik się od tamtej pory zmienił."""
    from app.routers.stats import latest_jobs_by_month
    from app.engine.doctors import doctor_key

    best = latest_jobs_by_month()
    history: dict[str, dict[str, float]] = {}
    names: dict[str, str] = {}
    for period, info in sorted(best.items()):
        cache = _load_cache(job_paths(info["job_id"]), "billing.json")
        if not cache or cache.get("empty"):
            continue
        for row in cache.get("by_doctor") or []:
            key = doctor_key(row.get("lekarz"))
            if not key:
                continue
            history.setdefault(key, {})[period] = round(float(row.get("wartosc") or 0), 2)
            names[key] = row.get("lekarz")  # nadpisywane w kolejności miesięcy -> zostaje najnowsza nazwa
    return {"doctors": history, "names": names}


@router.get("/compare/latest")
async def doctor_compare_latest():
    """Najnowsze ZAPISANE porównanie (wg computed_at) spośród wszystkich zadań —
    żeby na Porównaniu od razu pokazać ostatnio przeliczone wyniki. Musi być
    zadeklarowane PRZED /compare/{job_id} (inaczej „latest" złapie się jako job_id)."""
    # Ostatnio przeliczone zadanie najnowszego miesiąca — spójnie z Pulpitem.
    from app.routers.stats import latest_jobs_by_month
    best = latest_jobs_by_month()
    if not best:
        return {"empty": True, "reason": "not_computed"}
    latest = max(best.values(), key=lambda b: b["period"])
    job_id = latest["job_id"]
    out = _load_cache(job_paths(job_id), "compare.json")
    if not _compare_cache_fresh(out):
        return {"empty": True, "reason": "not_computed", "job_id": job_id}
    # Grupy jednostek (widok) — tak jak w /compare/{job_id}.
    if out.get("by_unit"):
        from app.engine.config import load_config, build_unit_group_map
        from app.engine.compare import regroup_by_unit
        gmap = build_unit_group_map(load_config().get("unit_groups", []))
        if gmap:
            out = {**out, "by_unit": regroup_by_unit(out["by_unit"], gmap)}
    return {**out, "job_id": job_id}


@router.get("/compare/{job_id}")
async def doctor_compare(job_id: str, recompute: bool = False, peek: bool = False):
    """Porównanie (marża) dla zadania — z takim samym zapisem/odczytem jak rozliczenie."""
    paths = job_paths(job_id)
    cached = None if recompute else _load_cache(paths, "compare.json")
    # Peek nigdy nie liczy — zwraca AKTUALNY zapis lub „brak" (stary zapis z innym
    # zestawem wyłączonych jednostek traktujemy jak niepoliczony).
    if peek:
        out = cached if _compare_cache_fresh(cached) else {"empty": True, "reason": "not_computed", "computed_at": None}
    # Przelicz, gdy brak cache, stary cache bez „rows_units" LUB nieaktualne wyłączenia.
    elif cached is not None and cached.get("empty"):
        out = cached
    elif cached is not None and "rows_units" in cached and _compare_cache_fresh(cached):
        out = cached
    else:
        _job, paths, slownik, cennik_lek = _resolve_job(job_id)
        from app.engine.compare import build_comparison
        out = build_comparison(paths["sprawdzone"], slownik, paths["cennik"], cennik_lek,
                               availability_by_doctor=_availability_totals(_job),
                               excluded_doctor_keys=_excluded_keys())
        if not out.get("empty"):
            out["computed_at"] = _now()
            _save_cache(paths, "compare.json", out)  # zapis ZAWSZE bez grupowania

    # Grupy jednostek (widok) — nakładane przy odczycie na świeżej kopii, bez zapisu,
    # żeby zmiana grupowania w Ustawieniach działała od razu (bez przeliczania).
    if isinstance(out, dict) and not out.get("empty") and out.get("by_unit"):
        from app.engine.config import load_config, build_unit_group_map
        from app.engine.compare import regroup_by_unit
        gmap = build_unit_group_map(load_config().get("unit_groups", []))
        if gmap:
            out = {**out, "by_unit": regroup_by_unit(out["by_unit"], gmap)}
    return out


def _xlsx_response(sheets: dict, filename: str) -> StreamingResponse:
    import pandas as pd
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet, rows in sheets.items():
            pd.DataFrame(rows or []).to_excel(writer, sheet_name=sheet[:31], index=False)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/billing/{job_id}/download")
async def doctor_billing_download(job_id: str):
    paths = job_paths(job_id)
    excluded = _excluded_keys()
    res = _load_cache(paths, "billing.json")
    need_full = res is None or res.get("_excluded_keys", []) != excluded
    if need_full or "category_counts" not in res:
        _job, paths, slownik, cennik_lek = _resolve_job(job_id)
        from app.engine.doctors import build_doctor_billing
        fresh = build_doctor_billing(paths["sprawdzone"], slownik, cennik_lek, excluded_keys=excluded)
        if fresh.get("empty"):
            raise HTTPException(400, fresh.get("reason", "Brak danych do rozliczenia lekarzy."))
        if need_full:
            res = fresh
            res["computed_at"] = _now()
            res["_excluded_keys"] = excluded
        else:
            # stary cache bez podsumowania — dołóż je, zachowując files_count itp.
            res["category_counts"] = fresh.get("category_counts", [])
        _save_cache(paths, "billing.json", res)

    # Arkusz „Podsumowanie": A=Lekarz, B=Kategoria (z cennika lekarzy),
    # C=Ilość wykonanych badań w tej kategorii.
    summary = [
        {"Lekarz": r.get("lekarz"), "Kategoria": r.get("kategoria"), "Ilość badań": r.get("ilosc")}
        for r in res.get("category_counts", [])
    ]
    return _xlsx_response(
        {"Podsumowanie": summary},
        "Podsumowanie_badan_lekarze.xlsx",
    )


@router.get("/billing/{job_id}/availability")
async def doctor_availability_download(job_id: str):
    """Roboczy plik: gotowość + triaż (TeamUp) SZCZEGÓŁOWO per lekarz, w jednym
    Excelu. Bierze wynik z cache rozliczenia lekarzy; gdy brak — liczy z miesiąca
    zadania. 400 gdy TeamUp nieskonfigurowany / brak pliku zobowiązań."""
    job = db.get_job(job_id)
    if not job:
        raise HTTPException(404, "Nie znaleziono zadania.")
    res = _load_cache(job_paths(job_id), "billing.json")
    av = (res or {}).get("availability")
    if not av:
        from app.engine.periods import period_from_filename
        period = period_from_filename(job.get("input_name"))
        if not period:
            raise HTTPException(400, "Nie rozpoznano miesiąca z nazwy pliku — brak danych gotowości.")
        try:
            from app.engine.teamup import compute_availability
            av = compute_availability(period)
        except RuntimeError as e:
            raise HTTPException(400, str(e))

    rows = []
    for doc in sorted(av.get("doctors", {}).values(), key=lambda d: d["name"].lower()):
        for it in doc.get("items", []):
            rows.append({
                "Lekarz": doc["name"], "Pozycja": it["label"],
                "Godziny": round(it["hours"], 2), "Stawka": it["rate"], "Wartość": it["amount"],
            })
        rows.append({"Lekarz": doc["name"], "Pozycja": "RAZEM", "Godziny": None,
                     "Stawka": None, "Wartość": doc.get("total", 0)})
        rows.append({})  # pusty wiersz-separator między lekarzami
    rows.append({"Lekarz": "SUMA — gotowość", "Godziny": av.get("hours_gotowosc", 0), "Wartość": av.get("sum_gotowosc", 0)})
    rows.append({"Lekarz": "SUMA — triaż", "Godziny": av.get("hours_triaz", 0), "Wartość": av.get("sum_triaz", 0)})
    rows.append({"Lekarz": "SUMA — razem", "Wartość": av.get("sum_total", 0)})

    # Diagnostyka: godziny NIEROZLICZONE (stawka 0/brak) oraz z nierozpoznanych tytułów.
    rows.append({})
    rows.append({"Lekarz": "GODZINY BEZ STAWKI (rozliczone na 0)", "Godziny": av.get("unbilled_hours", 0)})
    for u in av.get("unbilled", []):
        rows.append({"Lekarz": u["name"], "Pozycja": u["label"], "Godziny": u["hours"], "Stawka": 0, "Wartość": 0})
    if av.get("unmatched"):
        rows.append({})
        rows.append({"Lekarz": "GODZINY NIEROZPOZNANE (brak lekarza)", "Godziny": av.get("unmatched_hours", 0)})
        rows.append({"Lekarz": "Tytuły:", "Pozycja": ", ".join(av["unmatched"])})

    return _xlsx_response({f"Gotowość {av.get('period', '')}".strip(): rows},
                          f"Gotowosc_triaz_{av.get('period', 'okres')}.xlsx")


@router.get("/billing/{job_id}/files")
async def doctor_billing_files(job_id: str):
    """ZIP z osobnym plikiem Excel dla każdego lekarza (układ jak jednostki).
    Pliki powstają przy „Policz/Przelicz ponownie" rozliczenia lekarzy."""
    paths = job_paths(job_id)
    pliki_dir = os.path.join(_lekarze_dir(paths), "pliki")
    files = sorted(
        f for f in glob.glob(os.path.join(pliki_dir, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    )
    if not files:
        raise HTTPException(404, "Brak plików — najpierw policz rozliczenie lekarzy.")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in files:
            zf.write(p, os.path.basename(p))
    buf.seek(0)
    # octet-stream (nie „application/zip") — zniechęca Safari do automatycznego
    # rozpakowania po pobraniu, dzięki czemu można wysłać dalej TEN paczkowany ZIP
    # (ma poprawne polskie znaki w nazwach), bez ponownego pakowania w Finderze.
    return StreamingResponse(
        buf, media_type="application/octet-stream",
        headers={"Content-Disposition": 'attachment; filename="Rozliczenia_lekarzy.zip"'},
    )


@router.get("/compare/{job_id}/download")
async def doctor_compare_download(job_id: str):
    paths = job_paths(job_id)
    res = _load_cache(paths, "compare.json")
    if not _compare_cache_fresh(res):   # brak zapisu LUB nieaktualne wyłączenia jednostek
        _job, paths, slownik, cennik_lek = _resolve_job(job_id)
        from app.engine.compare import build_comparison
        res = build_comparison(paths["sprawdzone"], slownik, paths["cennik"], cennik_lek,
                               availability_by_doctor=_availability_totals(_job),
                               excluded_doctor_keys=_excluded_keys())
        if res.get("empty"):
            raise HTTPException(400, res.get("reason", "Brak danych do porównania."))
        res["computed_at"] = _now()
        _save_cache(paths, "compare.json", res)
    return _xlsx_response({
        "Marża per kategoria": res.get("rows", []),
        "Marża per lekarz": res.get("by_doctor", []),
        "Marża per jednostka": res.get("by_unit", []),
    }, "Porownanie_lekarze_jednostki.xlsx")


def _latest_full_job_id() -> str | None:
    for j in db.list_jobs(limit=100):
        if j["status"] == "done" and j["mode"] == "full":
            return j["id"]
    return None


def participants(job_id: str) -> dict:
    """
    Lekarze (kolumna „Opisujący") i jednostki (kolumna „Klient") występujący
    w danych zadania — z CACHE (participants.json w katalogu zadania). Zestaw
    jest STAŁY dla wgranego pliku, więc czytamy pliki Excel tylko RAZ; kolejne
    wejścia na Ustawienia dostają listę natychmiast.
    """
    paths = job_paths(job_id)
    cache = os.path.join(paths["base"], "participants.json")
    if os.path.isfile(cache):
        try:
            with open(cache, "r", encoding="utf-8") as f:
                data = json.load(f)
            # „consultants" doszło później — brak klucza = stary cache, przelicz raz.
            if "doctors" in data and "units" in data and "consultants" in data:
                return data
        except (OSError, ValueError):
            pass
    from app.engine.doctors import (
        read_verified_studies, doctor_key, _norm, OPISUJACY_COL, KONSULTUJACY_COL,
    )
    from app.engine.billing import _norm_unit
    doctors, units, consultants, seen, seen_c = [], [], [], set(), set()
    df = read_verified_studies(paths["sprawdzone"])
    if df is not None and not df.empty:
        if OPISUJACY_COL in df.columns:
            for val in df[OPISUJACY_COL].dropna().unique():
                disp = _norm(val)
                k = doctor_key(disp)
                if disp and k not in seen:
                    seen.add(k)
                    doctors.append({"name": disp, "key": k})
        # Konsultujący — osobna lista (lekarz może konsultować, nie opisując wcale).
        if KONSULTUJACY_COL in df.columns:
            for val in df[KONSULTUJACY_COL].dropna().unique():
                disp = _norm(val)
                k = doctor_key(disp)
                if disp and disp.lower() not in ("nan", "none") and k not in seen_c:
                    seen_c.add(k)
                    consultants.append({"name": disp, "key": k})
        if "Klient" in df.columns:
            for val in df["Klient"].dropna().unique():
                name = str(val).strip()
                if name:
                    units.append({"name": name, "key": _norm_unit(name)})
    doctors.sort(key=lambda d: d["name"].lower())
    consultants.sort(key=lambda d: d["name"].lower())
    units.sort(key=lambda u: u["name"].lower())
    data = {"doctors": doctors, "units": units, "consultants": consultants}
    try:
        with open(cache, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except OSError:
        pass
    return data


@router.get("/list")
async def doctors_list(job_id: str | None = None):
    """
    Lista lekarzy z danych zadania (domyślnie najnowszego pełnego rozliczenia),
    z flagą „wyłączony". Zestaw per zadanie z cache (participants) — bez czytania
    plików Excel przy każdym wejściu na Ustawienia.
    """
    jid = job_id or _latest_full_job_id()
    if not jid:
        return {"job_id": None, "doctors": []}
    excluded = set(_excluded_keys())
    doctors = [{**d, "excluded": d["key"] in excluded} for d in participants(jid)["doctors"]]
    return {"job_id": jid, "doctors": doctors}


@router.get("/names")
async def doctors_names(job_id: str | None = None):
    """
    Nazwiska lekarzy do WYBORU w ustawieniach konsultacji (dropdown zamiast
    wpisywania ręcznego — mniej pomyłek). Unia trzech źródeł, odduplikowana po
    kluczu (kolejność/wielkość liter bez znaczenia):
      • Opisujący  (z danych najnowszego zadania),
      • Konsultujący (z danych — bo konsultant nie musi sam opisywać),
      • Lekarz z aktywnego cennika lekarzy (nawet gdy nie ma jeszcze zadania).
    """
    from app.engine.doctors import doctor_key, _norm
    seen: dict[str, str] = {}

    def add(disp):
        disp = _norm(disp)
        if not disp or disp.lower() in ("nan", "none"):
            return
        k = doctor_key(disp)
        if k and k not in seen:
            seen[k] = disp

    jid = job_id or _latest_full_job_id()
    if jid:
        p = participants(jid)
        for d in p.get("doctors", []):
            add(d["name"])
        for d in p.get("consultants", []):
            add(d["name"])
    csv = _active_doctor_cennik_csv()
    if csv:
        try:
            import pandas as pd
            df = pd.read_csv(csv, sep=";", encoding="utf-8-sig", decimal=",")
            if "Lekarz" in df.columns:
                for v in df["Lekarz"].dropna().unique():
                    add(v)
        except Exception:  # noqa: BLE001
            pass
    return {"names": sorted(seen.values(), key=lambda s: s.lower())}


@router.put("/excluded")
async def set_excluded(payload: dict):
    """Zapisuje listę kluczy lekarzy wyłączonych z rozliczenia (w ustawieniach)."""
    keys = payload.get("keys", [])
    if not isinstance(keys, list):
        raise HTTPException(400, "Pole 'keys' musi być listą.")
    cfg = db.get_settings()
    cfg["doctors_excluded"] = sorted({str(k) for k in keys if k})
    db.save_settings(cfg)
    return {"ok": True, "doctors_excluded": cfg["doctors_excluded"]}
