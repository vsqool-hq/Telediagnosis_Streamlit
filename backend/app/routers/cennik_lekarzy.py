"""
Router konwertera cennika lekarzy (skoroszyt ZOBOWIĄZANIA LEKARZY → Lekarz;Kategoria;Cena).

Analogiczny do routera cennika jednostek, ale zapisuje wersje rodzaju
„cennik_lekarzy" i używa konwertera app.engine.cennik_lekarzy_convert.
"""

import os
import io
import uuid
import datetime

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse

from app import db
from app.storage import TMP_DIR, ensure_dirs, version_dir, safe_id, safe_filename
from app.engine.cennik_lekarzy_convert import convert_workbook, rows_to_csv

router = APIRouter(prefix="/api/cennik-lekarzy", tags=["cennik_lekarzy"])


def _now():
    return datetime.datetime.now().isoformat(timespec="seconds")


def _tmp_path(conv_id: str) -> str:
    return os.path.join(TMP_DIR, f"lek_{safe_id(conv_id)}.csv")


def _tmp_xlsx_path(conv_id: str) -> str:
    return os.path.join(TMP_DIR, f"lek_{safe_id(conv_id)}.xlsx")


@router.post("/convert")
async def convert(file: UploadFile = File(...)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Wgraj plik Excel (.xlsx lub .xls).")
    ensure_dirs()
    content = await file.read()

    try:
        result = convert_workbook(io.BytesIO(content))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"Nie udało się przetworzyć pliku: {e}")

    if not result["rows"]:
        raise HTTPException(400, "Nie znaleziono żadnych stawek w pliku. Sprawdź układ arkuszy.")

    conv_id = uuid.uuid4().hex[:12]
    with open(_tmp_path(conv_id), "w", encoding="utf-8") as f:
        f.write(rows_to_csv(result["rows"]))
    # Zachowujemy ORYGINALNY skoroszyt — to „plik zobowiązań", który po rozliczeniu
    # uzupełnimy o ilości okolic. Zapisywany na stałe przy zapisie wersji (save).
    with open(_tmp_xlsx_path(conv_id), "wb") as f:
        f.write(content)

    result_preview = [
        {"lekarz": l, "kategoria": k, "cena": p} for l, k, p in result["rows"][:60]
    ]
    return {
        "id": conv_id,
        "source_name": file.filename,
        "result_preview": result_preview,
        "validation": result["validation"],
    }


@router.get("/commitments-status")
async def commitments_status():
    """
    Czy aktywny cennik lekarzy ma zapisany „plik zobowiązań" (source.xlsx) do
    uzupełniania ilości okolic? Zwraca nazwę i liczbę arkuszy (lekarzy), żeby od
    razu było widać, że wgranie xlsx przez konwerter się powiodło.
    """
    from app.engine.commitments import active_commitments_workbook
    path, name = active_commitments_workbook()
    if not path:
        return {"available": False,
                "reason": "Aktywny cennik lekarzy nie ma zapisanego pliku .xlsx. "
                          "Wgraj ZOBOWIĄZANIA LEKARZY (.xlsx) przez konwerter."}
    info = {"available": True, "name": name, "size": os.path.getsize(path)}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        sheets = [s for s in wb.sheetnames if not s.strip().upper().startswith("ZBIORCZO")]
        info["doctor_sheets"] = len(sheets)
        wb.close()
    except Exception as e:  # noqa: BLE001
        info["read_error"] = str(e)
    return info


@router.get("/convert/{conv_id}/download")
async def download_converted(conv_id: str):
    path = _tmp_path(conv_id)
    if not os.path.isfile(path):
        raise HTTPException(404, "Wynik konwersji wygasł. Wgraj plik ponownie.")
    return FileResponse(path, filename="Cennik_lekarzy.csv", media_type="text/csv")


@router.post("/convert/{conv_id}/save")
async def save_converted(conv_id: str, label: str = Form(""), filename: str = Form("Cennik_lekarzy.csv")):
    path = _tmp_path(conv_id)
    if not os.path.isfile(path):
        raise HTTPException(404, "Wynik konwersji wygasł. Wgraj plik ponownie.")
    filename = safe_filename(filename, "Cennik_lekarzy.csv")
    if not filename.lower().endswith(".csv"):
        filename = "Cennik_lekarzy.csv"

    version_id = uuid.uuid4().hex[:12]
    vdir = version_dir("cennik_lekarzy", version_id)
    os.makedirs(vdir, exist_ok=True)
    dest = os.path.join(vdir, filename)
    with open(path, "rb") as src, open(dest, "wb") as out:
        data = src.read()
        out.write(data)

    # „Plik zobowiązań": zachowaj oryginalny skoroszyt jako source.xlsx (do
    # uzupełniania ilości okolic po rozliczeniu) wraz z jego oryginalną nazwą.
    xlsx_tmp = _tmp_xlsx_path(conv_id)
    if os.path.isfile(xlsx_tmp):
        with open(xlsx_tmp, "rb") as src, open(os.path.join(vdir, "source.xlsx"), "wb") as out:
            out.write(src.read())
        if label:
            with open(os.path.join(vdir, "source_name.txt"), "w", encoding="utf-8") as f:
                f.write(label if label.lower().endswith((".xlsx", ".xls")) else f"{label}.xlsx")

    make_active = db.get_active_version("cennik_lekarzy") is None
    db.add_version({
        "id": version_id, "kind": "cennik_lekarzy", "filename": filename,
        "original_name": filename, "label": label or "Z konwersji cennika lekarzy",
        "size": len(data), "is_active": 1 if make_active else 0, "uploaded_at": _now(),
    })
    for tmp in (path, xlsx_tmp):
        try:
            os.remove(tmp)
        except OSError:
            pass
    return db.get_version(version_id)
