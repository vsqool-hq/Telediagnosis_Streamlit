"""
Zamiana formuł Excela na WARTOŚCI — bez LibreOffice i bez zewnętrznych bibliotek.

Pliki lekarzy (i jednostek) generujemy w openpyxl z formułami. Ta sama paczka ma
zawierać drugą wersję każdego pliku, gdzie wszystkie formuły są już PRZELICZONE na
liczby (klient chce „martwy" plik do dalszej obróbki / wysyłki).

openpyxl nie liczy formuł, a obrazu produkcyjnego (python:3.12-slim) nie chcemy
obciążać LibreOffice. Na szczęście SAMI generujemy formuły, więc ich gramatyka jest
MAŁA i ZAMKNIĘTA — dokładnie:

    funkcje:  IF, IFERROR, VALUE, LEFT, OR, SUBTOTAL(9, zakres)
    operatory: >  +  -  *  /   oraz literał procentu „50%"
    operandy:  liczby, odwołania do komórek ($?LIT$?NUM), zakresy A1:A5

Wszystkie odwołania są w OBRĘBIE tego samego arkusza (brak „Arkusz!A1"). Dzięki temu
piszemy tu prosty, w pełni kontrolowany ewaluator (rekurencyjny zjazd + AST, żeby
IF/IFERROR/OR liczyły się LENIWIE), z memoizacją komórek. Poprawność weryfikujemy
na realnych danych, porównując przeliczone sumy z niezależnym rachunkiem silnika.
"""

import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


class FormulaError(Exception):
    """Błąd liczenia (jak #VALUE!/#DIV/0!) — łapany przez IFERROR / SUBTOTAL."""


# --- Tokenizer ---------------------------------------------------------------
_TOKEN_RE = re.compile(
    r"""
      (?P<WS>\s+)
    | (?P<NUM>\d+(?:\.\d+)?%?)
    | (?P<CELL>\$?[A-Z]{1,3}\$?\d+)          # musi być PRZED IDENT (ma cyfry)
    | (?P<IDENT>[A-Z]+)
    | (?P<OP>[(),:>+\-*/])
    """,
    re.VERBOSE,
)


def _tokenize(s: str):
    toks = []
    i, n = 0, len(s)
    while i < n:
        m = _TOKEN_RE.match(s, i)
        if not m:
            raise FormulaError(f"Nieznany znak w formule: {s[i:i+12]!r}")
        i = m.end()
        if m.lastgroup == "WS":
            continue
        toks.append((m.lastgroup, m.group()))
    toks.append(("EOF", ""))
    return toks


# --- Parser (rekurencyjny zjazd) → AST ---------------------------------------
# AST: ('num', float) | ('cell', 'D5') | ('range', 'A1', 'A5')
#    | ('func', 'IF', [arg,...]) | ('bin', '>', left, right)
class _Parser:
    def __init__(self, toks):
        self.toks = toks
        self.k = 0

    def _peek(self):
        return self.toks[self.k]

    def _next(self):
        t = self.toks[self.k]
        self.k += 1
        return t

    def _expect(self, val):
        kind, v = self._next()
        if v != val:
            raise FormulaError(f"Oczekiwano {val!r}, jest {v!r}")

    def parse(self):
        node = self._cmp()
        if self._peek()[0] != "EOF":
            raise FormulaError(f"Nadmiarowy token: {self._peek()[1]!r}")
        return node

    def _cmp(self):
        node = self._add()
        while self._peek()[1] == ">":
            self._next()
            node = ("bin", ">", node, self._add())
        return node

    def _add(self):
        node = self._mul()
        while self._peek()[1] in ("+", "-"):
            op = self._next()[1]
            node = ("bin", op, node, self._mul())
        return node

    def _mul(self):
        node = self._primary()
        while self._peek()[1] in ("*", "/"):
            op = self._next()[1]
            node = ("bin", op, node, self._primary())
        return node

    def _primary(self):
        kind, v = self._peek()
        if v == "(":
            self._next()
            node = self._cmp()
            self._expect(")")
            return node
        if kind == "NUM":
            self._next()
            if v.endswith("%"):
                return ("num", float(v[:-1]) / 100.0)
            return ("num", float(v))
        if kind == "IDENT":
            self._next()
            self._expect("(")
            args = []
            if self._peek()[1] != ")":
                args.append(self._cmp())
                while self._peek()[1] == ",":
                    self._next()
                    args.append(self._cmp())
            self._expect(")")
            return ("func", v, args)
        if kind == "CELL":
            self._next()
            c1 = v.replace("$", "")
            if self._peek()[1] == ":":
                self._next()
                k2, v2 = self._next()
                if k2 != "CELL":
                    raise FormulaError("Zakres wymaga drugiej komórki")
                return ("range", c1, v2.replace("$", ""))
            return ("cell", c1)
        raise FormulaError(f"Nieoczekiwany token: {v!r}")


# --- Ewaluator AST -----------------------------------------------------------
def _to_num(x):
    if x is None or x == "":
        return 0.0
    if isinstance(x, bool):
        return 1.0 if x else 0.0
    if isinstance(x, (int, float)):
        return float(x)
    raise FormulaError(f"Nie-liczba w działaniu: {x!r}")


def _truthy(x):
    if isinstance(x, bool):
        return x
    return _to_num(x) != 0.0


_CELL_RE = re.compile(r"^([A-Z]{1,3})(\d+)$")


def _expand_range(a, b):
    m1, m2 = _CELL_RE.match(a), _CELL_RE.match(b)
    if not (m1 and m2):
        raise FormulaError(f"Zły zakres {a}:{b}")
    c1, r1 = column_index_from_string(m1.group(1)), int(m1.group(2))
    c2, r2 = column_index_from_string(m2.group(1)), int(m2.group(2))
    for r in range(min(r1, r2), max(r1, r2) + 1):
        for c in range(min(c1, c2), max(c1, c2) + 1):
            yield f"{get_column_letter(c)}{r}"


class _SheetEval:
    """Liczy wartości komórek jednego arkusza z memoizacją (raw = surowe wartości/formuły)."""

    def __init__(self, raw: dict):
        self.raw = raw          # coord -> surowa wartość (liczba / tekst / '=formuła')
        self.memo = {}
        self.busy = set()

    def cell(self, coord):
        if coord in self.memo:
            return self.memo[coord]
        if coord in self.busy:
            raise FormulaError(f"Odwołanie cykliczne: {coord}")
        raw = self.raw.get(coord)
        if isinstance(raw, str) and raw.startswith("="):
            self.busy.add(coord)
            try:
                val = self.eval(_Parser(_tokenize(raw[1:])).parse())
            finally:
                self.busy.discard(coord)
        else:
            val = raw            # liczba, tekst albo None
        self.memo[coord] = val
        return val

    def eval(self, node):
        t = node[0]
        if t == "num":
            return node[1]
        if t == "cell":
            return self.cell(node[1])
        if t == "range":
            return [self.cell(c) for c in _expand_range(node[1], node[2])]
        if t == "bin":
            _, op, ln, rn = node
            if op == ">":
                return _to_num(self.eval(ln)) > _to_num(self.eval(rn))
            a, b = _to_num(self.eval(ln)), _to_num(self.eval(rn))
            if op == "+":
                return a + b
            if op == "-":
                return a - b
            if op == "*":
                return a * b
            if op == "/":
                if b == 0:
                    raise FormulaError("Dzielenie przez zero")
                return a / b
            raise FormulaError(f"Nieznany operator {op}")
        if t == "func":
            return self._func(node[1], node[2])
        raise FormulaError(f"Nieznany węzeł AST {t}")

    def _func(self, name, args):
        if name == "IF":
            cond = _truthy(self.eval(args[0]))
            if cond:
                return self.eval(args[1])
            return self.eval(args[2]) if len(args) > 2 else False
        if name == "IFERROR":
            try:
                return self.eval(args[0])
            except FormulaError:
                return self.eval(args[1])
        if name == "OR":
            return any(_truthy(self.eval(a)) for a in args)
        if name == "VALUE":
            s = self.eval(args[0])
            try:
                return float(str(s).strip().replace(" ", "").replace(",", "."))
            except (TypeError, ValueError):
                raise FormulaError(f"VALUE: nie-liczba {s!r}")
        if name == "LEFT":
            s = self.eval(args[0])
            n = int(_to_num(self.eval(args[1]))) if len(args) > 1 else 1
            if isinstance(s, float) and s.is_integer():
                s = str(int(s))
            elif s is None:
                s = ""
            else:
                s = str(s)
            return s[:n]
        if name == "SUBTOTAL":
            # SUBTOTAL(9, zakres) = SUMA; pojedyncze błędy/teksty pomijamy (jak AGGREGATE).
            total = 0.0
            for a in args[1:]:
                vals = self.eval(a)
                if not isinstance(vals, list):
                    vals = [vals]
                for v in vals:
                    try:
                        total += _to_num(v)
                    except FormulaError:
                        pass
            return total
        raise FormulaError(f"Nieobsługiwana funkcja {name}")


def evaluate_workbook_to_values(src_path: str, dst_path: str) -> dict:
    """Wczytuje skoroszyt z formułami, PRZELICZA każdą formułę na wartość i zapisuje
    kopię pod dst_path (style/formaty/kolory zostają — zmieniamy tylko .value).
    Zwraca {cells, errors} — liczba przeliczonych komórek i nieprzeliczonych (jeśli
    jakaś formuła wyjdzie poza znaną gramatykę, zostaje w pliku bez zmian)."""
    wb = load_workbook(src_path, data_only=False)
    cells_done, errors = 0, 0
    for ws in wb.worksheets:
        raw = {}
        formula_coords = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                raw[cell.coordinate] = v
                if isinstance(v, str) and v.startswith("="):
                    formula_coords.append(cell.coordinate)
        if not formula_coords:
            continue
        ev = _SheetEval(raw)
        for coord in formula_coords:
            try:
                val = ev.cell(coord)
                if isinstance(val, bool):       # Excel logiczne → PRAWDA/FAŁSZ jako liczba
                    val = 1 if val else 0
                ws[coord].value = val
                cells_done += 1
            except FormulaError:
                errors += 1                      # zostawiamy oryginalną formułę
    wb.save(dst_path)
    return {"cells": cells_done, "errors": errors}
