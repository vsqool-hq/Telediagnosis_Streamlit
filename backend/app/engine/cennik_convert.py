"""
Konwerter cennika zbiorczego (szeroki Excel) → prosty cennik 3-kolumnowy
(BADANIE;Jednostka;Cena), zgodny z tym, co dotychczas powstawało w Power Query.

Układ pliku źródłowego ("ZBIORCZO ..."):
  * wiersz 1: nagłówek — A = "BADANIE", kolejne kolumny = nazwy jednostek,
    a po nich kolumny pomocnicze (MEDIANA, ŚREDNIA, MAX, MIN, SUMA, marże, notatki).
  * wiersze: nazwa badania w kol. A + ceny w kolumnach jednostek. Występują też
    wiersze-nagłówki sekcji (RTG/TK/MR/MMG) i puste separatory.

Reguły konwersji (odtworzone z istniejącego cennika i ustaleń z użytkownikiem):
  * kolumny jednostek = od drugiej kolumny do PIERWSZEJ kolumny pomocniczej
    (nagłówek pusty lub jeden z: MEDIANA/ŚREDNIA/MAX/MIN/SUMA) — nazwy jednostek
    kończą się na kolumnie „Mediana",
  * wiersz jest "badaniem", jeśli ma ≥1 wartość liczbową w kolumnach jednostek
    (to odcina nagłówki sekcji oraz wiersze etykiet typu JEDNOSTKA/UWAGI),
  * puste wiersze-separatory sekcji są pomijane (pusta kolumna A),
  * wiersze z etykietą ze SKIP_LABELS (np. „Wpis SIMP") są pomijane,
  * wiersz „WSPARCIE" jest ostatnim importowanym — wszystko PONIŻEJ niego pomijamy,
  * w takim wierszu: liczba → cena, "-" → 0, pusta komórka → pomijana,
  * ceny zaokrąglane do 2 miejsc; "brudne" zapisy (np. "2 500,00 zł") naprawiane
    i raportowane.
"""

import re
import io

from openpyxl import load_workbook

STOP_HEADERS = {"MEDIANA", "ŚREDNIA", "SREDNIA", "MAX", "MIN", "SUMA"}

# Etykiety wierszy, których nie importujemy (porównanie bez wielkości liter).
SKIP_LABELS = {"WPIS SIMP"}

# Po przetworzeniu tego wiersza kończymy import — wszystko poniżej pomijamy.
STOP_AFTER_LABEL = "WSPARCIE"


def _clean_name(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _try_number(raw):
    """Zwraca (cena: float|None, repaired: bool, original: str).
    None oznacza wartość nienumeryczną, której nie udało się naprawić."""
    if raw is None:
        return None, False, ""
    if isinstance(raw, (int, float)):
        return round(float(raw), 2), False, str(raw)

    s = str(raw).strip()
    original = s
    if s == "":
        return None, False, s
    if s == "-":
        return 0.0, False, s

    # próba bezpośrednia
    try:
        return round(float(s.replace(",", ".")), 2) if ("," in s and "." not in s) else round(float(s), 2), False, original
    except ValueError:
        pass

    # naprawa "brudnych" zapisów: "2 500,00 zł", "2,500.00zł", "1 234"
    cleaned = (s.replace("zł", "").replace("PLN", "")
               .replace("\xa0", "").replace(" ", "").strip())
    if "," in cleaned and "." in cleaned:
        # przecinek = tysiące, kropka = dziesiętne (np. 2,500.00)
        cleaned = cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        return round(float(cleaned), 2), True, original
    except ValueError:
        return None, False, original


def format_price(p: float) -> str:
    """Format wyjściowy: liczba całkowita bez miejsc, ułamek z przecinkiem."""
    p = round(float(p), 2)
    if p == int(p):
        return str(int(p))
    return ("%.2f" % p).rstrip("0").rstrip(".").replace(".", ",")


def convert_workbook(path_or_bytes) -> dict:
    """
    Konwertuje cennik zbiorczy. Zwraca słownik:
      rows: [(BADANIE, Jednostka, cena_float)]
      units, badania, source_preview, validation
    """
    wb = load_workbook(path_or_bytes, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    grid = list(ws.iter_rows(values_only=True))
    if not grid:
        raise ValueError("Arkusz jest pusty.")

    header = [_clean_name(v) for v in grid[0]]

    # kolumny jednostek: od indeksu 1 do pierwszej kolumny pomocniczej
    # (nazwy jednostek kończą się na kolumnie „Mediana").
    unit_cols: list[int] = []
    for c in range(1, len(header)):
        h = header[c]
        if h == "" or h.upper() in STOP_HEADERS:
            break
        unit_cols.append(c)

    units = [header[c] for c in unit_cols]

    rows: list[tuple[str, str, float]] = []
    repaired: list[dict] = []
    errors: list[dict] = []
    zeros: list[list[str]] = []
    excluded_rows: list[str] = []
    seen: dict[tuple[str, str], int] = {}
    duplicates: list[dict] = []
    per_badanie: dict[str, int] = {}
    n_skipped_label = 0
    stopped = False

    for grow in grid[1:]:
        # Pusty wiersz-separator (pusta kolumna A) — pomijamy.
        badanie = _clean_name(grow[0] if len(grow) > 0 else "")
        if not badanie:
            continue

        # Etykiety do pominięcia (np. „Wpis SIMP").
        if badanie.upper() in SKIP_LABELS:
            n_skipped_label += 1
            excluded_rows.append(f"{badanie} (etykieta)")
            continue

        cells = [(c, grow[c] if c < len(grow) else None) for c in unit_cols]

        has_numeric = any(isinstance(v, (int, float)) or _looks_numeric(v) for _, v in cells)
        if not has_numeric:
            excluded_rows.append(badanie)
            # „WSPARCIE" zwykle ma liczby; gdyby nie miało — i tak kończymy poniżej.
            if badanie.upper() == STOP_AFTER_LABEL:
                stopped = True
                break
            continue

        for c, raw in cells:
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                continue
            price, was_repaired, original = _try_number(raw)
            jednostka = header[c]
            if price is None:
                errors.append({"badanie": badanie, "jednostka": jednostka, "wartosc": original})
                continue
            key = (badanie, jednostka)
            if key in seen:
                duplicates.append({"badanie": badanie, "jednostka": jednostka})
            seen[key] = seen.get(key, 0) + 1
            rows.append((badanie, jednostka, price))
            per_badanie[badanie] = per_badanie.get(badanie, 0) + 1
            if price == 0:
                zeros.append([badanie, jednostka])
            if was_repaired:
                repaired.append({"badanie": badanie, "jednostka": jednostka,
                                 "z": original, "na": format_price(price)})

        # „WSPARCIE" jest ostatnim importowanym wierszem — kończymy import.
        if badanie.upper() == STOP_AFTER_LABEL:
            stopped = True
            break

    # podgląd źródła: nagłówek + pierwsze 12 wierszy, max 8 pierwszych jednostek
    preview_cols = unit_cols[:8]
    source_preview = {
        "header": ["BADANIE"] + [header[c] for c in preview_cols],
        "rows": [
            [_clean_name(g[0] if len(g) else "")] +
            [("" if (c >= len(g) or g[c] is None) else str(g[c])) for c in preview_cols]
            for g in grid[1:13]
        ],
    }

    badania = list(per_badanie.keys())
    prices_only = [p for _, _, p in rows]
    validation = {
        "n_rows": len(rows),
        "n_badania": len(badania),
        "n_units": len(units),
        "n_zeros": len(zeros),
        "n_repaired": len(repaired),
        "n_errors": len(errors),
        "n_duplicates": len(duplicates),
        "n_skipped_label": n_skipped_label,
        "stopped_at_wsparcie": stopped,
        "price_min": min(prices_only) if prices_only else 0,
        "price_max": max(prices_only) if prices_only else 0,
        "zeros_sample": zeros[:50],
        "repaired": repaired[:50],
        "errors": errors[:50],
        "duplicates": duplicates[:50],
        "excluded_rows": excluded_rows,
        "units": units,
        "badania": badania,
        "rows_per_badanie": per_badanie,
    }

    return {"rows": rows, "units": units, "badania": badania,
            "source_preview": source_preview, "validation": validation}


def _looks_numeric(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip()
    if s in ("", "-"):
        return False
    try:
        float(s.replace(",", "."))
        return True
    except ValueError:
        return False


def rows_to_csv(rows: list[tuple[str, str, float]]) -> str:
    """Składa cennik 3-kolumnowy: BADANIE;Jednostka;Cena (przecinek dziesiętny, bez BOM)."""
    out = io.StringIO()
    out.write("BADANIE;Jednostka;Cena\n")
    for badanie, jednostka, price in rows:
        out.write(f"{badanie};{jednostka};{format_price(price)}\n")
    return out.getvalue()
