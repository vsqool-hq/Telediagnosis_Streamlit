"""
Generator uzupełnień cen jednostek z pliku ZOBOWIĄZANIA (SZPITALE).

IDEA (po korekcie 2026-08): współczynniki i dopisania do cennika mają WYPEŁNIAĆ LUKI —
czyli stawki, które są w ostatnim aneksie danej jednostki, ale których BRAK w aktualnym
cenniku. Jeśli cennik ma już daną stawkę wprost (kwotą) — NIC nie proponujemy (to była
przyczyna setek zbędnych, „dziwnych" ułamków: generator liczył derived/base nawet gdy
cennik miał okrągłą kwotę, a i tak wygrywała cena z cennika).

Dla KAŻDEJ luki (badanie POCHODNE: PORÓWNAWCZE / PORÓW. / ONKO / ANGIO — bazowych nie
ruszamy) sprawdzamy, jak stawka jest zapisana w ostatnim aneksie (kolumna stawek):
  • FORMUŁA `=<komórka_bazy>*k`  → WSPÓŁCZYNNIK {base, factor}: `factor` bierzemy WPROST
    z formuły (czyste 0.4 / 1.25 / 1.4), a `base` z wiersza, do którego formuła się
    odwołuje. To wierne odwzorowanie reguły umownej „pochodne = baza × %".
  • LITERAŁ (liczba)          → DODANIE DO CENNIKA {unit, key, amount} — do zatwierdzenia
    w oknie potwierdzenia (dopiero potem tworzymy nową wersję cennika).

Kanoniczną pisownię klucza (np. MR „stawy"/„angio" małą literą, „PORÓWNAWCZE" wielką)
bierzemy ze SŁOWNIKA KLUCZY istniejącego cennika (te same klucze produkuje build_price_key),
żeby współczynnik faktycznie zaskakiwał w rozliczeniu, a dodanie trafiało w istniejący format.

Stawki bierzemy WYŁĄCZNIE z NAJNOWSZEGO aneksu (ostatnia od prawej kolumna stawek z
dodatnimi wartościami). Wynik jest PROPOZYCJĄ do zatwierdzenia przez człowieka.
"""

import re
import datetime as dt

SKIP_SHEETS = {"LEKARZE", "TABELA", "ZBIORCZO 2026", "FORMUŁY (5)"}
_MODS = ("PORÓWNAWCZE", "PORÓW.", "PORÓW", "ONKO", "ANGIO")


def _cmp(s) -> str:
    """Klucz porównawczy badania: zwinięte spacje + wielkie litery (do dopasowań)."""
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip()).upper()


def _unorm(s) -> str:
    """Klucz porównawczy nazwy jednostki: bez diakrytyków, małe litery."""
    s = str(s if s is not None else "").replace("ł", "l").replace("Ł", "L")
    import unicodedata
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s.lower().strip()


def _is_derived(cmp_key: str) -> bool:
    toks = cmp_key.split(" ")
    return any(m in toks for m in _MODS)


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


# =… formuła typu „baza × współczynnik": jedna referencja komórki i jedna liczba, operator „*".
_CELL_RE = re.compile(r"\$?[A-Z]{1,3}\$?(\d+)")


def _parse_factor_formula(formula: str):
    """Z formuły `=B9*1.25` / `=1.25*B9` zwraca (base_row:int, factor:float) albo (None, None).
    Wymaga DOKŁADNIE jednej referencji komórki, jednej liczby i mnożenia."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return None, None
    f = formula[1:].replace("$", "")
    if "/" in f or "+" in f or "-" in f:      # tylko czyste mnożenie
        return None, None
    rows = _CELL_RE.findall(f)
    if len(rows) != 1:
        return None, None
    f_nolet = re.sub(r"[A-Z]{1,3}\d+", "", f)   # usuń referencję → zostaje mnożnik
    nums = re.findall(r"\d+(?:[.,]\d+)?", f_nolet)
    if len(nums) != 1:
        return None, None
    try:
        return int(rows[0]), float(nums[0].replace(",", "."))
    except (TypeError, ValueError):
        return None, None


def _build_cennik_index(cennik_rows):
    """cennik_rows: iterowalne (unit, badanie, cena). Zwraca:
      per_unit:  {unit_norm: {cmp_key: cena>0}},
      canon:     {cmp_key: kanoniczna_pisownia} (globalnie, ze wszystkich jednostek),
      unit_name: {unit_norm: pisownia_jednostki_z_cennika}."""
    per_unit, canon, unit_name = {}, {}, {}
    for unit, bad, cena in cennik_rows or []:
        u = _unorm(unit)
        try:
            c = float(cena)
        except (TypeError, ValueError):
            c = 0.0
        ck = _cmp(bad)
        if ck and ck not in canon:
            canon[ck] = re.sub(r"\s+", " ", str(bad).strip())
        unit_name.setdefault(u, str(unit).strip())
        if c > 0:
            per_unit.setdefault(u, {})[ck] = c
    return per_unit, canon, unit_name


def generate_adjustments(path_or_bytes, cennik_rows=None) -> dict:
    """Zwraca:
      { coefficients: { jednostka: { KLUCZ: {base, factor} } },
        cennik_additions: [ {unit, key, amount, base_row_label?} ],
        detail: [ {unit, key, kind, factor?, base?, amount?} ],
        current: {…},                # bieżące współczynniki (do różnicy w UI)
        stats: {…}, warning: str|None }
    """
    from openpyxl import load_workbook
    per_unit, canon, unit_name = _build_cennik_index(cennik_rows)

    # Dwa odczyty: wartości (do wykrycia aneksu/luki) i formuły (do rozpoznania baza×k).
    wbv = load_workbook(path_or_bytes, data_only=True)
    if hasattr(path_or_bytes, "seek"):
        path_or_bytes.seek(0)
    wbf = load_workbook(path_or_bytes, data_only=False)

    coefficients: dict = {}
    cennik_additions: list = []
    detail: list = []
    redundant = skipped_no_canon = sheets_scanned = 0

    for sheet in wbv.sheetnames:
        if _cmp(sheet) in SKIP_SHEETS:
            continue
        wsv = wbv[sheet]
        wsf = wbf[sheet] if sheet in wbf.sheetnames else None
        maxr, maxc = wsv.max_row, wsv.max_column
        if not maxr:
            continue

        # wiersz nagłówka = ten z 'RTG' w kol. A (pierwsze 6 wierszy)
        hdr_i = next((r for r in range(1, min(maxr, 6) + 1) if _cmp(wsv.cell(r, 1).value) == "RTG"), None)
        if hdr_i is None:
            continue
        # kolumny stawek: nagłówek zawiera 'STAWK' albo kolumna B (idx 2), byle nie data
        stawka_cols = sorted({
            c for c in range(2, maxc + 1)
            if not isinstance(wsv.cell(hdr_i, c).value, dt.datetime)
            and ("STAWK" in _cmp(wsv.cell(hdr_i, c).value) or c == 2)
        })
        if not stawka_cols:
            continue

        # wiersze badań: (excel_row, cmp_key, label_raw)
        body = []
        for r in range(hdr_i + 1, maxr + 1):
            label = wsv.cell(r, 1).value
            ck = _cmp(label)
            if not ck or "SUMA" in ck or ck in ("RTG", "TK", "MR", "MMG", "ILOŚĆ OKOLIC"):
                continue
            body.append((r, ck, re.sub(r"\s+", " ", str(label).strip())))

        # najnowszy aneks = ostatnia od prawej kolumna stawek z jakąkolwiek dodatnią stawką
        cur_col = next((c for c in reversed(stawka_cols)
                        if any((_num(wsv.cell(r, c).value) or 0) > 0 for r, _, _ in body)), None)
        if cur_col is None:
            continue
        sheets_scanned += 1

        u = _unorm(sheet)
        unit_cennik = per_unit.get(u, {})
        row_to_ck = {r: ck for r, ck, _ in body}

        rules: dict = {}
        for r, ck, label in body:
            if not _is_derived(ck):
                continue
            dv = _num(wsv.cell(r, cur_col).value)
            if not dv or dv <= 0:                      # brak stawki pochodnej w aneksie
                continue
            if unit_cennik.get(ck, 0) > 0:             # cennik ma już tę stawkę → nie proponujemy
                redundant += 1
                continue

            key = canon.get(ck)                        # kanoniczna pisownia (z cennika)
            if not key:
                # Klucz NIE występuje w żadnym cenniku → to nie jest realny klucz
                # rozliczeniowy (wiersze wolnotekstowe/notatki w ZOBOWIĄZANIACH, np.
                # „TK ANGIO CITO głowy i szyi !!"). Pomijamy — nie zaśmiecamy cennika.
                skipped_no_canon += 1
                continue

            formula = wsf.cell(r, cur_col).value if wsf is not None else None
            base_row, factor = _parse_factor_formula(formula)
            if base_row and factor and base_row in row_to_ck:
                base_ck = row_to_ck[base_row]
                base_key = canon.get(base_ck) or re.sub(r"\s+", " ", str(wsv.cell(base_row, 1).value or "").strip())
                rules[key] = {"base": base_key, "factor": round(factor, 6)}
                detail.append({"unit": sheet, "key": key, "kind": "coefficient",
                               "base": base_key, "factor": round(factor, 6), "amount": round(dv, 2)})
            else:
                # literał (kwota) albo formuła bez czytelnej postaci baza×k → do cennika
                cennik_additions.append({"unit": unit_name.get(u, str(sheet).strip()),
                                         "key": key, "amount": round(dv, 2)})
                detail.append({"unit": sheet, "key": key, "kind": "cennik", "amount": round(dv, 2)})
        if rules:
            coefficients[sheet] = rules

    wbv.close()
    wbf.close()

    warning = None
    if sheets_scanned == 0:
        warning = ("Nie znaleziono arkuszy jednostek ze stawkami. Upewnij się, że to pełny plik "
                   "ZOBOWIĄZANIA SZPITALE (arkusze per jednostka), a nie sam arkusz zbiorczy.")

    detail.sort(key=lambda d: (d["unit"], d["key"]))
    cennik_additions.sort(key=lambda a: (a["unit"], a["key"]))
    return {
        "coefficients": coefficients,
        "cennik_additions": cennik_additions,
        "detail": detail,
        "stats": {
            "units": len(coefficients),
            "coefficients": sum(len(v) for v in coefficients.values()),
            "cennik_additions": len(cennik_additions),
            "redundant_skipped": redundant,
            "no_canon": skipped_no_canon,
            "sheets_scanned": sheets_scanned,
        },
        "warning": warning,
    }
