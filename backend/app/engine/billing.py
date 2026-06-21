"""
Silnik rozliczeniowy (Etap 1: weryfikacja, Etap 2: rozliczenie).

Logika przeniesiona 1:1 z oryginalnego backend.py aplikacji desktopowej.
Jedyna zmiana: wartości wcześniej zahardkodowane są teraz wczytywane z
konfiguracji (app.engine.config), dzięki czemu można je edytować z panelu
Ustawienia bez modyfikacji kodu.
"""

import pandas as pd
import numpy as np
import re
import os
import glob
from pathlib import Path
from openpyxl.styles import PatternFill, NamedStyle, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import warnings
import multiprocessing

from app.engine.config import load_config

warnings.filterwarnings('ignore')

# Konfiguracja wczytywana raz, przy imporcie modułu (świeżo w każdym subprocessie).
CONFIG = load_config()

PRIORITY_DICT = CONFIG["priority_dict"]
PRIORITY_MAP = CONFIG["priority_map"]
TK_SUFFIX_MAP = CONFIG["tk_suffix_map"]
MR_GLKRG_KEYWORDS = CONFIG["mr_glkrg_keywords"]
MR_STAWY_KEYWORDS = CONFIG["mr_stawy_keywords"]
MASTER_PRIORITY_ORDER = CONFIG["master_priority_order"]
PRIORITY_COLORS = CONFIG["priority_colors"]
PROCEDURE_TYPE_COLORS = CONFIG["procedure_type_colors"]
# Jedna liczba rdzeni dla całego procesu. 0 = Auto (wszystkie dostępne).
# Niezależnie od ustawienia, wartość jest ograniczana do liczby fizycznych
# rdzeni maszyny — na komputerze z 4 rdzeniami wpisanie 8 nic nie przyspieszy.
_CPU = multiprocessing.cpu_count() or 1
_n = CONFIG.get("num_processes")
if _n is None:  # zgodność wstecz ze starą konfiguracją (osobne etapy)
    _n = max(int(CONFIG.get("num_processes_verify", 4)),
             int(CONFIG.get("num_processes_billing", 4)))
_n = int(_n)
if _n <= 0:     # 0 = Auto
    _n = _CPU
_n = max(1, min(_n, _CPU))
NUM_PROCESSES_VERIFY = _n
NUM_PROCESSES_BILLING = _n


# ###################################################################################
# ### ETAP 1: AGENT WERYFIKACJI POPRAWNOŚCI BADAŃ                                ###
# ###################################################################################

class MedicalVerificationAgent:
    def __init__(self, reference_dataframe, input_folder, output_folder):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.reference_data = self.validate_reference_data(reference_dataframe)

    def validate_reference_data(self, df):
        if df is None or df.empty:
            raise ValueError("Przekazany DataFrame wzorcowy jest pusty.")
        required_ref_cols = ['Procedura', 'Rodzaj procedury rozlicz.', 'Ilość Okolic']
        if not all(col in df.columns for col in required_ref_cols):
            missing_cols = [col for col in required_ref_cols if col not in df.columns]
            raise ValueError(f"W połączonych danych wzorcowych brakuje kluczowych kolumn: {missing_cols}")
        df_validated = df.dropna(how='all').copy()
        for col in required_ref_cols:
            df_validated[col] = df_validated[col].astype(str)

        if 'Rodzaj procedury sprawdzony' not in df_validated.columns:
            print("! OSTRZEŻENIE: Brak kolumny 'Rodzaj procedury sprawdzony' w plikach wzorcowych.", flush=True)
            df_validated['Rodzaj procedury sprawdzony'] = ''
        else:
            df_validated['Rodzaj procedury sprawdzony'] = df_validated['Rodzaj procedury sprawdzony'].astype(str)

        print(f"✓ Pomyślnie załadowano i zwalidowano {len(df_validated)} rekordów wzorcowych.", flush=True)
        return df_validated

    def extract_number_from_procedure(self, procedure_text):
        if pd.isna(procedure_text):
            return 1
        match = re.search(r'^(\d+)', str(procedure_text))
        return int(match.group(1)) if match else 1

    def find_recommended_regions(self, rodzaj_procedury, procedura):
        if self.reference_data is None or pd.isna(procedura) or pd.isna(rodzaj_procedury):
            return None
        rodzaj_procedury_str = str(rodzaj_procedury).lower().strip()
        procedura_str = str(procedura).lower().strip()
        matches = self.reference_data[
            (self.reference_data['Procedura'].str.lower().str.strip() == procedura_str) &
            (self.reference_data['Rodzaj procedury rozlicz.'].str.lower().str.strip() == rodzaj_procedury_str)
        ]
        if matches.empty:
            return None
        max_regions = pd.to_numeric(matches['Ilość Okolic'], errors='coerce').max()
        return int(max_regions) if pd.notna(max_regions) else None

    def get_procedure_type_priority(self, modalnosc, rodzaj_procedury):
        if pd.isna(modalnosc) or pd.isna(rodzaj_procedury):
            return None
        modalnosc_str = str(modalnosc).strip().upper()
        rodzaj_str = str(rodzaj_procedury).strip()
        if modalnosc_str in PRIORITY_DICT:
            return PRIORITY_DICT[modalnosc_str].get(rodzaj_str, None)
        return None

    def find_recommended_procedure_type(self, procedura, modalnosc):
        if self.reference_data is None or pd.isna(procedura) or pd.isna(modalnosc):
            return None
        procedura_str = str(procedura).lower().strip()
        modalnosc_str = str(modalnosc).strip().upper()
        matches = self.reference_data[
            (self.reference_data['Procedura'].str.lower().str.strip() == procedura_str) &
            (self.reference_data['Rodzaj procedury sprawdzony'].notna()) &
            (self.reference_data['Rodzaj procedury sprawdzony'].str.strip() != '')
        ]
        if matches.empty:
            return None
        best_priority = float('-inf')
        best_rodzaj = None
        for _, match_row in matches.iterrows():
            rodzaj = str(match_row['Rodzaj procedury sprawdzony']).strip()
            priority = self.get_procedure_type_priority(modalnosc_str, rodzaj)
            if priority is not None and priority > best_priority:
                best_priority = priority
                best_rodzaj = rodzaj
        return best_rodzaj

    def correct_procedure_types(self, df, col_map):
        corrections_made = []
        rows_corrected_indices = []
        if 'Modalność' not in df.columns:
            return corrections_made, rows_corrected_indices
        for index in df.index:
            row = df.loc[index]
            procedura = row.get(col_map.get('procedura'))
            modalnosc = row.get('Modalność')
            current_rodzaj = row.get(col_map.get('rodzaj_procedury_rozlicz'))
            if pd.isna(procedura) or pd.isna(modalnosc) or pd.isna(current_rodzaj):
                continue
            recommended_rodzaj = self.find_recommended_procedure_type(procedura, modalnosc)
            if recommended_rodzaj is None:
                continue
            current_priority = self.get_procedure_type_priority(modalnosc, current_rodzaj)
            recommended_priority = self.get_procedure_type_priority(modalnosc, recommended_rodzaj)
            if current_priority is None or recommended_priority is None:
                continue
            if recommended_priority > current_priority:
                df.loc[index, col_map['rodzaj_procedury_rozlicz']] = recommended_rodzaj
                rows_corrected_indices.append(index)
                corrections_made.append(
                    f"Wiersz {index + 2}: '{procedura}' -> Zmieniono rodzaj procedury z '{current_rodzaj}' "
                    f"(priorytet {current_priority}) na '{recommended_rodzaj}' (priorytet {recommended_priority})."
                )
        return corrections_made, rows_corrected_indices

    def format_corrected_procedure(self, count):
        if count == 1:
            return "1 okolica anatomiczna"
        return f"{count} okolice anatomiczne" if 2 <= count <= 4 else f"{count} okolic anatomicznych"

    def get_column_mapping(self, df_columns):
        mapping, variants = {}, {
            'procedura': ['Procedura', 'procedura', 'PROCEDURA'],
            'procedura_rozlicz': ['Procedura rozlicz.', 'Procedura rozlicz'],
            'rodzaj_procedury_rozlicz': ['Rodzaj procedury rozlicz.'],
            'klient': ['Klient'],
        }
        for standard, options in variants.items():
            for option in options:
                if option in df_columns:
                    mapping[standard] = option
                    break
        return mapping

    def process_client_data(self, client_df, client_name):
        logs = []
        try:
            logs.append(f"\n--- Weryfikuję dane dla klienta: {client_name} ---")
            df = client_df.copy()
            col_map = self.get_column_mapping(df.columns)
            required_cols_for_processing = ['procedura', 'procedura_rozlicz', 'rodzaj_procedury_rozlicz']
            if not all(k in col_map for k in required_cols_for_processing):
                missing = [k for k in required_cols_for_processing if k not in col_map]
                logs.append(f"! OSTRZEŻENIE: Pomijam klienta {client_name} z powodu braku wymaganych kolumn: {missing}.")
                return None, logs

            type_corrections, type_corrected_rows = self.correct_procedure_types(df, col_map)

            if type_corrections:
                logs.append(f"✓ Dokonano {len(type_corrections)} korekt rodzajów procedur według priorytetów:")
                for msg in type_corrections:
                    logs.append(f"  - {msg}")

            rows_to_correct_indices, corrections_made = [], []
            local_unmatched_records = []

            for index in df.index:
                row = df.loc[index]
                recommended_count = self.find_recommended_regions(
                    rodzaj_procedury=row.get(col_map['rodzaj_procedury_rozlicz']),
                    procedura=row.get(col_map['procedura'])
                )
                if recommended_count:
                    current_count = self.extract_number_from_procedure(row.get(col_map['procedura_rozlicz']))
                    if current_count < recommended_count:
                        corrected_text = self.format_corrected_procedure(recommended_count)
                        df.loc[index, col_map['procedura_rozlicz']] = corrected_text
                        rows_to_correct_indices.append(index)
                        corrections_made.append(f"Wiersz {index + 2}: '{row[col_map['procedura']]}' -> Poprawiono liczbę okolic na {recommended_count}.")
                else:
                    local_unmatched_records.append(row)

            safe_client_name = re.sub(r'[^\w\s-]', '', str(client_name)).strip().replace(' ', '_')
            output_filename = f"Sprawdzony_{safe_client_name}.xlsx"
            output_path = os.path.join(self.output_folder, output_filename)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Szczegółowe', index=False)

            all_corrected_rows = list(set(type_corrected_rows + rows_to_correct_indices))

            if corrections_made or type_corrections:
                if corrections_made:
                    logs.append(f"✓ Dokonano {len(corrections_made)} korekt ilości okolic dla klienta {client_name}:")
                    for msg in corrections_made:
                        logs.append(f"  - {msg}")

                wb = load_workbook(output_path)
                ws = wb['Szczegółowe']
                orange_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
                index_to_excel_row = {original_index: i + 2 for i, original_index in enumerate(df.index)}
                for original_idx in all_corrected_rows:
                    excel_row = index_to_excel_row.get(original_idx)
                    if excel_row:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=excel_row, column=col).fill = orange_fill
                wb.save(output_path)
            else:
                logs.append(f"✓ Dane dla klienta {client_name} poprawne, nie wymagały korekt.")

            logs.append(f"Zapisano zweryfikowany plik jako: {output_filename}")
            return local_unmatched_records, logs
        except Exception as e:
            logs.append(f"BŁĄD podczas weryfikacji danych dla klienta {client_name}: {e}")
            return None, logs

    @staticmethod
    def wrapper_process_client_data(args):
        verifier_instance, client_df, client_name = args
        return verifier_instance.process_client_data(client_df, client_name)

    def run_verification(self):
        all_files = glob.glob(os.path.join(self.input_folder, "*.xlsx")) + glob.glob(os.path.join(self.input_folder, "*.xls"))
        excel_files = [f for f in all_files if not os.path.basename(f).startswith('~$')]

        if not excel_files:
            print("Nie znaleziono pliku Excel w folderze wejściowym 'Jednostki'!", flush=True)
            return False
        if len(excel_files) > 1:
            print(f"OSTRZEŻENIE: Znaleziono {len(excel_files)} plików. Przetwarzam tylko pierwszy: {os.path.basename(excel_files[0])}", flush=True)

        master_file_path = excel_files[0]
        print(f"Wczytuję główny plik z danymi: {os.path.basename(master_file_path)}", flush=True)

        try:
            xls = pd.ExcelFile(master_file_path)
            sheet = "Szczegółowe" if "Szczegółowe" in xls.sheet_names else xls.sheet_names[0]
            if sheet != "Szczegółowe":
                print(f"OSTRZEŻENIE: Brak arkusza 'Szczegółowe' — używam '{sheet}'.", flush=True)
            master_df = pd.read_excel(xls, sheet_name=sheet, header=0)
            if 'Klient' not in master_df.columns:
                print("BŁĄD KRYTYCZNY: W pliku wejściowym brakuje kolumny 'Klient'.", flush=True)
                return False

            unique_clients = master_df['Klient'].dropna().unique()
            print(f"Znaleziono dane dla {len(unique_clients)} unikalnych klientów.", flush=True)

            num_processes = NUM_PROCESSES_VERIFY
            # Zadania niosą tylko dane klienta; słownik wzorcowy ładuje initializer
            # RAZ na workera (a nie 115× w payloadzie) — to oszczędza pamięć.
            tasks = [(master_df[master_df['Klient'] == client], client) for client in unique_clients]

            print(f"Rozpoczynam równoległą weryfikację na {num_processes} rdzeniach...", flush=True)

            all_unmatched_records = []
            processed_clients_count = 0

            # 'spawn' zamiast 'fork' — odporne na zakleszczenie przy forku po
            # wystartowaniu wątku (keep-alive); workery to świeże procesy.
            ctx = multiprocessing.get_context("spawn")
            with ctx.Pool(processes=num_processes,
                          initializer=_verify_init_worker,
                          initargs=(self.reference_data, self.input_folder, self.output_folder)) as pool:
                results = pool.map(_verify_one, tasks)

            for result in results:
                if result is not None:
                    unmatched, logs = result
                    print("\n".join(logs), flush=True)
                    if unmatched is not None:
                        all_unmatched_records.extend(unmatched)
                        processed_clients_count += 1

            print(f"\nZakończono weryfikację. Przetworzono dane dla {processed_clients_count} z {len(unique_clients)} klientów.", flush=True)

            if all_unmatched_records:
                print("\n" + "=" * 25, flush=True)
                print(f"Znaleziono {len(all_unmatched_records)} rekordów bez dopasowania w plikach wzorcowych.", flush=True)
                unmatched_df = pd.DataFrame(all_unmatched_records)
                output_path = os.path.join(self.output_folder, "Rekordy_Bez_Wzorca.xlsx")
                unmatched_df.to_excel(output_path, index=False)
                print("Zapisano je do zbiorczego pliku: Rekordy_Bez_Wzorca.xlsx", flush=True)
                print("=" * 25, flush=True)

            return True
        except Exception as e:
            print(f"BŁĄD KRYTYCZNY podczas wczytywania lub dzielenia pliku {os.path.basename(master_file_path)}: {e}", flush=True)
            return False


# --- Worker weryfikacji (Pool initializer) -----------------------------------
# Słownik wzorcowy ładujemy RAZ na proces workera (initargs), zamiast pakować go
# do każdego z ~115 zadań — to drastycznie obniża zużycie pamięci przy 'spawn'.
_WORKER_VERIFIER = None


def _verify_init_worker(reference_df, input_folder, output_folder):
    global _WORKER_VERIFIER
    _WORKER_VERIFIER = MedicalVerificationAgent(reference_df, input_folder, output_folder)


def _verify_one(args):
    client_df, client_name = args
    return _WORKER_VERIFIER.process_client_data(client_df, client_name)


# ###################################################################################
# ### ETAP 2: PRZETWARZANIE ROZLICZEŃ MEDYCZNYCH                                ###
# ###################################################################################

def bill_extract_multiplier(procedure_text):
    if pd.isna(procedure_text) or procedure_text == "":
        return 1
    numbers = re.findall(r'\d+', str(procedure_text))
    return int(numbers[0]) if numbers else 1


def mr_get_anatomical_category(procedura: str) -> str:
    proc_lower = str(procedura).lower()
    for kw in MR_GLKRG_KEYWORDS:
        if kw in proc_lower:
            return 'GŁ/KRG'
    for kw in MR_STAWY_KEYWORDS:
        if kw in proc_lower:
            return 'stawy'
    return 'inne'


def mr_get_type_suffix(rodzaj_procedury: str) -> str:
    r = str(rodzaj_procedury).strip()
    if r == 'MR Angiografia':
        return 'angio'
    if r == 'MR Onkologia':
        return 'ONKO'
    return ''


def build_price_key(row) -> str:
    modalnosc = str(row.get('Modalność', '')).strip().upper()
    rodzaj = str(row.get('Rodzaj procedury rozlicz.', '')).strip()
    priorytet_raw = str(row.get('Priorytet opisu', '')).strip()
    procedura = str(row.get('Procedura', '')).strip()
    porownawcze = pd.to_numeric(row.get('Badania do porównania', 0), errors='coerce')
    porownawcze = bool(porownawcze) if pd.notna(porownawcze) else False

    priorytet = PRIORITY_MAP.get(priorytet_raw, priorytet_raw)

    if modalnosc == 'RTG':
        return f"RTG {priorytet}"

    if modalnosc in ('MMG', 'MAMMOGRAFIA'):
        # Skryning → "MMG SKRINING", pozostałe mammografie → "MMG".
        # (W danych modalność bywa zapisana jako 'Mammografia' — obsługujemy oba.)
        proc_l = procedura.lower()
        if any(t in proc_l for t in ('skrin', 'skryn', 'skirin')):
            return "MMG SKRINING"
        return "MMG"

    if modalnosc == 'TK':
        type_suffix = TK_SUFFIX_MAP.get(rodzaj, '')
        if porownawcze:
            if type_suffix:
                return f"TK PORÓWNAWCZE {priorytet} {type_suffix}"
            return f"TK PORÓWNAWCZE {priorytet}"
        if type_suffix == 'ANGIO':
            return f"TK ANGIO {priorytet}"
        if type_suffix == 'ONKO':
            return f"TK {priorytet} ONKO"
        return f"TK {priorytet}"

    if modalnosc == 'MR':
        kat = mr_get_anatomical_category(procedura)
        type_suffix = mr_get_type_suffix(rodzaj)
        parts = ['MR', priorytet, kat]
        if type_suffix == 'angio':
            parts.append('angio')
        if porownawcze:
            if type_suffix == 'ONKO':
                parts.append('ONKO')
                parts.append('PORÓW.')
            else:
                parts.append('PORÓWNAWCZE')
        else:
            if type_suffix == 'ONKO':
                parts.append('ONKO')
        return ' '.join(parts)

    return f"{modalnosc} {priorytet}"


def bill_format_excel_sheet(workbook, sheet_name, data_sections, total_row=None, grand_totals=None):
    ws = workbook[sheet_name]
    priority_colors = PRIORITY_COLORS
    procedure_type_colors = PROCEDURE_TYPE_COLORS
    light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    medium_blue_fill = PatternFill(start_color="CCEAFF", end_color="CCEAFF", fill_type="solid")

    bold_font = Font(bold=True)

    SUMMABLE_KEYWORDS = ['#', 'Mnożnik', 'Ilość', 'Wartość', 'porównawcze']
    fills = {p: PatternFill(start_color=c, end_color=c, fill_type='solid') for p, c in priority_colors.items()}
    if "accounting" not in workbook.named_styles:
        accounting_style = NamedStyle(name="accounting", number_format='_-* #,##0.00 zł_-;-* #,##0.00 zł_-;_-* "-"??_-;_-@_-')
        workbook.add_named_style(accounting_style)

    for section in data_sections:
        billing_table = section['data']
        start_row = section['start_row']

        for c_idx, col_name in enumerate(billing_table.columns, 1):
            cell = ws.cell(row=start_row, column=c_idx)
            for p, fill in fills.items():
                if str(col_name).startswith(p):
                    cell.fill = fill
                    break

        data_start_row = start_row + 1
        for row_idx, (_, row_data) in enumerate(billing_table.iterrows(), start=data_start_row):
            rodzaj_procedury = str(row_data.get('Rodzaj procedury rozlicz.', ''))
            if rodzaj_procedury in procedure_type_colors:
                row_fill = PatternFill(start_color=procedure_type_colors[rodzaj_procedury],
                                       end_color=procedure_type_colors[rodzaj_procedury],
                                       fill_type='solid')
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.fill.start_color.rgb == '00000000' or cell.fill.start_color.rgb is None:
                        cell.fill = row_fill
        data_start_row, data_end_row = start_row + 1, start_row + len(billing_table)
        sum_row = data_end_row + 1
        ws.cell(row=sum_row, column=1).value = f"SUMA {section['modalnosc']}"
        for c_idx, col_name in enumerate(billing_table.columns, 1):
            col_letter = get_column_letter(c_idx)
            is_currency = 'Stawka' in str(col_name) or 'Wartość' in str(col_name)

            if is_currency:
                for r in range(data_start_row, data_end_row + 1):
                    ws.cell(row=r, column=c_idx).style = "accounting"

            if any(k in str(col_name) for k in SUMMABLE_KEYWORDS) and pd.api.types.is_numeric_dtype(billing_table.iloc[:, c_idx - 1]):
                if 'Stawka' not in str(col_name) and 'Mnożnik' not in str(col_name):
                    sum_cell = ws.cell(row=sum_row, column=c_idx)
                    # AGGREGATE(9, 6, …) = SUMA ignorująca błędy. Jeśli pojedyncza
                    # pozycja ma pustą liczbę okolic i jej formuła daje #ARG!/błąd,
                    # suma kolumny (a przez to całość szpitala) i tak się policzy.
                    sum_cell.value = f"=AGGREGATE(9,6,{col_letter}{data_start_row}:{col_letter}{data_end_row})"
                    if is_currency:
                        sum_cell.style = "accounting"

        for c_idx in range(1, len(billing_table.columns) + 1):
            ws.cell(row=sum_row, column=c_idx).fill = light_blue_fill

    if total_row:
        for c_idx, col_name in enumerate(data_sections[0]['data'].columns, 1):
            if 'Stawka' in str(col_name) or 'Wartość' in str(col_name):
                cell = ws.cell(row=total_row, column=c_idx)
                if cell.value is not None:
                    cell.style = "accounting"
        for c_idx in range(1, len(data_sections[0]['data'].columns) + 1):
            ws.cell(row=total_row, column=c_idx).fill = medium_blue_fill

    if grand_totals:
        last_row = total_row

        value_row = last_row + 2
        ws.cell(row=value_row, column=1).value = "RAZEM WARTOŚĆ (CAŁOŚĆ):"
        ws.cell(row=value_row, column=1).font = bold_font

        value_formula_parts = []
        for c_idx, col_name in enumerate(data_sections[0]['data'].columns, 1):
            if str(col_name).endswith("Wartość"):
                col_letter = get_column_letter(c_idx)
                value_formula_parts.append(f"{col_letter}{total_row}")

        if value_formula_parts:
            cell = ws.cell(row=value_row, column=2)
            cell.value = f"={'+'.join(value_formula_parts)}"
            cell.style = "accounting"
            cell.font = bold_font
            cell.fill = medium_blue_fill

        qty_row = last_row + 3
        ws.cell(row=qty_row, column=1).value = "RAZEM ILOŚĆ OKOLIC (CAŁOŚĆ):"
        ws.cell(row=qty_row, column=1).font = bold_font

        qty_formula_parts = []
        for c_idx, col_name in enumerate(data_sections[0]['data'].columns, 1):
            if str(col_name).endswith("Ilość"):
                col_letter = get_column_letter(c_idx)
                qty_formula_parts.append(f"{col_letter}{total_row}")

        if qty_formula_parts:
            cell = ws.cell(row=qty_row, column=2)
            cell.value = f"={'+'.join(qty_formula_parts)}"
            cell.font = bold_font
            cell.fill = medium_blue_fill

    for column in ws.columns:
        max_len = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 2, 50)


def bill_make_grouped(df_details, entity_col):
    """
    Grupuje arkusz „Szczegółowe" do tabeli rozliczeniowej. Wspólne dla jednostek
    i lekarzy — różni je tylko kolumna podmiotu (entity_col: 'Klient' / 'Opisujący').
    Zwraca (grouped, df_details) — df_details ze znormalizowanymi kolumnami.
    """
    if 'Badania do porównania' not in df_details.columns:
        df_details['Badania do porównania'] = 0
    df_details['Badania do porównania'] = pd.to_numeric(df_details['Badania do porównania'], errors='coerce').fillna(0)

    grouping_columns = ['Priorytet opisu', 'Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.', entity_col]
    for col in grouping_columns:
        if pd.api.types.is_string_dtype(df_details[col]):
            df_details[col] = df_details[col].astype(str).str.strip()

    grouped = df_details.groupby(grouping_columns).agg({'Nr badania': 'count', 'Badania do porównania': 'sum'}).reset_index()
    grouped.rename(columns={'Nr badania': '#', 'Badania do porównania': 'Porownawcze_Flag'}, inplace=True)
    grouped['Mnożnik'] = grouped['Procedura rozlicz.'].apply(bill_extract_multiplier)
    grouped['Porownawcze_Flag'] = grouped['Porownawcze_Flag'] * grouped['Mnożnik']
    return grouped, df_details


def bill_finalize_to_excel(merged, df_details, output_path, logs=None):
    """
    Z gotowej tabeli (z kolumną 'Cena') tworzy plik Excel: arkusz „Szczegółowe" +
    „Rozliczenie" z podziałem na priorytety, formułami i sumami. Identyczny układ
    jak rozliczenie jednostek — używany też dla rozliczeń lekarzy (inne źródło 'Cena').
    """
    logs = logs if logs is not None else []
    merged['Ilość'] = merged['#'] * merged['Mnożnik']
    merged['Wartość'] = np.nan

    billing_table = merged.sort_values(by=['Modalność', 'Rodzaj procedury rozlicz.'])
    priorities_in_data = merged['Priorytet opisu'].unique()
    priorities_for_this_report = [p for p in MASTER_PRIORITY_ORDER if p in priorities_in_data]

    final_billing_table = billing_table.drop_duplicates(['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.'])
    final_billing_table = final_billing_table[['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.']].copy()

    for p in priorities_for_this_report:
        priority_data = billing_table[billing_table['Priorytet opisu'] == p]
        p_cols = {
            'Cena': f'{p} Stawka', '#': f'{p} #', 'Porownawcze_Flag': f'{p} w tym porównawcze',
            'Mnożnik': f'{p} Mnożnik', 'Ilość': f'{p} Ilość', 'Wartość': f'{p} Wartość'
        }
        priority_subset = priority_data[['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.'] + list(p_cols.keys())].rename(columns=p_cols)
        final_billing_table = final_billing_table.merge(priority_subset, on=['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.'], how='left')

    billing_table = final_billing_table
    ordered_cols = ['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.']
    for p in priorities_for_this_report:
        ordered_cols.extend([f'{p} Stawka', f'{p} #', f'{p} w tym porównawcze', f'{p} Mnożnik', f'{p} Ilość', f'{p} Wartość'])

    billing_table = billing_table[[c for c in ordered_cols if c in billing_table.columns]]
    num_cols = [c for c in billing_table.columns if any(k in c for k in ['Stawka', '#', 'Mnożnik', 'Ilość', 'Wartość', 'porównawcze'])]
    billing_table[num_cols] = billing_table[num_cols].fillna(0)

    df_details_modified = df_details.copy()

    def transform_comparative_studies(row):
        if pd.to_numeric(row['Badania do porównania'], errors='coerce') == 1:
            return bill_extract_multiplier(row['Procedura rozlicz.'])
        return row['Badania do porównania']

    df_details_modified['Badania do porównania'] = df_details_modified.apply(transform_comparative_studies, axis=1)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_details_modified.to_excel(writer, sheet_name='Szczegółowe', index=False)

        wb = writer.book
        ws = wb.create_sheet('Rozliczenie')
        modalnosci = sorted(billing_table['Modalność'].unique())
        data_sections, current_row = [], 1
        subtotal_rows = []

        for m in modalnosci:
            m_data = billing_table[billing_table['Modalność'] == m].copy()
            if m_data.empty:
                continue

            col_map = {name: get_column_letter(i + 1) for i, name in enumerate(m_data.columns)}
            for c_idx, c_name in enumerate(m_data.columns, 1):
                ws.cell(row=current_row, column=c_idx).value = c_name
            data_sections.append({'modalnosc': m, 'data': m_data, 'start_row': current_row})
            current_row += 1

            for r_idx, (_, row_data) in enumerate(m_data.iterrows(), start=current_row):
                for c_idx, c_name in enumerate(m_data.columns, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    col_str = str(c_name)

                    if col_str.endswith(' Mnożnik'):
                        priority_prefix = col_str.replace(' Mnożnik', '')
                        hash_col = col_map.get(f'{priority_prefix} #')
                        if hash_col:
                            cell.value = f'=IF({hash_col}{r_idx}>0,IFERROR(VALUE(LEFT($D{r_idx},1)),1),0)'
                        else:
                            cell.value = f"=IFERROR(VALUE(LEFT($D{r_idx},1)),1)"

                    elif col_str.endswith(' Ilość'):
                        priority_prefix = col_str.replace(' Ilość', '')
                        hash_col = col_map.get(f'{priority_prefix} #')
                        mult_col = col_map.get(f'{priority_prefix} Mnożnik')
                        if hash_col and mult_col:
                            cell.value = f"={hash_col}{r_idx}*{mult_col}{r_idx}"

                    elif col_str.endswith(' Wartość'):
                        priority_prefix = col_str.replace(' Wartość', '')
                        stawka_col_letter = col_map.get(f'{priority_prefix} Stawka')
                        ilosc_col_letter = col_map.get(f'{priority_prefix} Ilość')
                        if stawka_col_letter and ilosc_col_letter:
                            cell.value = f"={stawka_col_letter}{r_idx}*{ilosc_col_letter}{r_idx}"
                    else:
                        cell.value = row_data[c_name]
            current_row += len(m_data)
            subtotal_rows.append(current_row)
            current_row += 2

        total_row = current_row
        ws.cell(row=total_row, column=1).value = "SUMA CAŁKOWITA"
        SUMMABLE_KEYWORDS = ['#', 'Mnożnik', 'Ilość', 'Wartość', 'porównawcze']
        for c_idx, c_name in enumerate(billing_table.columns, 1):
            if any(k in str(c_name) for k in SUMMABLE_KEYWORDS) and 'Mnożnik' not in str(c_name):
                col_letter = get_column_letter(c_idx)
                formula_parts = [f"{col_letter}{r}" for r in subtotal_rows]
                if formula_parts:
                    ws.cell(row=total_row, column=c_idx).value = f"={'+'.join(formula_parts)}"
        bill_format_excel_sheet(wb, 'Rozliczenie', data_sections, total_row, grand_totals=True)
    return logs


def bill_process_single_file(excel_path, csv_path, output_path):
    logs = []
    try:
        logs.append(f"\n--- Rozliczam plik: {os.path.basename(excel_path)} ---")
        df_details = pd.read_excel(excel_path, sheet_name='Szczegółowe')

        df_prices = pd.read_csv(csv_path, sep=';', encoding='utf-8-sig', decimal=',')
        df_prices['Cena'] = pd.to_numeric(df_prices['Cena'], errors='coerce')
        df_prices['BADANIE'] = df_prices['BADANIE'].str.replace(r'\s+', ' ', regex=True).str.strip()

        if 'Badania do porównania' not in df_details.columns:
            df_details['Badania do porównania'] = 0
        df_details['Badania do porównania'] = pd.to_numeric(df_details['Badania do porównania'], errors='coerce').fillna(0)

        grouping_columns = ['Priorytet opisu', 'Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.', 'Klient']

        for col in grouping_columns:
            if pd.api.types.is_string_dtype(df_details[col]):
                df_details[col] = df_details[col].astype(str).str.strip()

        agg_functions = {'Nr badania': 'count', 'Badania do porównania': 'sum'}

        grouped = df_details.groupby(grouping_columns).agg(agg_functions).reset_index()
        grouped.rename(columns={'Nr badania': '#', 'Badania do porównania': 'Porownawcze_Flag'}, inplace=True)

        grouped['Mnożnik'] = grouped['Procedura rozlicz.'].apply(bill_extract_multiplier)
        grouped['Porownawcze_Flag'] = grouped['Porownawcze_Flag'] * grouped['Mnożnik']

        grouped['CENA_KLUCZ'] = grouped.apply(build_price_key, axis=1)

        merged = grouped.merge(
            df_prices[['Jednostka', 'BADANIE', 'Cena']],
            left_on=['Klient', 'CENA_KLUCZ'],
            right_on=['Jednostka', 'BADANIE'],
            how='left'
        )
        merged.drop(columns=[c for c in ['Jednostka', 'BADANIE'] if c in merged.columns], inplace=True)

        if merged['Cena'].isna().any():
            logs.append(f"! OSTRZEŻENIE: Nie znaleziono cen dla {merged['Cena'].isna().sum()} pozycji (po dwóch próbach).")

        merged['Ilość'] = merged['#'] * merged['Mnożnik']
        merged['Wartość'] = np.nan

        billing_table = merged.sort_values(by=['Modalność', 'Rodzaj procedury rozlicz.'])

        priorities_in_data = merged['Priorytet opisu'].unique()
        priorities_for_this_report = [p for p in MASTER_PRIORITY_ORDER if p in priorities_in_data]

        logs.append(f"✓ Znaleziono aktywne priorytety dla tego klienta: {priorities_for_this_report}")

        final_billing_table = billing_table.drop_duplicates(['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.'])
        final_billing_table = final_billing_table[['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.']].copy()

        for p in priorities_for_this_report:
            priority_data = billing_table[billing_table['Priorytet opisu'] == p]

            p_cols = {
                'Cena': f'{p} Stawka', '#': f'{p} #', 'Porownawcze_Flag': f'{p} w tym porównawcze',
                'Mnożnik': f'{p} Mnożnik', 'Ilość': f'{p} Ilość', 'Wartość': f'{p} Wartość'
            }

            priority_subset = priority_data[['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.'] + list(p_cols.keys())].rename(columns=p_cols)
            final_billing_table = final_billing_table.merge(priority_subset, on=['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.'], how='left')

        billing_table = final_billing_table

        ordered_cols = ['Modalność', 'Procedura', 'Rodzaj procedury rozlicz.', 'Procedura rozlicz.']
        for p in priorities_for_this_report:
            ordered_cols.extend([f'{p} Stawka', f'{p} #', f'{p} w tym porównawcze', f'{p} Mnożnik', f'{p} Ilość', f'{p} Wartość'])

        billing_table = billing_table[[c for c in ordered_cols if c in billing_table.columns]]
        num_cols = [c for c in billing_table.columns if any(k in c for k in ['Stawka', '#', 'Mnożnik', 'Ilość', 'Wartość', 'porównawcze'])]
        billing_table[num_cols] = billing_table[num_cols].fillna(0)

        df_details_modified = df_details.copy()

        def transform_comparative_studies(row):
            if pd.to_numeric(row['Badania do porównania'], errors='coerce') == 1:
                return bill_extract_multiplier(row['Procedura rozlicz.'])
            return row['Badania do porównania']

        df_details_modified['Badania do porównania'] = df_details_modified.apply(transform_comparative_studies, axis=1)
        logs.append("✓ Zaktualizowano kolumnę 'Badania do porównania' w arkuszu 'Szczegółowe'.")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_details_modified.to_excel(writer, sheet_name='Szczegółowe', index=False)

            wb = writer.book
            ws = wb.create_sheet('Rozliczenie')
            modalnosci = sorted(billing_table['Modalność'].unique())
            data_sections, current_row = [], 1
            subtotal_rows = []

            for m in modalnosci:
                m_data = billing_table[billing_table['Modalność'] == m].copy()
                if m_data.empty:
                    continue

                col_map = {name: get_column_letter(i + 1) for i, name in enumerate(m_data.columns)}
                for c_idx, c_name in enumerate(m_data.columns, 1):
                    ws.cell(row=current_row, column=c_idx).value = c_name
                data_sections.append({'modalnosc': m, 'data': m_data, 'start_row': current_row})
                current_row += 1

                for r_idx, (_, row_data) in enumerate(m_data.iterrows(), start=current_row):
                    for c_idx, c_name in enumerate(m_data.columns, 1):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        col_str = str(c_name)

                        if col_str.endswith(' Mnożnik'):
                            priority_prefix = col_str.replace(' Mnożnik', '')
                            hash_col = col_map.get(f'{priority_prefix} #')
                            if hash_col:
                                cell.value = f'=IF({hash_col}{r_idx}>0,IFERROR(VALUE(LEFT($D{r_idx},1)),1),0)'
                            else:
                                cell.value = f"=IFERROR(VALUE(LEFT($D{r_idx},1)),1)"

                        elif col_str.endswith(' Ilość'):
                            priority_prefix = col_str.replace(' Ilość', '')
                            hash_col = col_map.get(f'{priority_prefix} #')
                            mult_col = col_map.get(f'{priority_prefix} Mnożnik')
                            if hash_col and mult_col:
                                cell.value = f"={hash_col}{r_idx}*{mult_col}{r_idx}"

                        elif col_str.endswith(' Wartość'):
                            priority_prefix = col_str.replace(' Wartość', '')
                            stawka_col_letter = col_map.get(f'{priority_prefix} Stawka')
                            ilosc_col_letter = col_map.get(f'{priority_prefix} Ilość')
                            if stawka_col_letter and ilosc_col_letter:
                                cell.value = f"={stawka_col_letter}{r_idx}*{ilosc_col_letter}{r_idx}"
                        else:
                            cell.value = row_data[c_name]
                current_row += len(m_data)
                subtotal_rows.append(current_row)
                current_row += 2

            total_row = current_row
            ws.cell(row=total_row, column=1).value = "SUMA CAŁKOWITA"
            SUMMABLE_KEYWORDS = ['#', 'Mnożnik', 'Ilość', 'Wartość', 'porównawcze']
            for c_idx, c_name in enumerate(billing_table.columns, 1):
                if any(k in str(c_name) for k in SUMMABLE_KEYWORDS) and 'Mnożnik' not in str(c_name):
                    col_letter = get_column_letter(c_idx)
                    formula_parts = [f"{col_letter}{r}" for r in subtotal_rows]
                    if formula_parts:
                        ws.cell(row=total_row, column=c_idx).value = f"={'+'.join(formula_parts)}"
            bill_format_excel_sheet(wb, 'Rozliczenie', data_sections, total_row, grand_totals=True)
        logs.append(f"✓ Rozliczenie zapisane w pliku: {os.path.basename(output_path)}")
        return logs
    except Exception as e:
        logs.append(f"BŁĄD podczas rozliczania pliku {os.path.basename(excel_path)}: {e}")
        return logs


def run_billing_process(input_dir, cennik_dir, output_dir):
    try:
        csv_files = glob.glob(os.path.join(cennik_dir, "*.csv"))
        if not csv_files:
            raise FileNotFoundError(f"Nie znaleziono pliku CSV w folderze Cennik: {cennik_dir}")
        all_files = glob.glob(os.path.join(input_dir, "*.xlsx")) + glob.glob(os.path.join(input_dir, "*.xls"))
        excel_files = [f for f in all_files if not os.path.basename(f).startswith('~$')]
        if not excel_files:
            print("Nie znaleziono zweryfikowanych plików do rozliczenia.", flush=True)
            return

        tasks = []
        for excel_path in excel_files:
            safe_client_name_from_filename = Path(excel_path).stem.replace('Sprawdzony_', '')
            output_filename = f"Rozliczenie_{safe_client_name_from_filename}.xlsx"
            output_path = os.path.join(output_dir, output_filename)
            tasks.append((excel_path, csv_files[0], output_path))

        num_processes = NUM_PROCESSES_BILLING
        print(f"Rozpoczynam równoległe tworzenie {len(tasks)} rozliczeń na {num_processes} rdzeniach...", flush=True)

        # 'spawn' — patrz uwaga w run_verification (odporne na deadlock fork+wątek).
        with multiprocessing.get_context("spawn").Pool(processes=num_processes) as pool:
            results = pool.starmap(bill_process_single_file, tasks)

        for log_list in results:
            if log_list:
                print("\n".join(log_list), flush=True)

        print("Zakończono tworzenie rozliczeń.", flush=True)

    except Exception as e:
        print(f"KRYTYCZNY BŁĄD w procesie rozliczania: {e}", flush=True)


# ###################################################################################
# ### GŁÓWNA FUNKCJA ORKIESTRUJĄCA                                          ###
# ###################################################################################

def load_single_reference_file(file_path):
    for header_row in [0, 1, 2, 3]:
        try:
            df = pd.read_excel(file_path, sheet_name=0, header=header_row)
            required_keywords = ['rodzaj procedury rozlicz.', 'procedura', 'ilość okolic']
            column_text = ' '.join([str(col).lower() for col in df.columns])
            if all(keyword in column_text for keyword in required_keywords):
                return df
        except Exception:
            continue
    print(f"! OSTRZEŻENIE: Nie udało się wczytać pliku wzorcowego {os.path.basename(file_path)} - brak oczekiwanych kolumn.", flush=True)
    return None


def main(jednostki_dir, wzorcowe_dir, cennik_dir, wynik_dir, sprawdzone_dir):
    for d in [jednostki_dir, wzorcowe_dir, cennik_dir]:
        if not os.path.isdir(d):
            print(f"BŁĄD KRYTYCZNY: Folder '{os.path.basename(d)}' nie istnieje. Proszę go utworzyć.", flush=True)
            return
    os.makedirs(wynik_dir, exist_ok=True)
    os.makedirs(sprawdzone_dir, exist_ok=True)

    print("=" * 60 + "\n ŁADOWANIE I ŁĄCZENIE PLIKÓW WZORCOWYCH\n" + "=" * 60, flush=True)
    all_ref_files_raw = glob.glob(os.path.join(wzorcowe_dir, "*.xlsx")) + glob.glob(os.path.join(wzorcowe_dir, "*.xls"))
    all_ref_files = [f for f in all_ref_files_raw if not os.path.basename(f).startswith('~$')]
    if not all_ref_files:
        print(f"BŁĄD KRYTYCZNY: Nie znaleziono plików wzorcowych Excel w folderze '{os.path.basename(wzorcowe_dir)}'.", flush=True)
        return
    list_of_dfs = [df for f in all_ref_files if (df := load_single_reference_file(f)) is not None]
    if not list_of_dfs:
        print(f"BŁĄD KRYTYCZNY: Żaden z plików w folderze '{os.path.basename(wzorcowe_dir)}' nie mógł być poprawnie wczytany.", flush=True)
        return
    combined_ref_df = pd.concat(list_of_dfs, ignore_index=True)

    print("\n" + "=" * 60 + "\n ROZPOCZYNAM ETAP 1: WERYFIKACJA PLIKÓW\n" + "=" * 60, flush=True)
    try:
        verifier = MedicalVerificationAgent(combined_ref_df, jednostki_dir, sprawdzone_dir)
        verification_success = verifier.run_verification()
    except Exception as e:
        print(f"KRYTYCZNY BŁĄD podczas inicjalizacji agenta weryfikacji: {e}", flush=True)
        verification_success = False

    if not verification_success:
        print("\nEtap 1 zakończony niepowodzeniem. Przerywam dalsze przetwarzanie.", flush=True)
        return

    print("\n\n" + "=" * 60 + "\n ROZPOCZYNAM ETAP 2: TWORZENIE ROZLICZEŃ\n" + "=" * 60, flush=True)
    run_billing_process(sprawdzone_dir, cennik_dir, wynik_dir)

    print("\n\n" + "=" * 60, flush=True)
    print(" PROCES ZAKOŃCZONY.", flush=True)
    print(f"Końcowe raporty znajdują się w folderze: '{os.path.basename(wynik_dir)}'", flush=True)
    print("=" * 60, flush=True)


def run_unmatched_only(jednostki_dir, wzorcowe_dir, sprawdzone_dir):
    """Tryb szybki – tylko plik Rekordy_Bez_Wzorca.xlsx, bez tworzenia plików na klienta."""
    print("=" * 60 + "\n TRYB SZYBKI: SZUKANIE REKORDÓW BEZ WZORCA\n" + "=" * 60, flush=True)

    all_ref_files_raw = glob.glob(os.path.join(wzorcowe_dir, "*.xlsx")) + glob.glob(os.path.join(wzorcowe_dir, "*.xls"))
    all_ref_files = [f for f in all_ref_files_raw if not os.path.basename(f).startswith('~$')]
    if not all_ref_files:
        print(f"BŁĄD: Nie znaleziono plików wzorcowych w '{os.path.basename(wzorcowe_dir)}'.", flush=True)
        return

    list_of_dfs = [df for f in all_ref_files if (df := load_single_reference_file(f)) is not None]
    if not list_of_dfs:
        print("BŁĄD: Żaden plik wzorcowy nie mógł być wczytany.", flush=True)
        return

    combined_ref_df = pd.concat(list_of_dfs, ignore_index=True)

    ref_keys = combined_ref_df[['Procedura', 'Rodzaj procedury rozlicz.']].copy()
    ref_keys['Procedura'] = ref_keys['Procedura'].str.lower().str.strip()
    ref_keys['Rodzaj procedury rozlicz.'] = ref_keys['Rodzaj procedury rozlicz.'].str.lower().str.strip()
    ref_keys = ref_keys.drop_duplicates()
    ref_keys['_matched'] = True

    all_files = glob.glob(os.path.join(jednostki_dir, "*.xlsx")) + glob.glob(os.path.join(jednostki_dir, "*.xls"))
    excel_files = [f for f in all_files if not os.path.basename(f).startswith('~$')]
    if not excel_files:
        print("BŁĄD: Nie znaleziono pliku Excel w folderze 'Jednostki'.", flush=True)
        return

    print(f"Wczytuję plik: {os.path.basename(excel_files[0])}", flush=True)

    try:
        xls = pd.ExcelFile(excel_files[0])
        sheet = "Szczegółowe" if "Szczegółowe" in xls.sheet_names else xls.sheet_names[0]
        if sheet != "Szczegółowe":
            print(f"OSTRZEŻENIE: Brak arkusza 'Szczegółowe' — używam '{sheet}'.", flush=True)
        master_df = pd.read_excel(xls, sheet_name=sheet, header=0)
    except Exception as e:
        print(f"BŁĄD wczytywania pliku: {e}", flush=True)
        return

    if 'Klient' not in master_df.columns:
        print("BŁĄD: Brak kolumny 'Klient' w pliku wejściowym.", flush=True)
        return

    required_cols = ['Procedura', 'Rodzaj procedury rozlicz.']
    if not all(c in master_df.columns for c in required_cols):
        print(f"BŁĄD: Brak kolumn {required_cols} w pliku wejściowym.", flush=True)
        return

    print(f"Wczytano {len(master_df)} wierszy. Szukam braków...", flush=True)

    master_df['_proc_key'] = master_df['Procedura'].astype(str).str.lower().str.strip()
    master_df['_rodzaj_key'] = master_df['Rodzaj procedury rozlicz.'].astype(str).str.lower().str.strip()

    merged = master_df.merge(
        ref_keys,
        left_on=['_proc_key', '_rodzaj_key'],
        right_on=['Procedura', 'Rodzaj procedury rozlicz.'],
        how='left',
        suffixes=('', '_ref')
    )

    unmatched_mask = merged['_matched'].isna()
    unmatched_df = master_df[unmatched_mask.values].drop(columns=['_proc_key', '_rodzaj_key'])

    master_df.drop(columns=['_proc_key', '_rodzaj_key'], inplace=True)

    os.makedirs(sprawdzone_dir, exist_ok=True)
    output_path = os.path.join(sprawdzone_dir, "Rekordy_Bez_Wzorca.xlsx")

    if not unmatched_df.empty:
        unmatched_df.to_excel(output_path, index=False)
        print(f"\nZnaleziono {len(unmatched_df)} rekordów bez dopasowania.", flush=True)
        print(f"Zapisano: {output_path}", flush=True)
    else:
        print("\nWszystkie rekordy mają dopasowanie w plikach wzorcowych.", flush=True)

    print("=" * 60 + "\n GOTOWE.\n" + "=" * 60, flush=True)
