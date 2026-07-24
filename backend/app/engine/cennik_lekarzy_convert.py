"""
Konwerter cennika lekarzy (ZOBOWIĄZANIA LEKARZY) → prosty cennik 3-kolumnowy
(Lekarz;Kategoria;Cena).

Układ pliku źródłowego:
  * jeden skoroszyt, KAŻDA ZAKŁADKA = jeden lekarz; nazwa arkusza w formacie
    „NAZWISKO IMIĘ" (wielkimi literami). Zakładkę „ZBIORCZO ..." pomijamy.
  * w arkuszu lekarza: kolumna z etykietą kategorii + kolejna kolumna ze stawką.
    Sekcje (RTG/TK/MR) to nagłówki bez ceny; są też wiersze SUMA, ILOŚĆ OKOLIC,
    metadane (DATA ZAWARCIA…, kod umowy, FAKTURA, TERMIN…) i pozycje GOTOWOŚĆ/TRIAŻ.
  * lekarz może mieć KILKA bloków cen obok siebie (aneksy). Bierzemy NAJNOWSZY,
    czyli skrajnie prawy blok (zgodnie z ustaleniem).

Reguły:
  * dla każdego lekarza znajdujemy kolumny-etykiety (te, w których występują
    nagłówki sekcji RTG/TK/MR); najnowszy aneks = NAJWIĘKSZY indeks takiej kolumny,
    stawka jest w kolumnie bezpośrednio po niej (label+1),
  * importujemy wiersze, w których etykieta to kategoria badania, a stawka jest
    liczbą (0 zł zostaje — lekarz nie jest opłacany za tę kategorię),
  * pomijamy nagłówki sekcji, SUMA/ILOŚĆ OKOLIC, metadane i pozycje GOTOWOŚĆ/TRIAŻ
    (dyżury rozliczamy osobno — poza zakresem per-badanie),
  * kategorie spoza standardu (literówki w pliku) są raportowane jako „nietypowe".

Klucz dopasowania lekarza (do kolumny „Opisujący" w pliku miesięcznym):
  uppercase + zwinięcie spacji + posortowane tokeny (kolejność imię/nazwisko bez
  znaczenia). „Mariusz Mróz" ↔ arkusz „MRÓZ MARIUSZ" → ten sam klucz.
"""

import re
import io
import datetime

from openpyxl import load_workbook

SECTION_HEADERS = {"RTG", "TK", "MR", "MMG"}

# Wiersze, których nie traktujemy jako kategorii badań.
SKIP_EXACT = {
    "SUMA", "RTG SUMA", "TK SUMA", "MR SUMA", "MMG SUMA",
    "ILOŚĆ OKOLIC", "FAKTURA", "ANEKSY", "UWAGI DO ROZLICZEŃ:",
    "TERMIN PŁATNOŚCI", "DATA ZAWARCIA", "-", "",
}
SKIP_PREFIX = ("GOTOWOŚĆ", "GODZINA", "TRIAŻ", "DATA ZAWARCIA", "TERMIN")

# Rozpoznawanie standardowych kategorii (do raportu nietypowych/literówek).
_STD_PATTERNS = [
    re.compile(r"^RTG (CITO|PILNE|PLANOWE)( DZIECI)?$"),
    re.compile(r"^(TK|MR) (CITO|PILNE|PLANOWE) [ABCD]$"),
    re.compile(r"^MMG (SKRINING|ZWYKŁA)$"),
]


def _clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


# Łatka na drobne literówki w kategoriach cennika lekarzy (np. "PLINE" → "PILNE").
# Stosowana po obu stronach dopasowania (cennik i słownik), więc literówka po
# którejkolwiek stronie nie psuje wyceny. Łatwo rozszerzyć o kolejne pary.
_CATEGORY_TYPOS = {"PLINE": "PILNE"}
_TYPO_RE = re.compile(r"\b(" + "|".join(_CATEGORY_TYPOS) + r")\b", re.IGNORECASE)


# „MMG ZWYKŁA"/„MMG ZWYKŁE" to ta sama pozycja rozliczeniowa co goła „MMG"
# (słownik mapuje całą mammografię na kategorię „MMG"). Część lekarzy ma w cenniku
# wariant „ZWYKŁA" — sprowadzamy go do „MMG", by stawka się dopasowała.
_MMG_ZWYKLA_RE = re.compile(r"^MMG\s+ZWYK[ŁL][A-ZĄĘ]*$", re.IGNORECASE)


def fix_category_typos(cat: str) -> str:
    if not cat:
        return cat
    # Wyprostuj nazwę: zwiń wielokrotne spacje i przytnij brzegi
    # (np. „MR PLANOWE  B" → „MR PLANOWE B"), żeby dopasowanie i rozpoznawanie działało.
    cat = _clean(cat)
    cat = _TYPO_RE.sub(lambda m: _CATEGORY_TYPOS[m.group(0).upper()], cat)
    if _MMG_ZWYKLA_RE.match(cat):
        return "MMG"
    return cat


def doctor_key(name) -> str:
    """
    Klucz dopasowania lekarza — niewrażliwy na:
      • kolejność imię/nazwisko (sortujemy człony),
      • wielkość liter,
      • łączniki vs spacje w nazwiskach dwuczłonowych oraz różne rodzaje myślników
        ('Szalacha-Tarała' == 'Szalacha Tarała'; działa też po zamianie kolejności).
    """
    s = str(name or "")
    s = re.sub(r"[-‐-―−]", " ", s)  # łączniki/myślniki → spacja
    s = re.sub(r"\s+", " ", s).strip().upper()
    if not s:
        return ""
    return " ".join(sorted(s.split(" ")))


def is_standard_category(cat: str) -> bool:
    c = _clean(cat).upper()  # zwiń wielokrotne spacje przed dopasowaniem wzorca
    return any(p.match(c) for p in _STD_PATTERNS)


def _try_number(raw):
    """(cena|None, repaired, original). None = nieliczbowe (np. '-', '#DIV/0!')."""
    if raw is None:
        return None, False, ""
    if isinstance(raw, (int, float)):
        return round(float(raw), 2), False, str(raw)
    s = str(raw).strip()
    original = s
    if s == "" or s == "-" or s.startswith("#"):
        return None, False, s
    try:
        val = float(s.replace(",", ".")) if ("," in s and "." not in s) else float(s)
        return round(val, 2), False, original
    except ValueError:
        pass
    cleaned = s.replace("zł", "").replace("PLN", "").replace("\xa0", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        return round(float(cleaned), 2), True, original
    except ValueError:
        return None, False, original


def format_price(p: float) -> str:
    p = round(float(p), 2)
    if p == int(p):
        return str(int(p))
    return ("%.2f" % p).rstrip("0").rstrip(".").replace(".", ",")


def _label_cols(rows, ncols):
    """Kolumny-etykiety bloków (aneksów): te, w których ≥2 razy występuje nagłówek
    sekcji (RTG/TK/MR/MMG). Każdy taki blok ma stawkę w kolumnie label+1."""
    out = []
    for c in range(ncols):
        hits = 0
        for r in rows[:60]:
            v = r[c] if c < len(r) else None
            if v is not None and _clean(v).upper() in SECTION_HEADERS:
                hits += 1
        if hits >= 2:
            out.append(c)
    return out


def _cell_year_month(v):
    """(rok, mies) dla komórki będącej datą; inaczej None."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.year, v.month)
    return None


def _month_columns(rows):
    """Wykrywa WIERSZ nagłówków miesięcy (najwięcej komórek-dat spośród pierwszych
    wierszy) i zwraca {kolumna: (rok, mies)}. To kolumny ILOŚCI per miesiąc — po ich
    położeniu poznajemy, który blok-aneks obowiązuje w danym miesiącu."""
    best = {}
    for r in rows[:6]:
        cols = {}
        for c, v in enumerate(r):
            ym = _cell_year_month(v)
            if ym:
                cols[c] = ym
        if len(cols) > len(best):
            best = cols
    return best


def _find_block_for_period(rows, ncols, period):
    """(label_col, price_col) bloku-aneksu obowiązującego dla `period` ('YYYY-MM').

    Aneksy leżą OBOK SIEBIE chronologicznie (najnowszy skrajnie prawy). Skrajnie
    prawy bywa PRZYGOTOWANY NA PRZYSZŁOŚĆ (np. od lipca), więc dla czerwca trzeba
    wziąć wcześniejszy. Szukamy KOLUMNY MIESIĄCA = `period` (nagłówki miesięcy),
    a stawka to blok, którego etykieta jest najdalej na prawo, ale ≤ tej kolumny
    (czyli „znajdź miesiąc i przesuwaj się w lewo"). Bez `period`/nagłówków —
    skrajnie prawy blok (jak dotąd)."""
    label_cols = _label_cols(rows, ncols)
    if not label_cols:
        return None, None
    if period:
        try:
            y, m = int(str(period)[:4]), int(str(period)[5:7])
        except (ValueError, IndexError):
            y = None
        if y:
            mcols = _month_columns(rows)  # {kol: (rok, mies)}
            target_col = None
            if mcols:
                exact = [c for c, ym in mcols.items() if ym == (y, m)]
                if exact:
                    target_col = max(exact)
                else:  # brak dokładnego miesiąca → najpóźniejszy ≤ (y, m)
                    earlier = [(ym[0], ym[1], c) for c, ym in mcols.items() if ym <= (y, m)]
                    if earlier:
                        target_col = max(earlier)[2]
            if target_col is not None:
                applicable = [lc for lc in label_cols if lc <= target_col]
                if applicable:
                    lc = max(applicable)
                    return lc, lc + 1
    label_col = max(label_cols)  # fallback: najnowszy aneks (skrajnie prawy)
    return label_col, label_col + 1


def convert_workbook(path_or_bytes, period=None) -> dict:
    """
    Konwertuje cennik lekarzy. Zwraca:
      rows: [(Lekarz, Kategoria, cena_float)]  (Lekarz = nazwa arkusza)
      doctors, categories, validation

    period ('YYYY-MM', opcjonalny): wybieramy blok-aneks obowiązujący dla tego
    miesiąca (a nie skrajnie prawy — który bywa przygotowany na przyszłe okresy).
    Bez period — jak dotąd (skrajnie prawy blok).
    """
    wb = load_workbook(path_or_bytes, data_only=True, read_only=True)

    rows_out: list[tuple[str, str, float]] = []
    repaired: list[dict] = []
    nonstandard: list[dict] = []
    zeros = 0
    doctors_ok: list[str] = []
    doctors_empty: list[str] = []
    skipped_sheets: list[str] = []
    categories: set[str] = set()

    for sheet in wb.sheetnames:
        sname = _clean(sheet)
        if sname.upper().startswith("ZBIORCZO"):
            skipped_sheets.append(sheet)
            continue

        ws = wb[sheet]
        grid = list(ws.iter_rows(values_only=True))
        if not grid:
            doctors_empty.append(sname)
            continue
        ncols = max((len(r) for r in grid), default=0)

        label_col, price_col = _find_block_for_period(grid, ncols, period)
        if label_col is None:
            doctors_empty.append(sname)
            continue

        count_before = len(rows_out)
        seen_cats: set[str] = set()
        for r in grid:
            cat = _clean(r[label_col] if label_col < len(r) else None)
            if not cat:
                continue
            up = cat.upper()
            if up in SKIP_EXACT or up.startswith(SKIP_PREFIX):
                continue
            raw = r[price_col] if price_col < len(r) else None
            price, was_rep, original = _try_number(raw)
            # Nagłówek sekcji (RTG/TK/MR/MMG) pomijamy TYLKO gdy nie ma przy nim ceny.
            # Pozycja rozliczeniowa o tej samej nazwie co nagłówek — np. osobny wiersz
            # „MMG" ze stawką 40 zł, pod nagłówkiem „MMG" — JEST realną pozycją i nie
            # wolno jej zgubić (była przyczyną braku MMG w cenniku).
            if up in SECTION_HEADERS and price is None:
                continue
            if price is None:
                continue
            cat = fix_category_typos(cat)  # literówki (PLINE→PILNE) + MMG ZWYKŁA→MMG
            if cat in seen_cats:   # ten sam lekarz+kategoria w bloku — bierzemy pierwsze
                continue
            seen_cats.add(cat)
            rows_out.append((sname, cat, price))
            categories.add(cat)
            if price == 0:
                zeros += 1
            if was_rep:
                repaired.append({"lekarz": sname, "kategoria": cat, "z": original, "na": format_price(price)})
            if not is_standard_category(cat):
                nonstandard.append({"lekarz": sname, "kategoria": cat})

        if len(rows_out) > count_before:
            doctors_ok.append(sname)
        else:
            doctors_empty.append(sname)

    prices = [p for _, _, p in rows_out]
    validation = {
        "n_rows": len(rows_out),
        "n_doctors": len(doctors_ok),
        "n_categories": len(categories),
        "n_zeros": zeros,
        "n_repaired": len(repaired),
        "n_nonstandard": len(nonstandard),
        "n_doctors_empty": len(doctors_empty),
        "price_min": min(prices) if prices else 0,
        "price_max": max(prices) if prices else 0,
        "repaired": repaired[:50],
        "nonstandard": nonstandard[:80],
        "doctors_empty": doctors_empty[:50],
        "skipped_sheets": skipped_sheets,
        "categories": sorted(categories),
        "doctors": doctors_ok,
    }
    return {"rows": rows_out, "doctors": doctors_ok,
            "categories": sorted(categories), "validation": validation}


def rows_to_csv(rows: list[tuple[str, str, float]]) -> str:
    out = io.StringIO()
    out.write("Lekarz;Kategoria;Cena\n")
    for lekarz, kategoria, price in rows:
        out.write(f"{lekarz};{kategoria};{format_price(price)}\n")
    return out.getvalue()
