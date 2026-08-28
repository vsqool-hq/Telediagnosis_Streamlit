"""
Silnik rozliczenia LEKARZY (moduł niezależny od rozliczenia jednostek).

Wejście:
  * zweryfikowane dane badań (pliki „sprawdzone", arkusz „Szczegółowe") — te same,
    które powstają w Etapie 1 rozliczenia jednostek; nie uruchamiamy nic ponownie,
  * słownik (plik wzorcowy) z wypełnioną kolumną „Rodzaj procedury lekarz",
    która mapuje (Procedura, Rodzaj procedury rozlicz.) → kategoria lekarska,
  * cennik lekarzy w formacie Lekarz;Kategoria;Cena (po konwersji skoroszytu).

Wynik: rozliczenie per lekarz i kategoria (ilość, stawka, wartość) + diagnostyka
(niedopasowani lekarze, procedury bez kategorii, kategorie bez stawki).

Składanie kategorii:
  Kolumna „Rodzaj procedury lekarz" zawiera BAZĘ bez priorytetu („RTG", „TK A",
  „MR C"…), a cennik lekarzy ma priorytet w środku („RTG CITO", „TK CITO A").
  Priorytet bierzemy z kolumny „Priorytet opisu" badania i wstawiamy między
  modalność a rozmiar — patrz resolve_category().
"""

import os
import re
import glob

from app.engine.cennik_lekarzy_convert import doctor_key, fix_category_typos

LEKARZ_COL_SLOWNIK = "Rodzaj procedury lekarz"
OPISUJACY_COL = "Opisujący"

_PRIORITIES = {"CITO", "PILNE", "PLANOWE"}


def _priority_from_study(value) -> str:
    """Mapuje 'Priorytet opisu' badania na priorytet cennika lekarzy (CITO/PILNE/PLANOWE)."""
    raw = str(value or "").strip().upper()
    if not raw:
        return ""
    if "CITO" in raw or "RATUN" in raw or "UDAR" in raw:
        return "CITO"
    if "PIL" in raw:               # Pilny, Bardzo pilny
        return "PILNE"
    if "PLAN" in raw:              # Planowy
        return "PLANOWE"
    return ""


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s if s is not None else "")).strip()


def _key(s) -> str:
    return _norm(s).lower()


def load_doctor_prices(csv_path: str) -> dict:
    """Czyta Lekarz;Kategoria;Cena → {(klucz_lekarza, kategoria_norm): cena}."""
    import pandas as pd
    df = pd.read_csv(csv_path, sep=";", encoding="utf-8-sig", decimal=",")
    prices = {}
    for _, r in df.iterrows():
        kat = fix_category_typos(_norm(r["Kategoria"])).upper()  # łatka PLINE→PILNE
        prices[(doctor_key(r["Lekarz"]), kat)] = float(r["Cena"])
    return prices


def load_lekarz_categories(slownik_path: str) -> dict:
    """
    Czyta słownik → {(procedura_klucz, rodzaj_klucz): kategoria_lekarska}.
    Klucze jak w weryfikacji jednostek: Procedura i Rodzaj procedury rozlicz. (lower+strip).
    Zwraca tylko wiersze z niepustą kolumną „Rodzaj procedury lekarz".
    """
    import pandas as pd
    df = pd.read_excel(slownik_path, sheet_name="Szczegółowe")
    if LEKARZ_COL_SLOWNIK not in df.columns:
        return {}
    mapping = {}
    for _, r in df.iterrows():
        kat = _norm(r.get(LEKARZ_COL_SLOWNIK))
        if not kat or kat.lower() in ("none", "nan"):
            continue
        kat = fix_category_typos(kat)  # łatka PLINE→PILNE również po stronie słownika
        mapping[(_key(r.get("Procedura")), _key(r.get("Rodzaj procedury rozlicz.")))] = kat
    return mapping


def resolve_category(study_row, slownik_category: str) -> str:
    """
    Składa finalną kategorię cennika lekarzy z bazy ze słownika + priorytetu badania.

    Słownik podaje bazę BEZ priorytetu: "RTG", "TK A", "MR C"…
    Cennik lekarzy ma priorytet w środku: "RTG CITO", "TK CITO A", "MR PLANOWE C".
    Wstawiamy więc priorytet (z 'Priorytet opisu') między modalność a rozmiar:
      "RTG" + CITO        -> "RTG CITO"
      "TK A" + PLANOWE     -> "TK PLANOWE A"

    Jeśli baza nie jest modalnością RTG/TK/MR albo już zawiera priorytet
    (np. gotowe "MMG SKRINING"), zwracamy ją bez zmian.

    RTG DZIECI: dzieci mają w cenniku OSOBNE stawki ("RTG CITO DZIECI" itd.).
    Rozpoznajemy je po kolumnie „Rodzaj procedury rozlicz." = „RTG Dzieci"
    (niezależnie od tego, czy słownik ma „DZIECI" w bazie) i doklejamy „DZIECI".
    """
    base = _norm(slownik_category)
    if not base:
        return ""
    toks = base.upper().split(" ")
    modality = toks[0]
    if modality not in ("RTG", "TK", "MR") or any(t in _PRIORITIES for t in toks):
        return base  # gotowa kategoria / inna modalność — bez składania
    priority = _priority_from_study(study_row.get("Priorytet opisu"))
    if not priority:
        return base  # brak priorytetu → zostaw bazę (trafi do „bez stawki")
    rest = " ".join(toks[1:])  # rozmiar: A/B/C/D (lub puste dla RTG)
    cat = f"{modality} {priority} {rest}".strip()
    # RTG dzieci → stawka „…DZIECI" (tylko RTG ma taki wariant w cenniku).
    if modality == "RTG":
        rodzaj = str(study_row.get("Rodzaj procedury rozlicz.", "") or "").upper()
        if "DZIEC" in rodzaj and "DZIECI" not in cat.split():
            cat = f"{cat} DZIECI"
    return cat


_PRIORITY_LADDER = ["CITO", "PILNE", "PLANOWE"]
_SIZE_LADDER = ["D", "C", "B", "A"]   # rozmiary od najwyższego do najniższego


def _prio_order(prio: str) -> list:
    """Kolejność prób priorytetów: bieżący → niższe → wyższe (najbliższy najpierw).
    PILNE → [PILNE, PLANOWE, CITO]; CITO → [CITO, PILNE, PLANOWE]."""
    i = _PRIORITY_LADDER.index(prio)
    return [prio] + _PRIORITY_LADDER[i + 1:] + list(reversed(_PRIORITY_LADDER[:i]))


def resolve_doctor_price(prices: dict, lek_key: str, category: str):
    """
    Stawka lekarza dla kategorii z dziedziczeniem, gdy stawka jest 0/pusta:
      1) ta sama literka (rozmiar), niższy priorytet: 'TK CITO A' → 'TK PILNE A'
         → 'TK PLANOWE A';
      2) gdy dla tej literki nie ma stawki na ŻADNYM priorytecie (także wyższym),
         schodzimy po literkach od najwyższej (D → C → B → A, z pominięciem
         własnej), a w obrębie każdej literki: bieżący priorytet → niższe → wyższe.
         Np. 'TK PILNE B' → (B: PLANOWE, CITO) → 'TK PILNE D' → 'TK PLANOWE D'
         → 'TK CITO D' → literka C → literka A.
    Kategorie bez literki (RTG/MMG): tylko krok 1 (priorytety w dół, jak dotąd).
    Zwraca pierwszą stawkę > 0, a gdy brak — oryginalny wynik lookupu (None/0/NaN),
    żeby diagnostyka braków działała jak dotąd.
    """
    cat = str(category or "").strip().upper()
    if not cat:
        return None
    c = prices.get((lek_key, cat))
    if c is not None and c == c and c > 0:   # c == c odrzuca NaN
        return c
    toks = cat.split(" ")
    prio = next((p for p in _PRIORITY_LADDER if p in toks), None)
    if prio is None:
        return c
    pidx = toks.index(prio)
    size = toks[-1] if toks[-1] in _SIZE_LADDER and len(toks) > pidx + 1 else None

    def _try(p, s):
        t = toks[:pidx] + [p] + toks[pidx + 1:]
        if s is not None:
            t = t[:-1] + [s]
        cc = prices.get((lek_key, " ".join(t)))
        return cc if (cc is not None and cc == cc and cc > 0) else None

    if size is None:
        # bez literki — jak dotąd: tylko priorytety W DÓŁ
        for lower in _PRIORITY_LADDER[_PRIORITY_LADDER.index(prio) + 1:]:
            cc = _try(lower, None)
            if cc is not None:
                return cc
        return c

    # 1) własna literka: wszystkie priorytety (bieżący pominięty — już sprawdzony)
    for p in _prio_order(prio)[1:]:
        cc = _try(p, size)
        if cc is not None:
            return cc
    # 2) pozostałe literki od najwyższej (D→A), każda po wszystkich priorytetach
    for s in _SIZE_LADDER:
        if s == size:
            continue
        for p in _prio_order(prio):
            cc = _try(p, s)
            if cc is not None:
                return cc
    return c


KONSULTUJACY_COL = "Konsultujący"


def load_consult_config():
    """Konfiguracja dopłat za KONSULTACJE z ustawień, znormalizowana do doctor_key.
    Zwraca listę grup: [(konsultujacy_key, {opisujacy_key,...}, stawka_or_None)].

    stawka (opcjonalna, „zaraz pod grupą"):
      • podana  → konsultujący w PARZE z opisującym z tej grupy dostaje TĘ stawkę
        za badanie (× okolice × liczba), zamiast 50% stawki opisowej;
      • brak    → w parze liczymy 50% stawki opisowej konsultanta (dawny schemat).
    Poza grupą (opisujący spoza niej lub konsultant bez grup) — zawsze 100% stawki
    opisowej. Źródło: ustawienia z bazy (db.get_settings) — panel Ustawień."""
    from app import db
    cfg = db.get_settings()
    groups = []
    for g in cfg.get("consult_groups", []) or []:
        if not isinstance(g, dict):
            continue
        k = doctor_key(g.get("konsultujacy", ""))
        if not k:
            continue
        opis = {doctor_key(o) for o in (g.get("opisujacy") or []) if doctor_key(o)}
        if not opis:
            continue
        rate = g.get("stawka", None)
        try:
            rate = float(rate) if rate is not None and str(rate).strip() != "" else None
        except (TypeError, ValueError):
            rate = None
        if rate is not None and rate <= 0:
            rate = None
        groups.append((k, opis, rate))
    return groups


def per_study_consultations(df, cat_map, prices, groups, excluded_keys=None):
    """Jeden wiersz na KONSULTOWANE badanie z policzoną dopłatą — wspólne dla widoku
    (build_doctor_billing) i plików per lekarz (generate_doctor_billing_files), żeby
    liczyły identycznie. `groups` = wynik load_consult_config()
    (lista (kons_key, {opis_key}, stawka_or_None)). Reguła dopłaty per badanie:
      • para (konsultujący + opisujący z tej samej grupy) i grupa MA stawkę
        → stawka_grupy × okolice   (tryb „stawka grupy", pct=100%);
      • para bez stawki grupy       → 50% stawki opisowej konsultanta × okolice;
      • poza parą (opisujący spoza grup / konsultant bez grup)
        → 100% stawki opisowej konsultanta × okolice.
    Kolumny wynikowe: Modalność, Procedura, Rodzaj procedury rozlicz., Procedura
    rozlicz., Priorytet opisu, _kons_key, _kons_disp, _kategoria, _okolice, _stawka,
    _tryb, _pct, _wartosc, _src_idx (indeks wiersza w df — do doliczenia kosztu w
    porównaniu). Pomijamy: konsultację własnego opisu (kons==opis), lekarzy
    wyłączonych, badania bez kategorii oraz — gdy stawka opisowa jest potrzebna
    (para bez stawki grupy / poza parą) — brak stawki konsultanta w cenniku."""
    import pandas as pd
    from app.engine.billing import bill_extract_multiplier
    if df is None or KONSULTUJACY_COL not in getattr(df, "columns", []):
        return pd.DataFrame()
    # indeks grup per konsultant: {kk: [({opis_key}, stawka_or_None), ...]}
    groups_by_kk = {}
    for kk_g, opis_set, rate in (groups or []):
        groups_by_kk.setdefault(kk_g, []).append((opis_set, rate))
    d = df.copy()
    if "_kategoria" not in d.columns:
        d["_proc_key"] = d["Procedura"].map(_key)
        d["_rodzaj_key"] = d["Rodzaj procedury rozlicz."].map(_key)
        d["_kategoria"] = [
            resolve_category(r, cat_map.get((r["_proc_key"], r["_rodzaj_key"]), ""))
            for _, r in d.iterrows()
        ]
    if "_lek_key" not in d.columns:
        d["_lek_key"] = d[OPISUJACY_COL].map(doctor_key)
    d["_lek_disp"] = d[OPISUJACY_COL].map(_norm)   # nazwa lekarza OPISUJĄCEGO (do raportu konsultacji)
    d["_kons_key"] = d[KONSULTUJACY_COL].map(doctor_key)
    d["_kons_disp"] = d[KONSULTUJACY_COL].map(_norm)
    d = d[(d["_kons_key"] != "") & (d["_kons_key"] != d["_lek_key"]) & (d["_kategoria"] != "")]
    if excluded_keys:
        d = d[~d["_kons_key"].isin(set(excluded_keys))]
    if d.empty:
        return pd.DataFrame()
    d["_okolice"] = d["Procedura rozlicz."].map(bill_extract_multiplier)
    keep = ["Modalność", "Procedura", "Rodzaj procedury rozlicz.", "Procedura rozlicz.", "Priorytet opisu"]
    recs = []
    for idx, r in d.iterrows():
        kk, okc, kat, opis = r["_kons_key"], int(r["_okolice"]), r["_kategoria"], r["_lek_key"]
        # w parze? jeśli tak — jaka stawka grupy (może być None)?
        in_group, group_rate = False, None
        for opis_set, rate in groups_by_kk.get(kk, []):
            if opis in opis_set:
                in_group, group_rate = True, rate
                break
        if in_group and group_rate is not None:
            stawka, pct, tryb = group_rate, 1.0, "stawka grupy"
            val = stawka * okc
        else:
            stawka = resolve_doctor_price(prices, kk, kat)
            if stawka is None or stawka != stawka:   # brak stawki opisowej w cenniku
                continue
            pct = 0.5 if in_group else 1.0
            tryb = f"{int(pct * 100)}%"
            val = stawka * okc * pct
        rec = {c: r[c] for c in keep}
        rec.update({"_kons_key": kk, "_kons_disp": r["_kons_disp"], "_lek_disp": r["_lek_disp"],
                    "_kategoria": kat, "_okolice": okc, "_stawka": float(stawka), "_tryb": tryb,
                    "_pct": pct, "_wartosc": round(val, 2), "_src_idx": idx})
        recs.append(rec)
    return pd.DataFrame(recs)


def read_verified_studies(sprawdzone_dir: str):
    """Wczytuje i łączy arkusze „Szczegółowe" ze wszystkich plików sprawdzonych."""
    import pandas as pd
    frames = []
    for path in sorted(glob.glob(os.path.join(sprawdzone_dir, "*.xlsx"))):
        if os.path.basename(path).startswith("~$"):
            continue
        try:
            frames.append(pd.read_excel(path, sheet_name="Szczegółowe"))
        except Exception:  # noqa: BLE001
            continue
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)


def build_doctor_billing(sprawdzone_dir: str, slownik_path: str, doctor_cennik_csv: str,
                         excluded_keys=None) -> dict:
    """
    Liczy rozliczenie lekarzy. Zwraca:
      rows: [{lekarz, kategoria, ilosc, stawka, wartosc}]
      by_doctor: [{lekarz, ilosc, wartosc}]
      validation: diagnostyka niedopasowań

    excluded_keys: zbiór kluczy lekarzy (doctor_key) do POMINIĘCIA w rozliczeniu —
    np. lekarze rozliczani osobno. Ich badania nie są wyceniane ani zgłaszane jako
    braki; pokazujemy tylko ile pominięto.
    """
    import pandas as pd
    excluded_keys = set(excluded_keys or [])

    df = read_verified_studies(sprawdzone_dir)
    if df is None or df.empty:
        return {"empty": True, "reason": "Brak zweryfikowanych danych (plików sprawdzonych)."}
    if OPISUJACY_COL not in df.columns:
        return {"empty": True, "reason": f"Brak kolumny '{OPISUJACY_COL}' w danych."}

    # Częściowo (a nawet w ogóle nie) wypełniona kolumna „Rodzaj procedury lekarz"
    # NIE blokuje liczenia — badania bez przypisanej kategorii są pomijane i
    # raportowane (studies_without_category). Pełna pustka = po prostu 0 wycenionych.
    cat_map = load_lekarz_categories(slownik_path)

    prices = load_doctor_prices(doctor_cennik_csv)

    # mapowanie badanie → kategoria lekarska
    df = df.copy()
    # LEKARZE: brak jakiegokolwiek podciągania (w górę/dół). Używamy ORYGINALNego
    # rodzaju procedury i liczby okolic (kolumny zachowane w Etapie 1 PRZED korektami
    # jednostek), a kategorię dobieramy po PARZE (opis procedury, rodzaj procedury) —
    # ten sam opis przy różnych rodzajach bywa w innej kategorii cenowej lekarza.
    # Spójne z generate_doctor_billing_files (raporty per lekarz).
    if "Rodzaj procedury rozlicz. (oryg.)" in df.columns:
        df["Rodzaj procedury rozlicz."] = df["Rodzaj procedury rozlicz. (oryg.)"]
    if "Procedura rozlicz. (oryg.)" in df.columns:
        df["Procedura rozlicz."] = df["Procedura rozlicz. (oryg.)"]
    if "Priorytet opisu" in df.columns:
        df["Priorytet opisu"] = df["Priorytet opisu"].replace({"Bardzo pilny": "Pilny"})
    df["_proc_key"] = df["Procedura"].map(_key)
    df["_rodzaj_key"] = df["Rodzaj procedury rozlicz."].map(_key)
    df["_kategoria"] = [
        resolve_category(r, cat_map.get((r["_proc_key"], r["_rodzaj_key"]), ""))
        for _, r in df.iterrows()
    ]
    df["_lek_key"] = df[OPISUJACY_COL].map(doctor_key)
    df["_lek_disp"] = df[OPISUJACY_COL].map(_norm)

    # Wyłączeni lekarze (z ustawień) — pomijamy ich badania w rozliczeniu.
    excluded_studies = int(df["_lek_key"].isin(excluded_keys).sum()) if excluded_keys else 0
    if excluded_keys:
        df = df[~df["_lek_key"].isin(excluded_keys)].copy()

    # diagnostyka
    no_category = df[df["_kategoria"] == ""]
    studies_no_cat = int(len(no_category))
    # Rozbicie „bez kategorii" na pary (Procedura, Rodzaj procedury rozlicz.) — to
    # DOKŁADNIE te wiersze słownika, którym brakuje wpisu w kolumnie „Rodzaj procedury
    # lekarz". Lekarz też, żeby było wiadomo kogo dotyczy. Sortujemy malejąco po liczbie.
    _nc_cols = [c for c in ["Modalność", "Procedura", "Rodzaj procedury rozlicz."] if c in no_category.columns]
    if _nc_cols and not no_category.empty:
        no_cat_pairs = (
            no_category.groupby(_nc_cols).size().reset_index(name="n")
            .rename(columns={"Modalność": "modalnosc", "Procedura": "procedura",
                             "Rodzaj procedury rozlicz.": "rodzaj"})
            .sort_values("n", ascending=False)
        )
        no_cat_records = no_cat_pairs.head(300).to_dict("records")
    else:
        no_cat_records = []

    priced = df[df["_kategoria"] != ""].copy()
    priced["_stawka"] = [
        resolve_doctor_price(prices, lk, kat)
        for lk, kat in zip(priced["_lek_key"], priced["_kategoria"])
    ]

    missing_price = priced[priced["_stawka"].isna()]
    pairs_no_price = (
        missing_price.groupby(["_lek_disp", "_kategoria"]).size().reset_index(name="n")
        .sort_values("n", ascending=False)
    )

    # Badania rozliczone po stawce 0 zł — pozycja W CENNIKU istnieje, ale stawka = 0
    # (lekarz nieopłacany za tę kategorię). Osobna lista od „braku stawki" (NaN wyżej).
    zero_rate = priced[pd.to_numeric(priced["_stawka"], errors="coerce") == 0]
    zero_rate_pairs = (
        zero_rate.groupby(["_lek_disp", "_kategoria"]).size().reset_index(name="n")
        .rename(columns={"_lek_disp": "lekarz", "_kategoria": "kategoria"})
        .sort_values("n", ascending=False)
    )

    doctors_in_data = set(df["_lek_key"]) - {""}
    doctors_in_cennik = {k for (k, _) in prices.keys()}
    doctors_unmatched = sorted(
        df[df["_lek_key"].isin(doctors_in_data - doctors_in_cennik)]["_lek_disp"].unique().tolist()
    )

    from app.engine.billing import bill_extract_multiplier
    ok = priced[priced["_stawka"].notna()].copy()
    ok["ilosc"] = 1  # licznik badań (do kolumny „Badania")
    # Wartość = stawka × liczba okolic (jak u jednostek) — okolice z „Procedura rozlicz.".
    ok["_okolice"] = ok["Procedura rozlicz."].map(bill_extract_multiplier)
    ok["wartosc"] = ok["_okolice"] * ok["_stawka"]

    by_cat = (
        ok.groupby(["_lek_disp", "_kategoria"])
        .agg(ilosc=("ilosc", "sum"), stawka=("_stawka", "first"), wartosc=("wartosc", "sum"))
        .reset_index().rename(columns={"_lek_disp": "lekarz", "_kategoria": "kategoria"})
        .sort_values(["lekarz", "kategoria"])
    )
    by_doctor = (
        ok.groupby("_lek_disp").agg(ilosc=("ilosc", "sum"), wartosc=("wartosc", "sum"))
        .reset_index().rename(columns={"_lek_disp": "lekarz"})
        .sort_values("wartosc", ascending=False)
    )

    # ---- KONSULTACJE: dopłata dla lekarza w roli „Konsultujący" (dodatkowa) --------
    # Dla każdego badania z niepustym konsultującym doliczamy JEMU (opisujący bez zmian):
    #   • para z grupy + stawka grupy → stawka_grupy × okolice,
    #   • para z grupy bez stawki     → 50% stawki opisowej konsultanta × okolice,
    #   • poza grupą                  → 100% stawki opisowej konsultanta × okolice.
    # Konsultacja własnego opisu (ten sam lekarz) pomijana.
    consult_groups = load_consult_config()
    cons_df = per_study_consultations(df, cat_map, prices, consult_groups, excluded_keys)
    consult_detail = []
    consult_by_key = {}
    if not cons_df.empty:
        for _, r in cons_df.iterrows():
            acc = consult_by_key.setdefault(r["_kons_key"],
                                            {"lekarz": r["_kons_disp"], "ilosc": 0, "okolice": 0, "wartosc": 0.0})
            acc["ilosc"] += 1
            acc["okolice"] += int(r["_okolice"])
            acc["wartosc"] += float(r["_wartosc"])
            consult_detail.append({"konsultujacy": r["_kons_disp"], "kategoria": r["_kategoria"],
                                   "tryb": r["_tryb"], "okolice": int(r["_okolice"]),
                                   "wartosc": round(float(r["_wartosc"]), 2)})
    consultations = [
        {"lekarz": v["lekarz"], "ilosc": int(v["ilosc"]), "okolice": int(v["okolice"]),
         "wartosc": round(v["wartosc"], 2)}
        for v in sorted(consult_by_key.values(), key=lambda x: -x["wartosc"])
    ]
    consult_total = round(sum(v["wartosc"] for v in consult_by_key.values()), 2)

    # Doklejamy konsultacje do sumy per lekarz (by_doctor): dopasowanie po doctor_key.
    bd = by_doctor.to_dict("records")
    by_key_idx = {doctor_key(row["lekarz"]): row for row in bd}
    for row in bd:
        row["wartosc_opis"] = round(float(row["wartosc"]), 2)
        row["wartosc_konsultacje"] = 0.0
    for kk, v in consult_by_key.items():
        row = by_key_idx.get(kk)
        if row is None:
            row = {"lekarz": v["lekarz"], "ilosc": 0, "wartosc": 0.0,
                   "wartosc_opis": 0.0, "wartosc_konsultacje": 0.0}
            bd.append(row); by_key_idx[kk] = row
        row["wartosc_konsultacje"] = round(row["wartosc_konsultacje"] + v["wartosc"], 2)
        row["wartosc"] = round(float(row["wartosc"]) + v["wartosc"], 2)
    bd.sort(key=lambda r: -float(r["wartosc"]))
    by_doctor_records = bd

    # Podsumowanie ilościowe: liczba WYKONANYCH badań per lekarz i kategoria
    # (kategoria w formacie cennika lekarzy). Liczymy ze WSZYSTKICH badań mających
    # kategorię — niezależnie od tego, czy cennik ma dla niej stawkę (to licznik
    # badań, nie wartość). Badania bez kategorii są raportowane osobno
    # (studies_without_category) i tu się nie pojawiają.
    cat_counts = (
        priced.groupby(["_lek_disp", "_kategoria"]).size()
        .reset_index(name="ilosc")
        .rename(columns={"_lek_disp": "lekarz", "_kategoria": "kategoria"})
        .sort_values(["lekarz", "kategoria"])
    )

    # Liczba OKOLIC per lekarz i kategoria (count × mnożnik z „Procedura rozlicz.") —
    # do uzupełniania pliku zobowiązań lekarzy (tam wpisuje się okolice, nie sztuki).
    priced["_okolice"] = priced["Procedura rozlicz."].map(bill_extract_multiplier)
    cat_okolice = (
        priced.groupby(["_lek_disp", "_kategoria"])["_okolice"].sum()
        .reset_index()
        .rename(columns={"_lek_disp": "lekarz", "_kategoria": "kategoria", "_okolice": "okolice"})
        .sort_values(["lekarz", "kategoria"])
    )

    # Rozbicie okolic po DACIE BADANIA (dzień) — potrzebne tylko dla miesięcy
    # rozbitych aneksem w połowie (w pliku zobowiązań są wtedy dwie kolumny:
    # np. „01-17.05" i „18-31.05" z różnymi stawkami). Lekka agregacja
    # lekarz × kategoria × data → suma okolic; dla zwykłych miesięcy nieużywana
    # (wpisujemy sumę z category_okolice). Data: data badania (zapas: zatwierdzenia).
    cat_okolice_daily = []
    _date_col = next((c for c in ("Data badania (UTC)", "Data 1. zatwierdzenia")
                      if c in priced.columns), None)
    if _date_col is not None and not priced.empty:
        _pr = priced[["_lek_disp", "_kategoria", "_okolice", _date_col]].copy()
        _pr["_data"] = pd.to_datetime(_pr[_date_col], errors="coerce")
        _pr = _pr[_pr["_data"].notna()]
        if not _pr.empty:
            _pr["_data"] = _pr["_data"].dt.strftime("%Y-%m-%d")
            cat_okolice_daily = (
                _pr.groupby(["_lek_disp", "_kategoria", "_data"])["_okolice"].sum()
                .reset_index()
                .rename(columns={"_lek_disp": "lekarz", "_kategoria": "kategoria",
                                 "_data": "data", "_okolice": "okolice"})
                .to_dict("records")
            )

    # --- KONSULTACJE 100% → ILOŚCI do pliku ZOBOWIĄZAŃ (jak własne opisy) ----------
    # Konsultacje rozliczane PO 100% (poza parą: konsultant dostaje 100% SWOJEJ stawki
    # opisowej — jak gdyby sam opisał to badanie) doliczamy KONSULTANTOWI do okolic per
    # kategoria. Dzięki temu w pliku zobowiązań (wiersz okolic × stawka opisowa) wchodzą
    # tak samo jak opisy. Trybów „50%" i „stawka grupy" NIE dokładamy — tam stawka jest
    # inna niż opisowa, a wiersz zobowiązań liczy po opisowej, więc tylko 100% wyceni się
    # w nim poprawnie. Ilości sczytujemy zawsze (raport Konsultacje.xlsx niżej).
    cat_okolice_records = cat_okolice.to_dict("records")
    if not cons_df.empty:
        c100 = cons_df[cons_df["_tryb"] == "100%"]
        if not c100.empty:
            merged_ok = {(r["lekarz"], r["kategoria"]): dict(r) for r in cat_okolice_records}
            add = c100.groupby(["_kons_disp", "_kategoria"])["_okolice"].sum().reset_index()
            for _, r in add.iterrows():
                key = (r["_kons_disp"], r["_kategoria"])
                if key in merged_ok:
                    merged_ok[key]["okolice"] = int(merged_ok[key]["okolice"]) + int(r["_okolice"])
                else:
                    merged_ok[key] = {"lekarz": r["_kons_disp"], "kategoria": r["_kategoria"],
                                      "okolice": int(r["_okolice"])}
            cat_okolice_records = sorted(merged_ok.values(), key=lambda x: (x["lekarz"], x["kategoria"]))
            # rozbicie dzienne (dla miesięcy rozbitych aneksem) — data badania z df po _src_idx
            if _date_col is not None:
                for _, cr in c100.iterrows():
                    try:
                        d = pd.to_datetime(df.loc[cr["_src_idx"], _date_col], errors="coerce")
                    except (KeyError, TypeError):
                        d = None
                    if d is not None and pd.notna(d):
                        cat_okolice_daily.append({
                            "lekarz": cr["_kons_disp"], "kategoria": cr["_kategoria"],
                            "data": d.strftime("%Y-%m-%d"), "okolice": int(cr["_okolice"])})

    # --- Raport KONSULTACJE (wszystkie rozliczone konsultacje, wiersz = badanie) -----
    # Kto konsultował, komu (opisujący), po jakiej stawce, jakiego badania i priorytetu —
    # do pliku Konsultacje.xlsx w paczce. „Stawka za konsultację" = stawka × % (kwota
    # płacona konsultantowi za okolicę); Wartość = × okolice (żeby łatwo zsumować „ile
    # trafiło do lekarzy z konsultacji").
    consultations_report = []
    if not cons_df.empty:
        _rep = cons_df.copy()
        _rep["_rate_eff"] = (_rep["_stawka"] * _rep["_pct"]).round(2)
        _rep = _rep.sort_values(["_kons_disp", "Modalność", "Rodzaj procedury rozlicz.", "Priorytet opisu"])
        for _, r in _rep.iterrows():
            consultations_report.append({
                "Lekarz konsultujący": r["_kons_disp"],
                "Lekarz opisujący": r.get("_lek_disp", ""),
                "Stawka za konsultację": float(r["_rate_eff"]),
                "Tryb": r["_tryb"],
                "Rodzaj procedury rozlicz.": r["Rodzaj procedury rozlicz."],
                "Opis badania": r["Procedura"],
                "Priorytet": r["Priorytet opisu"],
                "Okolice": int(r["_okolice"]),
                "Wartość": round(float(r["_wartosc"]), 2),
            })

    value_opis = round(float(ok["wartosc"].sum()), 2)
    return {
        "empty": False,
        "rows": by_cat.to_dict("records"),
        "by_doctor": by_doctor_records,
        "consultations": consultations,
        "consult_detail": consult_detail,
        "consultations_report": consultations_report,
        "category_counts": cat_counts.to_dict("records"),
        "category_okolice": cat_okolice_records,
        "category_okolice_daily": cat_okolice_daily,
        "validation": {
            "total_studies": int(len(df)),
            "priced_studies": int(len(ok)),
            "studies_without_category": studies_no_cat,
            "categories_missing": no_cat_records,
            "slownik_categories": len(cat_map),
            "n_doctors": len(by_doctor_records),
            # total_value = opisy + konsultacje (gotowość dokłada doctors_job osobno)
            "total_value": round(value_opis + consult_total, 2),
            "value_opis": value_opis,
            "value_consultations": consult_total,
            "excluded_studies": excluded_studies,
            "doctors_unmatched": doctors_unmatched[:100],
            "pairs_without_price": pairs_no_price.head(100).to_dict("records"),
            "zero_rate_studies": int(len(zero_rate)),
            "zero_rate_pairs": zero_rate_pairs.head(200).to_dict("records"),
        },
    }


def _doctor_row_color(cat_map):
    """Funkcja koloru wiersza w pliku LEKARZA: (modalnosc, procedura, rodzaj) → HEX|None.
    Kolor wg GRUPY lekarskiej A/B/C/D w TK/MR (litera z bazy kategorii ze słownika,
    np. „TK A"→A); „RTG Dzieci" ma własny kolor. Grupa B oraz kategorie bez litery
    (RTG dorośli, MMG…) — bez wypełnienia. Jednostki mają własne kolory (bez tej funkcji)."""
    TK = {"A": "C0E6F5", "C": "FBE2D5", "D": "E8E8E8"}   # B → brak wypełnienia
    MR = {"A": "A6C9EC", "C": "F7C7AC", "D": "FFE699"}   # B → brak wypełnienia

    def color(modalnosc, procedura, rodzaj):
        r = _norm(rodzaj).upper()
        if r.startswith("RTG") and "DZIEC" in r:
            return "FFE5FF"
        base = _norm(cat_map.get((_key(procedura), _key(rodzaj)), "")).upper()
        toks = base.split()
        if len(toks) >= 2 and toks[-1] in ("A", "B", "C", "D"):
            table = TK if toks[0] == "TK" else (MR if toks[0] == "MR" else None)
            if table is not None:
                return table.get(toks[-1])
        return None
    return color


def _doctor_cat_sortkey(cat_map):
    """Klucz sortowania wiersza pliku lekarza W OBRĘBIE modalności: baza kategorii
    lekarskiej ze słownika (np. „MR A"). Dzięki temu wiersze tej samej kategorii
    (=tego samego koloru) są RAZEM i w kolejności A→B→C→D, a nie porozrzucane po
    „Rodzaj procedury rozlicz.". Bez kategorii → na koniec bloku modalności."""
    def key(modalnosc, procedura, rodzaj):
        base = _norm(cat_map.get((_key(procedura), _key(rodzaj)), "")).upper()
        return base or "ZZZ"
    return key


def _surname_first(name: str) -> str:
    """„Imię Nazwisko" → „Nazwisko Imię" (np. 'Tomasz Bujalski' → 'Bujalski Tomasz')."""
    toks = _norm(name).split()
    return " ".join(toks[1:] + toks[:1]) if len(toks) >= 2 else _norm(name)


def _safe_filename(s: str) -> str:
    """Usuwa znaki niedozwolone w nazwach plików; zachowuje spacje i myślniki."""
    return re.sub(r'[\\/:*?"<>|]+', "", s).strip()


def _period_mmyyyy(frame) -> str:
    """Okres rozliczenia jako MMRRRR — z daty zatwierdzenia opisu (fallback: data badania)."""
    import pandas as pd
    for col in ("Data 1. zatwierdzenia", "Data badania (UTC)"):
        if col in frame.columns:
            dt = pd.to_datetime(frame[col], errors="coerce").dropna()
            if not dt.empty:
                return dt.dt.strftime("%m%Y").mode().iloc[0]
    return ""


def _aggregate_consultations(sub_c):
    """Z konsultacji per-badanie (per_study_consultations) → wiersze ZAGREGOWANE do
    wpięcia w tabelę „Rozliczenie" pliku lekarza: liczba konsultacji per
    (Modalność, Procedura, Rodzaj procedury rozlicz., Procedura rozlicz., Priorytet
    opisu, stawka, %). Kolumny wyniku: te 5 kluczy + Stawka, Konsultacje (liczba), _pct.
    Rozbicie po _pct daje osobne wiersze 50% / 100% / ryczałt (stawka konsultanta)."""
    import pandas as pd
    if sub_c is None or getattr(sub_c, "empty", True):
        return pd.DataFrame()
    g = (sub_c.groupby(["Modalność", "Procedura", "Rodzaj procedury rozlicz.",
                        "Procedura rozlicz.", "Priorytet opisu", "_stawka", "_pct"], dropna=False)
         .size().reset_index(name="Konsultacje"))
    return g.rename(columns={"_stawka": "Stawka"})


def write_consultations_report(rows: list, out_path: str) -> int:
    """Zapisuje raport KONSULTACJE (Konsultacje.xlsx) — jeden wiersz na rozliczoną
    konsultację. Kolumny: Lekarz konsultujący, Lekarz opisujący, Stawka za konsultację,
    Tryb, Rodzaj procedury rozlicz., Opis badania, Priorytet, Okolice, Wartość + wiersz
    RAZEM (suma okolic i wartości). Zwraca liczbę pozycji. `rows` = result['consultations_report']."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = ["Lekarz konsultujący", "Lekarz opisujący", "Stawka za konsultację", "Tryb",
            "Rodzaj procedury rozlicz.", "Opis badania", "Priorytet", "Okolice", "Wartość"]
    money = '_-* #,##0.00 zł_-;-* #,##0.00 zł_-;_-* "-"??_-;_-@_-'
    bold = Font(bold=True)
    head_fill = PatternFill(start_color="CCEAFF", end_color="CCEAFF", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Konsultacje"
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = bold
        cell.fill = head_fill
    r = 2
    for row in rows or []:
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=r, column=j, value=row.get(c))
            if c in ("Stawka za konsultację", "Wartość"):
                cell.number_format = money
        r += 1
    if rows:
        ws.cell(row=r, column=1, value="RAZEM").font = bold
        for name in ("Okolice", "Wartość"):
            col = cols.index(name) + 1
            L = get_column_letter(col)
            c = ws.cell(row=r, column=col, value=f"=SUBTOTAL(9,{L}2:{L}{r-1})")
            c.font = bold
            if name == "Wartość":
                c.number_format = money
    for j, c in enumerate(cols, 1):
        maxlen = max([len(str(c))] + [len(str((row or {}).get(c, ""))) for row in (rows or [])[:800]])
        ws.column_dimensions[get_column_letter(j)].width = min(maxlen + 2, 46)
    wb.save(out_path)
    return len(rows or [])


def generate_doctor_billing_files(sprawdzone_dir: str, slownik_path: str, doctor_cennik_csv: str,
                                  out_dir: str, excluded_keys=None, period_mmyyyy=None,
                                  availability=None) -> dict:
    """
    Tworzy OSOBNY plik Excel dla KAŻDEGO lekarza — w układzie identycznym jak
    rozliczenia jednostek (arkusz „Szczegółowe" z jego badaniami + „Rozliczenie"
    z podziałem na priorytety, liczbą okolic i sumami), ale wyceniony cennikiem
    lekarzy (stawka per Lekarz+Kategoria; Wartość = stawka × ilość_okolic).

    Braki kategorii/stawki → stawka 0 (jak braki cen u jednostek). Wyłączeni
    lekarze pomijani. Zwraca listę utworzonych plików.
    """
    import os as _os
    import shutil
    import numpy as np
    from app.engine.billing import bill_make_grouped, bill_finalize_to_excel
    from app.engine.formula_eval import evaluate_workbook_to_values

    excluded = set(excluded_keys or [])
    df = read_verified_studies(sprawdzone_dir)
    if df is None or df.empty or OPISUJACY_COL not in df.columns:
        return {"files": [], "count": 0}

    cat_map = load_lekarz_categories(slownik_path)
    prices = load_doctor_prices(doctor_cennik_csv)
    row_color_fn = _doctor_row_color(cat_map)   # kolory wierszy wg grupy lekarskiej
    row_sort_fn = _doctor_cat_sortkey(cat_map)  # sortowanie wg kategorii (koloru) w modalności

    import pandas as pd
    df = df.copy()
    df["_lek_key"] = df[OPISUJACY_COL].map(doctor_key)
    if excluded:
        df = df[~df["_lek_key"].isin(excluded)]
    # Oryginalny rodzaj/okolice + „Bardzo pilny"→„Pilny" (spójnie z build_doctor_billing),
    # żeby kategoria konsultacji i okolice liczyły się tak samo jak w widoku.
    if "Rodzaj procedury rozlicz. (oryg.)" in df.columns:
        df["Rodzaj procedury rozlicz."] = df["Rodzaj procedury rozlicz. (oryg.)"]
    if "Procedura rozlicz. (oryg.)" in df.columns:
        df["Procedura rozlicz."] = df["Procedura rozlicz. (oryg.)"]
    if "Priorytet opisu" in df.columns:
        df["Priorytet opisu"] = df["Priorytet opisu"].replace({"Bardzo pilny": "Pilny"})
    df["_kons_key"] = (df[KONSULTUJACY_COL].map(doctor_key)
                       if KONSULTUJACY_COL in df.columns else pd.Series([""] * len(df), index=df.index))

    # Konsultacje — per badanie (wspólny helper, jak widok), pogrupowane per konsultant.
    consult_groups = load_consult_config()
    cons_df = per_study_consultations(df, cat_map, prices, consult_groups, excluded)
    cons_by_kons, consult_disp = {}, {}
    if not cons_df.empty:
        for kk, sub_c in cons_df.groupby("_kons_key"):
            if kk in excluded:
                continue
            cons_by_kons[kk] = sub_c
            consult_disp[kk] = sub_c["_kons_disp"].iloc[0]

    def _raw_consulted(kk):
        return df[(df["_kons_key"] == kk) & (df["_kons_key"] != df["_lek_key"])]

    # świeży katalog wyjściowy z DWIEMA wersjami każdego pliku:
    #   formuły/  — pliki z formułami (jak dotąd),
    #   wartości/ — te same pliki z formułami PRZELICZONYMI na liczby (do dalszej obróbki).
    # Zobowiązania i Konsultacje.xlsx dokłada doctors_job LUZEM (na górze paczki).
    shutil.rmtree(out_dir, ignore_errors=True)
    formuly_dir = _os.path.join(out_dir, "formuły")
    wartosci_dir = _os.path.join(out_dir, "wartości")
    _os.makedirs(formuly_dir, exist_ok=True)
    _os.makedirs(wartosci_dir, exist_ok=True)

    def _write_pair(fn, writer):
        """writer(path) tworzy plik z formułami w formuły/<fn>; my dorabiamy wartości/<fn>."""
        fpath = _os.path.join(formuly_dir, fn)
        writer(fpath)
        try:
            evaluate_workbook_to_values(fpath, _os.path.join(wartosci_dir, fn))
        except Exception as e:  # noqa: BLE001
            print(f'! Wersja „wartości" dla {fn}: {e}', flush=True)

    files = []
    described_keys = set()
    for lek_key, sub in df.groupby("_lek_key"):
        if not lek_key:
            continue
        disp = _norm(sub[OPISUJACY_COL].iloc[0])
        if not disp or disp.lower() in ("nan", "none"):
            continue  # puste/nieznane nazwisko — nie twórz pliku „nan"
        sub = sub.copy()
        # LEKARZE: brak jakiegokolwiek podciągania — przywróć ORYGINALNY rodzaj
        # procedury i liczbę okolic (kolumny zachowane w Etapie 1 w
        # billing.process_client_data, przed korektami priorytetów/okolic).
        if "Rodzaj procedury rozlicz. (oryg.)" in sub.columns:
            sub["Rodzaj procedury rozlicz."] = sub["Rodzaj procedury rozlicz. (oryg.)"]
        if "Procedura rozlicz. (oryg.)" in sub.columns:
            sub["Procedura rozlicz."] = sub["Procedura rozlicz. (oryg.)"]
        # „Bardzo pilny" rozliczamy razem z „Pilny" (ta sama stawka, jeden blok).
        if "Priorytet opisu" in sub.columns:
            sub["Priorytet opisu"] = sub["Priorytet opisu"].replace({"Bardzo pilny": "Pilny"})
        grouped, det = bill_make_grouped(sub, OPISUJACY_COL)

        def _cena(r):
            base = cat_map.get((_key(r["Procedura"]), _key(r["Rodzaj procedury rozlicz."])), "")
            category = resolve_category(r, base)
            if not category:
                return np.nan
            c = resolve_doctor_price(prices, lek_key, category)   # 0/brak → niższy priorytet
            return np.nan if c is None else c

        grouped["Cena"] = grouped.apply(_cena, axis=1)

        # Stawka dla DOWOLNEGO priorytetu danej procedury (także bez badań w tym
        # priorytecie) — żeby w raporcie stawka była wszędzie, nawet gdy ilość=0.
        def _rate(procedura, rodzaj, priorytet, _lk=lek_key):
            base = cat_map.get((_key(procedura), _key(rodzaj)), "")
            category = resolve_category(
                {"Priorytet opisu": priorytet, "Rodzaj procedury rozlicz.": rodzaj}, base)
            if not category:
                return np.nan
            c = resolve_doctor_price(prices, _lk, category)       # 0/brak → niższy priorytet
            return np.nan if c is None else c

        # Nazwa pliku wg szablonu: „MMRRRR dr Nazwisko Imię" (np. „052026 dr Bujalski Tomasz").
        # Miesiąc w nazwie pliku: z nazwy pliku wejściowego (period_mmyyyy),
        # a w razie braku — z dat w danych (zapas).
        period = period_mmyyyy or _period_mmyyyy(sub)
        fname = (_safe_filename(f"{period} dr {_surname_first(disp)}") or "dr lekarz") + ".xlsx"
        # Kwota gotowości + triażu (TeamUp) do wiersza GOTOWOŚĆ w pliku lekarza.
        got_amount = None
        if availability and lek_key in availability:
            got_amount = round(float(availability[lek_key].get("total") or 0), 2)
        # Konsultacje tego lekarza WPINAMY w tabelę „Rozliczenie" (osobne wiersze per
        # priorytet/%, kolumna „{P} Konsultacje"), a konsultowane badania trafiają na
        # „Szczegółowe" pod jego opisami — wszystko w jednym pliku (jedna spójna suma).
        cons_agg = _aggregate_consultations(cons_by_kons.get(lek_key))
        cons_raw = _raw_consulted(lek_key) if lek_key in cons_by_kons else None
        try:
            _write_pair(fname, lambda path: bill_finalize_to_excel(
                grouped, det, path,
                for_doctor=True, rate_resolver=_rate, gotowosc_amount=got_amount,
                consultations=cons_agg, consult_raw=cons_raw,
                row_color=row_color_fn, row_sort_key=row_sort_fn))
            files.append(fname)
            described_keys.add(lek_key)
        except Exception as e:  # noqa: BLE001
            print(f"BŁĄD tworzenia pliku lekarza {disp}: {e}", flush=True)

    # Lekarze, którzy TYLKO konsultowali (bez własnych opisów) — plik z pustą częścią
    # opisów i wpiętymi konsultacjami (ten sam układ tabeli). Szczegółowe = konsultowane
    # badania (df_details), Rozliczenie = wiersze konsultacji.
    empty_cols = ['Priorytet opisu', 'Modalność', 'Procedura', 'Rodzaj procedury rozlicz.',
                  'Procedura rozlicz.', '#', 'Porownawcze_Flag', 'Mnożnik', 'Cena']
    for kk, sub_c in cons_by_kons.items():
        if kk in described_keys:
            continue
        disp = consult_disp[kk]
        if not disp or disp.lower() in ("nan", "none"):
            continue
        raw = _raw_consulted(kk)
        period = period_mmyyyy or _period_mmyyyy(raw)
        fname = (_safe_filename(f"{period} dr {_surname_first(disp)}") or "dr lekarz") + ".xlsx"
        empty_grouped = pd.DataFrame(columns=empty_cols)
        try:
            _write_pair(fname, lambda path, _sc=sub_c: bill_finalize_to_excel(
                empty_grouped, raw, path,
                for_doctor=True, consultations=_aggregate_consultations(_sc),
                row_color=row_color_fn, row_sort_key=row_sort_fn))
            files.append(fname)
        except Exception as e:  # noqa: BLE001
            print(f"BŁĄD tworzenia pliku konsultanta {disp}: {e}", flush=True)

    return {"files": sorted(files), "count": len(files)}
