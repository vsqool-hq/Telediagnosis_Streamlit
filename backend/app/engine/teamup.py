"""
Integracja z TeamUp — godziny gotowości i triażu lekarzy.

Źródła:
  * TeamUp API (https://apidocs.teamup.com): JEDEN kalendarz (grafik dyżurów).
    Lekarz jest w TYTULE wydarzenia (np. „Piłat Krzysztof", czasem z dopiskiem
    typu „RTG"). TRIAŻ to wydarzenie z dopiskiem „TRIAGE"/„TRIAŻ" w tytule
    (np. „TRIAGE Jan Kowalski") — nie osobny kalendarz. Gdy triaż nakłada się z
    dyżurem tej samej osoby, w godzinach wspólnych płacimy tylko triaż.
    Wydarzenia całodniowe i wpisy URLOP/L4 pomijamy. Klucz API: zmienna
    środowiskowa TEAMUP_API_KEY (sekret Fly) albo plik /data/teamup.json.
  * Stawki: plik ZOBOWIĄZAŃ (source.xlsx aktywnego cennika lekarzy) — każda
    zakładka to lekarz, na dole wiersze GOTOWOŚĆ/GODZINA TRIAŻ w wariantach
    (powszedni/weekend/święta × 8:00-21:00/21:00-8:00).

Klasyfikacja godzin: każdą godzinę (ułamkowo, minutowo) przypisujemy do
wariantu wg DATY tej godziny: święto (polskie ustawowe) > weekend > powszedni;
pasmo 8:00–21:00 = dzień, reszta = noc. Triaż ma tylko dzień/noc.
"""

import os
import re
import json
import datetime as dt
import urllib.request
import urllib.parse

API_BASE = "https://api.teamup.com"
# Domyślne kalendarze (z linków klienta) — można nadpisać w /data/teamup.json.
DEFAULT_CAL_GOTOWOSC = "ks4dfo7jwkbaqg2he8"
DEFAULT_CAL_TRIAZ = "ks9cti68r1xnjj5m9m"

# Tytuły wydarzeń pomijanych (nieobecności) — dopasowanie po fragmencie, bez liter wielkości.
SKIP_TOKENS = ("urlop", "l4", "zwolnien", "chorob", "nieobecn")

# Warianty rozliczeniowe: (rodzaj, pasmo, typ dnia) → etykieta do raportów.
VARIANTS = {
    ("G", "D", "POW"): "Gotowość 8:00–21:00",
    ("G", "N", "POW"): "Gotowość 21:00–8:00",
    ("G", "D", "WKD"): "Gotowość weekend 8:00–21:00",
    ("G", "N", "WKD"): "Gotowość weekend 21:00–8:00",
    ("G", "D", "SW"): "Gotowość święta 8:00–21:00",
    ("G", "N", "SW"): "Gotowość święta 21:00–8:00",
    ("T", "D", "*"): "Godzina triaż 8:00–21:00",
    ("T", "N", "*"): "Godzina triaż 21:00–8:00",
}


def _config_path() -> str:
    from app.storage import DATA_DIR
    return os.path.join(DATA_DIR, "teamup.json")


def _env_key() -> str:
    """Klucz z ENV — odporny na wielkość liter nazwy (sekret na Fly bywa
    'teamup_api_key' zamiast 'TEAMUP_API_KEY', a zmienne środowiskowe są
    case-sensitive)."""
    v = os.environ.get("TEAMUP_API_KEY")
    if v:
        return v.strip()
    for k, val in os.environ.items():
        if k.lower() == "teamup_api_key" and val:
            return val.strip()
    return ""


def load_config() -> dict:
    cfg = {}
    try:
        with open(_config_path(), "r", encoding="utf-8") as f:
            cfg = json.load(f) or {}
    except (OSError, ValueError):
        cfg = {}
    env_key = _env_key()
    file_key = (cfg.get("api_key") or "").strip()
    return {
        "api_key": env_key or file_key,
        "cal_gotowosc": cfg.get("cal_gotowosc") or DEFAULT_CAL_GOTOWOSC,
        "cal_triaz": cfg.get("cal_triaz") or DEFAULT_CAL_TRIAZ,
        "key_from_env": bool(env_key),
        "key_source": "env" if env_key else ("file" if file_key else "none"),
        # Nazwy zmiennych środowiskowych zawierające „teamup" (BEZ wartości) —
        # do diagnostyki „mam sekret, a apka go nie widzi".
        "env_names": sorted(k for k in os.environ if "teamup" in k.lower()),
    }


def save_config(api_key: str | None = None, cal_gotowosc: str | None = None,
                cal_triaz: str | None = None) -> dict:
    try:
        with open(_config_path(), "r", encoding="utf-8") as f:
            cfg = json.load(f) or {}
    except (OSError, ValueError):
        cfg = {}
    if api_key is not None:
        cfg["api_key"] = api_key.strip()
    if cal_gotowosc is not None:
        cfg["cal_gotowosc"] = cal_gotowosc.strip()
    if cal_triaz is not None:
        cfg["cal_triaz"] = cal_triaz.strip()
    os.makedirs(os.path.dirname(_config_path()), exist_ok=True)
    with open(_config_path(), "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False)
    return load_config()


# --- Święta polskie (ustawowe) ---------------------------------------------

def _easter(year: int) -> dt.date:
    """Niedziela wielkanocna (algorytm Meeusa)."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month, day = divmod(h + l - 7 * m + 114, 31)
    return dt.date(year, month, day + 1)


def polish_holidays(year: int) -> set:
    e = _easter(year)
    days = {
        dt.date(year, 1, 1), dt.date(year, 1, 6),
        e, e + dt.timedelta(days=1),                      # Wielkanoc + poniedziałek
        dt.date(year, 5, 1), dt.date(year, 5, 3),
        e + dt.timedelta(days=49),                        # Zielone Świątki
        e + dt.timedelta(days=60),                        # Boże Ciało
        dt.date(year, 8, 15), dt.date(year, 11, 1), dt.date(year, 11, 11),
        dt.date(year, 12, 25), dt.date(year, 12, 26),
    }
    if year >= 2025:
        days.add(dt.date(year, 12, 24))                   # Wigilia — ustawowe od 2025
    return days


def _day_type(d: dt.date, holidays: set) -> str:
    if d in holidays:
        return "SW"
    if d.weekday() >= 5:
        return "WKD"
    return "POW"


# --- TeamUp API --------------------------------------------------------------

def fetch_events(cal_key: str, start: dt.date, end: dt.date, api_key: str) -> list:
    """Wydarzenia kalendarza w przedziale dat (włącznie). Rzuca RuntimeError z
    czytelnym komunikatem przy problemach z API/kluczem."""
    q = urllib.parse.urlencode({"startDate": start.isoformat(), "endDate": end.isoformat()})
    url = f"{API_BASE}/{cal_key}/events?{q}"
    req = urllib.request.Request(url, headers={"Teamup-Token": api_key})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.load(resp)
    except urllib.error.HTTPError as e:  # noqa: PERF203
        try:
            detail = json.load(e).get("error", {}).get("message", "")
        except Exception:  # noqa: BLE001
            detail = ""
        raise RuntimeError(f"TeamUp HTTP {e.code}: {detail or e.reason}")
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"TeamUp — błąd połączenia: {e}")
    return data.get("events", [])


def _skip_title(title: str) -> bool:
    t = str(title or "").lower()
    return any(tok in t for tok in SKIP_TOKENS)


# Triaż i zwykły dyżur są w TYM SAMYM kalendarzu — triaż poznajemy po dopisku
# „TRIAGE"/„TRIAŻ" w tytule (np. „TRIAGE Jan Kowalski").
_TRIAGE_RE = re.compile(r"\bTRIA\w*", re.IGNORECASE)


def _is_triage(title) -> bool:
    return bool(_TRIAGE_RE.search(str(title or "")))


def _strip_triage(title) -> str:
    """Usuwa dopisek TRIAGE z tytułu → zostaje sam lekarz
    („TRIAGE Jan Kowalski" → „Jan Kowalski"). Sprząta też puste nawiasy/interpunkcję."""
    s = _TRIAGE_RE.sub(" ", str(title or ""))
    s = re.sub(r"\(\s*\)|\[\s*\]", " ", s)          # puste nawiasy po usunięciu dopisku
    s = re.sub(r"\s+", " ", s).strip(" -–—:;,/()[]")
    return s


def _first_str(v) -> str:
    """Wyciąga tekst z wartości pola własnego TeamUp (str / lista / słownik)."""
    if isinstance(v, str):
        return v
    if isinstance(v, (list, tuple)):
        return _first_str(v[0]) if v else ""
    if isinstance(v, dict):
        for x in v.values():
            s = _first_str(x)
            if s:
                return s
        return ""
    return "" if v is None else str(v)


# Pole z typem dnia: „tryb" (u klienta „tryb_dyz_uru2"), ewentualnie „weekend"/
# „święto"/„wolne". UWAGA: NIE dopasowujemy „dzień" — bo koliduje z polem-flagą
# „dyżur dzienny" (dyz_ur_dzienny), które nie mówi nic o weekendzie.
_DAYTYPE_KEY_RE = re.compile(r"tryb|weekend|świ|swi|wolne", re.IGNORECASE)


def _tryb_dyzuru(ev) -> str | None:
    """
    Typ dnia z POLA WŁASNEGO o nazwie „tryb…"/„weekend"/„święto"/„wolne":
      'W' = weekend (tag „W"), 'S' = święto (tag „Ś"/„S"), '' = takie pole jest,
      ale bez W/Ś (np. puste albo „dodatkowy" → powszedni). None gdy TAKIEGO pola
      w ogóle nie ma → klasyfikujemy z daty (fallback).
    Wartości W/Ś mają PIERWSZEŃSTWO — gdyby pasowało kilka pól, wygrywa to z W/Ś.
    NIE skanujemy przypadkowych wartości i NIE zwracamy 'T' (triaż jest z tytułu).
    """
    custom = ev.get("custom") or {}
    if not isinstance(custom, dict):
        return None
    found = False
    for k, v in custom.items():
        if not _DAYTYPE_KEY_RE.search(str(k).replace("_", " ")):
            continue
        found = True
        dt_ = _classify_daytype_value(_first_str(v))
        if dt_:
            return dt_
    return "" if found else None            # pole trybu bez W/Ś → powszedni; brak → z daty


# Mapowanie polskich znaków na bazowe — odporność na to, JAK TeamUp zwróci „Ś"
# (może przyjść jako „Ś", „ś", „S" albo „święto"/„swieto"; nazwy pól i tak są
# ze slugami „_", ale WARTOŚCI czytamy jako UTF-8).
_PL_STRIP = str.maketrans("śżęąółćńźŚŻĘĄÓŁĆŃŹ", "szeaolcnzSZEAOLCNZ")


def _classify_daytype_value(raw) -> str:
    """Wartość pola trybu → 'W' (weekend) / 'S' (święto) / '' (inne, np. „dodatkowy")."""
    v = str(raw or "").strip()
    if not v:
        return ""
    low = v.lower()
    if low[0] == "w" or "weekend" in low:
        return "W"
    strp = low.translate(_PL_STRIP)              # ś→s, święto→swieto itd.
    if v in ("Ś", "ś", "S", "s") or strp.startswith("swi") or "swiet" in strp:
        return "S"
    return ""


def _slices(start: dt.datetime, end: dt.datetime):
    """Tnie przedział na kawałki w obrębie pełnych godzin zegarowych.
    Zwraca (data, godzina, ułamek_godziny)."""
    cur = start
    while cur < end:
        nxt = (cur.replace(minute=0, second=0, microsecond=0) + dt.timedelta(hours=1))
        nb = min(end, nxt)
        yield cur.date(), cur.hour, (nb - cur).total_seconds() / 3600.0
        cur = nb


def _merge_intervals(intervals: list) -> list:
    """Sortuje i scala nakładające/stykające się przedziały (start, koniec)."""
    if not intervals:
        return []
    ivs = sorted(intervals)
    merged = [ivs[0]]
    for s, e in ivs[1:]:
        ls, le = merged[-1]
        if s <= le:
            merged[-1] = (ls, max(le, e))
        else:
            merged.append((s, e))
    return merged


def _subtract_covered(s: dt.datetime, e: dt.datetime, covering: list) -> list:
    """Zwraca fragmenty [s,e) NIEpokryte żadnym przedziałem z `covering`
    (posortowana, scalona lista (start,koniec))."""
    out = []
    cur = s
    for cs, ce in covering:
        if ce <= cur or cs >= e:
            continue
        if cs > cur:
            out.append((cur, min(cs, e)))
        cur = max(cur, ce)
        if cur >= e:
            break
    if cur < e:
        out.append((cur, e))
    return out


def triaz_intervals_by_title(events: list) -> dict:
    """Przedziały czasu triażu per lekarz (klucz: doctor_key tytułu), scalone.
    Używane do wykluczenia pokrywającego się dyżuru — gdy triaż biegnie w tych
    samych godzinach co dyżur tej samej osoby, płacimy tylko za triaż."""
    from app.engine.cennik_lekarzy_convert import doctor_key
    raw: dict = {}
    for ev in events:
        if ev.get("all_day"):
            continue
        try:
            s = dt.datetime.fromisoformat(ev["start_dt"])
            e = dt.datetime.fromisoformat(ev["end_dt"])
        except (KeyError, ValueError):
            continue
        if e.hour == 23 and e.minute == 59:
            e = e + dt.timedelta(minutes=1)
        if e <= s:
            continue
        key = doctor_key(ev.get("title") or "")
        if not key:
            continue
        raw.setdefault(key, []).append((s, e))
    return {t: _merge_intervals(v) for t, v in raw.items()}


def hours_by_variant(events: list, kind: str, month_start: dt.date, month_end: dt.date,
                     holidays: set, exclude_by_title: dict | None = None) -> dict:
    """Sumuje godziny wydarzeń per (tytuł, wariant). kind: 'G' (gotowość) / 'T' (triaż).
    Godziny spoza miesiąca (wydarzenia na styku) są przycinane do miesiąca.

    exclude_by_title (tylko kind='G'): {tytuł: [(start,koniec), ...]} — godziny
    dyżuru pokrywające się z triażem TEJ SAMEJ osoby są wycinane przed liczeniem,
    bo jeśli triaż jest prowadzony podczas dyżuru RTG, płacimy tylko za triaż
    (część dyżuru bywa pokryta tylko częściowo — stąd odejmowanie przedziałów,
    nie całych zdarzeń)."""
    out: dict = {}
    for ev in events:
        if ev.get("all_day") or _skip_title(ev.get("title")):
            continue
        try:
            s = dt.datetime.fromisoformat(ev["start_dt"])
            e = dt.datetime.fromisoformat(ev["end_dt"])
        except (KeyError, ValueError):
            continue
        # Koniec 23:59 traktujemy jak 24:00 (tak jak kalkulator TeamUp) — inaczej
        # każdy taki dyżur gubi 1 minutę, a takich rekordów są dziesiątki.
        if e.hour == 23 and e.minute == 59:
            e = e + dt.timedelta(minutes=1)
        if e <= s:
            continue
        title = str(ev.get("title") or "").strip()
        if not title:
            continue
        # Weekend/święto bierzemy z ręcznego pola „Tryb dyżuru" (jak kalkulator
        # TeamUp) — całe zdarzenie ma jeden typ dnia. Fallback na datę, gdy pola brak.
        tryb = _tryb_dyzuru(ev) if kind == "G" else None

        segments = [(s, e)]
        if kind == "G" and exclude_by_title:
            from app.engine.cennik_lekarzy_convert import doctor_key
            covering = exclude_by_title.get(doctor_key(title))
            if covering:
                segments = _subtract_covered(s, e, covering)

        for seg_s, seg_e in segments:
            for d, hour, frac in _slices(seg_s, seg_e):
                if not (month_start <= d <= month_end):
                    continue
                band = "D" if 8 <= hour < 21 else "N"
                if kind != "G":
                    day_t = "*"
                elif tryb is None:
                    day_t = _day_type(d, holidays)        # brak pola Tryb → z daty
                elif tryb == "S":
                    day_t = "SW"
                elif tryb == "W":
                    day_t = "WKD"
                else:                                     # "" (powszedni) lub inne
                    day_t = "POW"
                key = (kind, band, day_t)
                per = out.setdefault(title, {})
                per[key] = per.get(key, 0.0) + frac
    return out


# --- Stawki z pliku ZOBOWIĄZAŃ ----------------------------------------------

def parse_availability_rates(workbook_path: str, period: str | None = None) -> dict:
    """
    Czyta stawki gotowości/triażu z KAŻDEJ zakładki (zakładka = lekarz).
    Zwraca { doctor_key(nazwa zakładki): {"name": zakładka, "rates": {wariant: stawka}} }.

    Stawki bierzemy z ANEKSU OBOWIĄZUJĄCEGO w `period` ('YYYY-MM') — dokładnie tak
    samo jak ceny badań: kolumnę-etykietę i stawki wyznacza _find_block_for_period
    (blok, którego kolumna miesiąca = period; skrajnie prawy bloku bywa przyszłym
    aneksem). Bez period — skrajnie prawy blok (jak dotąd). Wiersze GOTOWOŚĆ/TRIAŻ są
    powielone pod kategoriami badań w tym samym bloku, więc czytamy je z tej samej
    pary kolumn. Wariant: godzina startu pasma (8→dzień, inaczej→noc) + typ dnia.
    """
    import re
    from openpyxl import load_workbook
    from app.engine.cennik_lekarzy_convert import (
        doctor_key, _find_block_for_period, _clean, _try_number,
    )

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    out = {}
    for sheet in wb.sheetnames:
        if _clean(sheet).upper().startswith("ZBIORCZO"):
            continue
        grid = list(wb[sheet].iter_rows(values_only=True))
        if not grid:
            continue
        ncols = max((len(r) for r in grid), default=0)
        label_col, price_col = _find_block_for_period(grid, ncols, period)
        if label_col is None:
            continue

        rates = {}
        for r in grid:
            label = _clean(r[label_col] if label_col < len(r) else None).upper()
            if not label or ("GOTOW" not in label and "TRIA" not in label):
                continue
            m = re.search(r"(\d{1,2})[:.]?\d{0,2}\s*[-–]", label)
            if not m:
                continue
            band = "D" if int(m.group(1)) == 8 else "N"
            rate, _rep, _orig = _try_number(r[price_col] if price_col < len(r) else None)
            if rate is None:
                rate = 0.0
            if "TRIA" in label:
                rates[("T", band, "*")] = rate
            else:
                day_t = "SW" if ("ŚWI" in label or "SWI" in label) else \
                        ("WKD" if "WEEKEND" in label else "POW")
                rates[("G", band, day_t)] = rate
        if rates:
            out[doctor_key(sheet)] = {"name": _clean(sheet), "rates": rates}
    wb.close()
    return out


# --- Dopasowanie lekarza z tytułu wydarzenia ---------------------------------

def match_doctor(title: str, known: dict) -> str | None:
    """
    Dopasowuje tytuł wydarzenia do klucza lekarza (known: {doctor_key: ...}).
    1) dokładny klucz; 2) klucz lekarza zawarty w tytule (tytuł ma dopiski, np.
    „Dreżewski Karol RTG"); 3) tytuł zawarty w kluczu (samo nazwisko) — tylko
    gdy pasuje DOKŁADNIE jeden lekarz.
    """
    from app.engine.cennik_lekarzy_convert import doctor_key
    tk = doctor_key(title)
    if not tk:
        return None
    if tk in known:
        return tk
    t_tokens = set(tk.split(" "))
    contained = [k for k in known if set(k.split(" ")) <= t_tokens]
    if len(contained) == 1:
        return contained[0]
    if len(contained) > 1:   # np. dwóch lekarzy o tym samym nazwisku — wybierz dłuższe dopasowanie
        contained.sort(key=lambda k: -len(set(k.split(" ")) & t_tokens))
        if len(set(contained[0].split(" ")) & t_tokens) > len(set(contained[1].split(" ")) & t_tokens):
            return contained[0]
        return None
    reverse = [k for k in known if t_tokens <= set(k.split(" "))]
    return reverse[0] if len(reverse) == 1 else None


# --- Główne liczenie ---------------------------------------------------------

def compute_availability(period: str, excluded_keys=None) -> dict:
    """
    Gotowość + triaż za miesiąc rozliczenia „YYYY-MM": godziny z TeamUp × stawki
    z pliku ZOBOWIĄZAŃ. Zwraca:
      { doctors: {lek_key: {name, items:[{label,hours,rate,amount}], total}},
        sum_total, sum_gotowosc, sum_triaz, unmatched: [tytuły], period }
    Rzuca RuntimeError z czytelnym powodem (brak klucza / pliku zobowiązań / API).
    """
    cfg = load_config()
    if not cfg["api_key"]:
        raise RuntimeError("Brak klucza API TeamUp (ustaw sekret TEAMUP_API_KEY lub wpisz w Ustawieniach).")
    from app.engine.commitments import active_commitments_workbook
    wb_path, _disp = active_commitments_workbook()
    if not wb_path:
        raise RuntimeError("Brak pliku zobowiązań (wgraj cennik lekarzy jako .xlsx przez konwerter).")

    y, m = int(period[:4]), int(period[5:7])
    month_start = dt.date(y, m, 1)
    month_end = (dt.date(y + 1, 1, 1) if m == 12 else dt.date(y, m + 1, 1)) - dt.timedelta(days=1)
    holidays = polish_holidays(y) | polish_holidays(y + 1)

    rates_by_doc = parse_availability_rates(wb_path, period)

    # Pobieramy z 1-dniowym marginesem (dyżur nocny zaczęty ostatniego dnia
    # poprzedniego miesiąca) — i tak przycinamy godziny do miesiąca.
    fetch_from = month_start - dt.timedelta(days=1)
    fetch_to = month_end + dt.timedelta(days=1)

    # JEDEN kalendarz (grafik dyżurów). Triaż vs zwykły dyżur rozróżniamy po dopisku
    # „TRIAGE"/„TRIAŻ" w tytule; z tytułu triażu zdejmujemy dopisek → zostaje lekarz.
    g_events: list = []
    t_events: list = []
    if cfg["cal_gotowosc"]:
        for ev in fetch_events(cfg["cal_gotowosc"], fetch_from, fetch_to, cfg["api_key"]):
            if _is_triage(ev.get("title")):
                t_events.append({**ev, "title": _strip_triage(ev.get("title"))})
            else:
                g_events.append(ev)

    # Gdy triaż biegnie w tych samych godzinach co dyżur TEJ SAMEJ osoby — płacimy
    # tylko triaż: pokrywające się przedziały wycinamy z dyżuru.
    triaz_exclude = triaz_intervals_by_title(t_events)

    per_title: dict = {}
    if g_events:
        part = hours_by_variant(g_events, "G", month_start, month_end, holidays,
                                 exclude_by_title=triaz_exclude)
        for title, variants in part.items():
            agg = per_title.setdefault(title, {})
            for k, v in variants.items():
                agg[k] = agg.get(k, 0.0) + v
    if t_events:
        part = hours_by_variant(t_events, "T", month_start, month_end, holidays)
        for title, variants in part.items():
            agg = per_title.setdefault(title, {})
            for k, v in variants.items():
                agg[k] = agg.get(k, 0.0) + v

    excl = set(excluded_keys or [])
    doctors: dict = {}
    unmatched: list = []
    unmatched_hours = 0.0
    for title, variants in per_title.items():
        lk = match_doctor(title, rates_by_doc)
        if lk is not None and lk in excl:
            continue                                    # lekarz wyłączony — pomijamy jego gotowość
        if lk is None:
            unmatched.append(title)
            unmatched_hours += sum(variants.values())   # godziny nierozliczone (brak lekarza)
            continue
        doc = doctors.setdefault(lk, {"name": rates_by_doc[lk]["name"], "variants": {}})
        for k, v in variants.items():
            doc["variants"][k] = doc["variants"].get(k, 0.0) + v

    sum_g = sum_t = 0.0
    hours_g = hours_t = 0.0        # łączne GODZINY (niezależnie od stawki)
    unbilled_hours = 0.0          # godziny ze stawką ≤ 0 (nierozliczone kwotowo)
    unbilled = []                 # [{name, label, hours}]
    for lk, doc in doctors.items():
        rates = rates_by_doc[lk]["rates"]
        items, total = [], 0.0
        for k, hours in sorted(doc.pop("variants").items()):
            # Zaokrąglenie do PEŁNYCH godzin, 30 min w górę (29 min → 0, 30 min → 1).
            # Per wariant, bo każdy ma inną stawkę — to jednostka rozliczeniowa.
            hrs = int(hours + 0.5)   # hours ≥ 0, więc int(x+0.5) = zaokrąglenie w górę od połowy
            if hrs == 0:
                continue             # poniżej 30 min w tym wariancie → 0, pomijamy
            rate = float(rates.get(k, 0.0))
            amount = round(hrs * rate, 2)
            total += amount
            if k[0] == "G":
                sum_g += amount; hours_g += hrs
            else:
                sum_t += amount; hours_t += hrs
            if rate <= 0:
                unbilled_hours += hrs
                unbilled.append({"name": doc["name"], "label": VARIANTS.get(k, str(k)), "hours": hrs})
            items.append({"label": VARIANTS.get(k, str(k)), "hours": hrs,
                          "rate": rate, "amount": amount, "no_rate": rate <= 0})
        doc["items"] = items
        doc["total"] = round(total, 2)
    return {
        "period": period,
        "doctors": doctors,
        "sum_total": round(sum_g + sum_t, 2),
        "sum_gotowosc": round(sum_g, 2),
        "sum_triaz": round(sum_t, 2),
        "hours_gotowosc": round(hours_g, 2),
        "hours_triaz": round(hours_t, 2),
        "unbilled_hours": round(unbilled_hours, 2),
        "unbilled": sorted(unbilled, key=lambda x: -x["hours"])[:80],
        "unmatched_hours": round(unmatched_hours, 2),
        "unmatched": sorted(set(unmatched)),
    }
