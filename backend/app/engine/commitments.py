"""
Plik zobowiązań lekarzy — uzupełnianie ILOŚCI OKOLIC po rozliczeniu.

„Plik zobowiązań" to szeroki skoroszyt (zakładka = lekarz), wgrywany przez
konwerter cennika lekarzy i przechowywany jako source.xlsx przy aktywnej wersji
cennika lekarzy. Po policzeniu rozliczenia danego miesiąca wpisujemy do niego
liczbę OKOLIC per lekarz i kategoria, w kolumnie tego miesiąca — KUMULATYWNIE
(kolejne liczone miesiące dopisują się do tego samego pliku; nie ruszamy nic poza
komórkami ilości — formuły SUMA/ŚREDNIA przeliczą się same w Excelu).

Mapowanie:
  * lekarz → zakładka: doctor_key() (niewrażliwe na kolejność imię/nazwisko),
  * miesiąc → kolumna: data w WIERSZU 2. Gdy ten sam miesiąc występuje w kilku
    blokach (aneksy obok siebie), bierzemy NAJNOWSZY = NAJBARDZIEJ NA PRAWO.
  * kategoria → wiersz: etykieta w kolumnie A (wiersze wspólne dla bloków aneksów).

Miesiąc ROZBITY aneksem w połowie:
  gdy w połowie miesiąca podpisano aneks ze zmianą stawek, miesiąc w pliku jest
  podzielony na DWIE kolumny z zakresem dni w nagłówku (np. „01-17.05.2026" i
  „18-31.05.2026"). Wtedy nie wpisujemy jednej sumy — rozdzielamy okolice po
  DACIE BADANIA: badania z dni ≤ granicy idą do pierwszej kolumny, pozostałe do
  drugiej (dane dzienne z build_doctor_billing → category_okolice_daily).
"""

import os
import re
import shutil
import datetime
from collections import defaultdict

from openpyxl import load_workbook

from app import db
from app.storage import version_dir
from app.engine.cennik_lekarzy_convert import doctor_key


def _ncat(s) -> str:
    return re.sub(r"\s+", " ", str(s if s is not None else "")).strip().upper()


# Wiersze OKOLIC (kategorie badań: RTG/TK/MR/MMG …) — te i tylko te czyścimy w
# kolumnie miesiąca przed wpisaniem bieżących ilości, żeby nie zostały NIEAKTUALNE
# wartości z wcześniejszego przeliczenia (inna wersja słownika = inny podział
# kategorii → w kolumnie miesiąca zostawały „osierocone" ilości). NIE ruszamy
# nagłówków sekcji, SUMA/ILOŚĆ OKOLIC, GOTOWOŚĆ/TRIAŻ, DATA/FAKTURA/TERMIN.
def _is_okolice_row(label) -> bool:
    u = _ncat(label)
    if not u:
        return False
    if u.split()[0] not in ("RTG", "TK", "MR", "MMG"):
        return False
    if u in ("RTG", "TK", "MR", "MMG"):          # sam nagłówek sekcji
        return False
    if any(x in u for x in ("SUMA", "OKOLIC", "GOTOW", "TRIA", "FAKTUR", "ANEKS", "DATA", "TERMIN")):
        return False
    return True


# Nagłówek kolumny rozbitego miesiąca: „DD-DD.MM.YYYY" (np. „01-17.05.2026").
_RANGE_RE = re.compile(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*\.\s*(\d{4})\s*$")


def active_commitments_workbook():
    """Zwraca (ścieżka_source.xlsx, nazwa_wyświetlana) aktywnego pliku zobowiązań,
    albo (None, None) jeśli aktywny cennik lekarzy nie był wgrany jako .xlsx."""
    v = db.get_active_version("cennik_lekarzy")
    if not v:
        return None, None
    vdir = version_dir("cennik_lekarzy", v["id"])
    path = os.path.join(vdir, "source.xlsx")
    if not os.path.isfile(path):
        return None, None
    name = "ZOBOWIĄZANIA LEKARZY.xlsx"
    nf = os.path.join(vdir, "source_name.txt")
    if os.path.isfile(nf):
        try:
            name = open(nf, encoding="utf-8").read().strip() or name
        except OSError:
            pass
    return path, name


def _month_columns(ws, y: int, m: int):
    """Zwraca kolumny miesiąca z wiersza 2:
      whole  — [(col)] kolumny ze zwykłą datą tego miesiąca (zwykle 1, czasem
               kilka przy powtórce w blokach aneksów),
      splits — [(col, dzień_od, dzień_do)] kolumny rozbitego miesiąca
               (nagłówek „DD-DD.MM.YYYY"), posortowane wg dnia początkowego.
    """
    whole, splits = [], []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if isinstance(v, datetime.datetime):
            if v.year == y and v.month == m:
                whole.append(c)
        elif isinstance(v, str):
            mt = _RANGE_RE.match(v)
            if mt:
                d_from, d_to, mm, yy = (int(mt.group(i)) for i in range(1, 5))
                if yy == y and mm == m:
                    splits.append((c, d_from, d_to))
    splits.sort(key=lambda t: t[1])
    return whole, splits


def fill_workbook(wb_path: str, period_ym: str, okolice_rows: list,
                  daily_rows: list = None) -> dict:
    """
    Uzupełnia w skoroszycie (w miejscu) liczbę okolic dla danego miesiąca.
      period_ym    — „YYYY-MM" (miesiąc rozliczenia),
      okolice_rows — [{lekarz, kategoria, okolice}] — suma okolic (zwykły miesiąc),
      daily_rows   — [{lekarz, kategoria, data:"YYYY-MM-DD", okolice}] — rozbicie
                     dzienne; używane TYLKO dla miesięcy rozbitych aneksem.
    Nadpisuje TYLKO komórki (kategoria × kolumna miesiąca); reszty nie rusza.
    """
    y, m = (int(x) for x in period_ym.split("-")[:2])

    by_doc = defaultdict(dict)   # doctor_key -> {NCAT: okolice}
    disp_by_key = {}
    for r in okolice_rows:
        k = doctor_key(r["lekarz"])
        if not k:
            continue
        by_doc[k][_ncat(r["kategoria"])] = r["okolice"]
        disp_by_key.setdefault(k, r["lekarz"])

    # rozbicie dzienne: doctor_key -> {NCAT: {dzień_miesiąca: suma_okolic}}
    daily = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in (daily_rows or []):
        k = doctor_key(r["lekarz"])
        if not k:
            continue
        try:
            d = datetime.date.fromisoformat(str(r["data"])[:10])
        except (ValueError, TypeError):
            continue
        if d.year == y and d.month == m:
            daily[k][_ncat(r["kategoria"])][d.day] += float(r["okolice"] or 0)

    report = {"period": period_ym, "doctors_written": 0, "cells_written": 0,
              "doctors_no_sheet": [], "no_month_column": [], "categories_not_found": [],
              "split_doctors": [], "split_unallocated": []}

    wb = load_workbook(wb_path)  # data_only=False — zachowujemy formuły

    sheet_by_key = {}
    for sn in wb.sheetnames:
        if sn.strip().upper().startswith("ZBIORCZO"):
            continue
        sheet_by_key.setdefault(doctor_key(sn), sn)

    for k, cats in by_doc.items():
        sn = sheet_by_key.get(k)
        if not sn:
            report["doctors_no_sheet"].append(disp_by_key.get(k, k))
            continue
        ws = wb[sn]

        whole_cols, split_cols = _month_columns(ws, y, m)
        if not whole_cols and not split_cols:
            report["no_month_column"].append(sn)
            continue

        # kategoria → wiersz (etykieta w kolumnie A; wiersze wspólne dla bloków)
        cat_row = {}
        okolice_row_idx = []
        for rr in range(1, ws.max_row + 1):
            a = ws.cell(row=rr, column=1).value
            if a is not None:
                cat_row.setdefault(_ncat(a), rr)
                if _is_okolice_row(a):
                    okolice_row_idx.append(rr)

        wrote = 0
        if split_cols:
            # Miesiąc rozbity aneksem — rozdziel okolice po dniu badania na zakresy.
            doc_daily = daily.get(k, {})
            for ncat, total in cats.items():
                row = cat_row.get(ncat)
                if not row:
                    report["categories_not_found"].append({"lekarz": disp_by_key.get(k, k), "kategoria": ncat})
                    continue
                day_map = doc_daily.get(ncat, {})
                if not day_map:
                    # Brak danych dziennych dla tej kategorii — NIE nadpisujemy
                    # komórek zerami; zgłaszamy do ręcznego sprawdzenia.
                    report["split_unallocated"].append({
                        "lekarz": disp_by_key.get(k, k), "kategoria": ncat,
                        "suma_miesiac": float(total or 0), "rozdzielono": 0.0})
                    continue
                allocated = 0.0
                for (col, d_from, d_to) in split_cols:
                    s = sum(ok for day, ok in day_map.items() if d_from <= day <= d_to)
                    ws.cell(row=row, column=col).value = int(round(s))
                    allocated += s
                    wrote += 1
                # suma z zakresów ≠ suma miesięczna → niepełne dane dzienne (np. NaT)
                if abs(allocated - float(total or 0)) > 0.5:
                    report["split_unallocated"].append({
                        "lekarz": disp_by_key.get(k, k), "kategoria": ncat,
                        "suma_miesiac": float(total or 0), "rozdzielono": allocated})
            report["split_doctors"].append({
                "lekarz": disp_by_key.get(k, k),
                "zakresy": [{"od": d_from, "do": d_to} for (_, d_from, d_to) in split_cols]})
        else:
            # Zwykły miesiąc — jedna suma do NAJNOWSZEJ (skrajnie prawej) kolumny.
            target_col = max(whole_cols)
            # Najpierw WYCZYŚĆ tę kolumnę we wszystkich wierszach okolic — inaczej
            # zostają nieaktualne ilości z wcześniejszego przeliczenia (np. kategoria,
            # której lekarz już w tym miesiącu nie ma; inna wersja słownika = inny
            # podział). Nie ruszamy SUMA/GOTOWOŚĆ/innych miesięcy.
            for rr in okolice_row_idx:
                ws.cell(row=rr, column=target_col).value = None
            for ncat, okolice in cats.items():
                row = cat_row.get(ncat)
                if not row:
                    report["categories_not_found"].append({"lekarz": disp_by_key.get(k, k), "kategoria": ncat})
                    continue
                ws.cell(row=row, column=target_col).value = int(okolice)
                wrote += 1

        if wrote:
            report["doctors_written"] += 1
            report["cells_written"] += wrote

    wb.save(wb_path)
    return report


def fill_and_package(wb_path: str, period_ym: str, okolice_rows: list,
                     package_dir: str, display_name: str, daily_rows: list = None) -> dict:
    """Uzupełnia przechowywany plik (kumulatywnie) i wrzuca jego kopię do paczki."""
    report = fill_workbook(wb_path, period_ym, okolice_rows, daily_rows)
    os.makedirs(package_dir, exist_ok=True)
    base = re.sub(r'[\\/:*?"<>|]+', "", display_name or "ZOBOWIĄZANIA LEKARZY")
    base = re.sub(r"\.(xlsx|xls)$", "", base, flags=re.I).strip() or "ZOBOWIĄZANIA LEKARZY"
    out = os.path.join(package_dir, f"{base} - ilosci.xlsx")
    shutil.copy2(wb_path, out)
    report["package_file"] = os.path.basename(out)
    return report
